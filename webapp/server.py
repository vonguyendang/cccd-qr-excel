import os
import sys
import json
import base64
import uuid
import datetime
import asyncio
import httpx
import re
from io import BytesIO
import numpy as np
import logging
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
try:
    from vietocr_engine import extract_text_from_image
except ImportError as e:
    print(f"Warning: Could not import vietocr_engine. OCR might not work: {e}")

from fastapi import FastAPI, UploadFile, File, Request, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any

import cv2
from pyzbar.pyzbar import decode, ZBarSymbol
import openpyxl
from openpyxl.styles import Font

app = FastAPI(title="CCCD QR API")

# Setup CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global model variables
detector = None

@app.on_event("startup")
async def startup_event():
    cleanup_old_sessions()
    await load_models()

async def load_models():

    global detector
    model_paths = [
        'models/detect.prototxt', 'models/detect.caffemodel',
        'models/sr.prototxt', 'models/sr.caffemodel'
    ]
    base_dir = os.path.dirname(os.path.abspath(__file__))
    abs_paths = [os.path.join(base_dir, p) for p in model_paths]
    if all(os.path.exists(p) for p in abs_paths):
        try:
            detector = cv2.wechat_qrcode_WeChatQRCode(*abs_paths)
            print("Loaded WeChat QRCode detector.")
        except Exception as e:
            print(f"Error loading detector: {e}")
    else:
        print("Model files not found.")

# ---- Room & WebSocket Sync Manager ----
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = {}

    async def connect(self, websocket: WebSocket, room_id: str):
        await websocket.accept()
        if room_id not in self.active_connections:
            self.active_connections[room_id] = []
        self.active_connections[room_id].append(websocket)

    def disconnect(self, websocket: WebSocket, room_id: str):
        if room_id in self.active_connections:
            if websocket in self.active_connections[room_id]:
                self.active_connections[room_id].remove(websocket)
            if not self.active_connections[room_id]:
                del self.active_connections[room_id]

    async def broadcast(self, message: dict, room_id: str):
        if room_id in self.active_connections:
            for connection in self.active_connections[room_id]:
                try:
                    await connection.send_json(message)
                except Exception:
                    pass

manager = ConnectionManager()
rooms: Dict[str, Dict[str, Any]] = {}
import time
import glob

SESSIONS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sessions")
os.makedirs(SESSIONS_DIR, exist_ok=True)

def cleanup_old_sessions():
    try:
        now = time.time()
        count = 0
        for filepath in glob.glob(os.path.join(SESSIONS_DIR, "*.json")):
            if os.path.isfile(filepath):
                # Check file modified time or read JSON
                # 10 days = 10 * 24 * 60 * 60 seconds = 864000
                if now - os.path.getmtime(filepath) > 864000:
                    os.remove(filepath)
                    count += 1
        if count > 0:
            print(f"Cleaned up {count} old session(s) > 10 days.")
    except Exception as e:
        print(f"Error cleaning up old sessions: {e}")


def save_room(room_id: str):
    try:
        if room_id in rooms:
            filepath = os.path.join(SESSIONS_DIR, f"{room_id}.json")
            data_to_save = {
                "items": rooms[room_id]["items"],
                "seen_cccds": list(rooms[room_id]["seen_cccds"]),
                "duplicate_files": rooms[room_id].get("duplicate_files", [])
            }
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving session {room_id}: {e}")

ROOMS_BACKUP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rooms_backup.json")

def cleanup_old_sessions():
    global rooms
    if os.path.exists(ROOMS_BACKUP_FILE):
        try:
            with open(ROOMS_BACKUP_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                for room_id, room_data in data.items():
                    rooms[room_id] = {
                        "items": room_data.get("items", []),
                        "seen_cccds": set(room_data.get("seen_cccds", [])),
                        "duplicate_files": room_data.get("duplicate_files", [])
                    }
            print(f"Loaded {len(rooms)} rooms from backup.")
        except Exception as e:
            print(f"Error loading rooms backup: {e}")

def save_room(room_id):
    try:
        data_to_save = {}
        for room_id, room_data in rooms.items():
            data_to_save[room_id] = {
                "items": room_data["items"],
                "seen_cccds": list(room_data["seen_cccds"]),
                "duplicate_files": room_data.get("duplicate_files", [])
            }
        with open(ROOMS_BACKUP_FILE, "w", encoding="utf-8") as f:
            json.dump(data_to_save, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving rooms backup: {e}")

def get_room(room_id: str):
    if room_id not in rooms:
        rooms[room_id] = {
            "items": [],
            "seen_cccds": set(),
            "duplicate_files": []
        }
    return rooms[room_id]


@app.websocket("/ws/room/{room_id}")
async def websocket_endpoint(websocket: WebSocket, room_id: str):
    await manager.connect(websocket, room_id)
    try:
        while True:
            _ = await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket, room_id)

class ExportItem(BaseModel):
    filename: Optional[str] = None
    qrData: Optional[str] = None
    error: Optional[str] = None
    fromOCR: Optional[bool] = False
    ocrData: Optional[Dict[str, str]] = None
    imageBase64: Optional[str] = None
    isDuplicate: Optional[bool] = False
    duplicateWith: Optional[str] = None

class RoomAddRequest(BaseModel):
    room_id: str
    item: ExportItem

@app.post("/api/room/add")
async def room_add(req: RoomAddRequest):
    room_id = req.room_id
    item = req.item
    room = get_room(room_id)
    
    if "duplicate_files" not in room:
        room["duplicate_files"] = []
        
    if item.imageBase64 and item.filename:
        room_img_dir = os.path.join(SESSIONS_DIR, room_id, "images")
        os.makedirs(room_img_dir, exist_ok=True)
        img_path = os.path.join(room_img_dir, item.filename)
        base64_str = item.imageBase64
        if "," in base64_str:
            base64_str = base64_str.split(",")[1]
        try:
            img_data = base64.b64decode(base64_str)
            with open(img_path, "wb") as f:
                f.write(img_data)
        except Exception as e:
            print(f"Error saving image: {e}")
            
    # Remove imageBase64 from item before broadcasting to save bandwidth
    item.imageBase64 = None
    
    if item.isDuplicate:
        if item.filename:
            room["duplicate_files"].append({
                "duplicate": item.filename,
                "original": item.duplicateWith or ""
            })
        save_room(room_id)
        return {"success": True, "total_count": len(room["items"])}
    
    cccd_num = None
    if item.qrData:
        parts = item.qrData.split('|')
        cccd_num = parts[0] if parts else None
    elif item.fromOCR and item.ocrData:
        cccd_num = item.ocrData.get('CCCD')
        
    if cccd_num:
        if cccd_num in room["seen_cccds"]:
            # If an OCR is already there and we found a QR, we should replace it.
            # But for simplicity, if it's in seen_cccds, we check if we need to upgrade from OCR to QR.
            # In Python, we can find the index and replace.
            existing_idx = next((i for i, v in enumerate(room["items"]) 
                                 if (v.get("qrData") and v["qrData"].startswith(cccd_num)) or 
                                    (v.get("ocrData") and v["ocrData"].get("CCCD") == cccd_num)), None)
            
            if existing_idx is not None:
                existing_item = room["items"][existing_idx]
                if not existing_item.get("qrData") and item.qrData:
                    # Upgrade OCR to QR
                    old_filename = existing_item.get("filename")
                    if old_filename:
                        room["duplicate_files"].append({
                            "duplicate": old_filename,
                            "original": item.filename or ""
                        })
                    room["items"][existing_idx] = item.dict()
                    await manager.broadcast({
                        "type": "update_item",
                        "items": room["items"],
                        "total_count": len(room["items"])
                    }, room_id)
                    save_room(room_id)
                    return {"success": True, "total_count": len(room["items"])}
                else:
                    # Both are OCR or both are QR, check which has more info
                    def count_info(d):
                        if not d: return 0
                        keys = ['Họ tên', 'CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD']
                        return sum(1 for k in keys if d.get(k))
                    
                    old_count = 0
                    new_count = 0
                    
                    # For OCR, compare ocrData
                    if not existing_item.get("qrData") and not item.qrData:
                        old_count = count_info(existing_item.get("ocrData"))
                        new_count = count_info(item.ocrData)
                    
                    if new_count > old_count:
                        old_filename = existing_item.get("filename")
                        if old_filename:
                            room["duplicate_files"].append({
                                "duplicate": old_filename,
                                "original": item.filename or ""
                            })
                        room["items"][existing_idx] = item.dict()
                        await manager.broadcast({
                            "type": "update_item",
                            "items": room["items"],
                            "total_count": len(room["items"])
                        }, room_id)
                        save_room(room_id)
                        return {"success": True, "total_count": len(room["items"])}
                    else:
                        return {"success": False, "error": "Duplicate CCCD"}
        else:
            room["seen_cccds"].add(cccd_num)
            
    item_dict = item.dict()
    room["items"].append(item_dict)
    
    await manager.broadcast({
        "type": "new_item",
        "item": item_dict,
        "total_count": len(room["items"])
    }, room_id)
    save_room(room_id)
    return {"success": True, "total_count": len(room["items"])}

@app.get("/api/room/state/{room_id}")
async def room_state(room_id: str):
    room = get_room(room_id)
    return {"success": True, "items": room["items"], "total_count": len(room["items"])}

class RoomClearRequest(BaseModel):
    room_id: str

@app.post("/api/room/clear")
async def room_clear(req: RoomClearRequest):
    room_id = req.room_id
    if room_id in rooms:
        rooms[room_id] = {"items": [], "seen_cccds": set(), "duplicate_files": []}
        save_room(room_id)
        await manager.broadcast({
            "type": "clear",
            "total_count": 0
        }, room_id)
    return {"success": True}

# ---- End Room Manager ----

class ScanQRRequest(BaseModel):
    imageBase64: str
    filename: Optional[str] = None

class LogCameraRequest(BaseModel):
    message: str

@app.post("/api/log_camera")
async def log_camera(req: LogCameraRequest):
    print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {req.message}", flush=True)
    return {"success": True}

@app.post("/api/scan_qr")
async def scan_qr(req: ScanQRRequest):
    fname = f"'{req.filename}'" if req.filename else "từ Live Camera"
    print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Nhận yêu cầu quét QR ảnh {fname}...", flush=True)
    if not detector:
        print("-> Lỗi: Mô hình WeChat QRCode chưa được tải.", flush=True)
        return {"success": False, "error": "Model not loaded"}
    
    # decode base64
    base64_str = req.imageBase64
    if "," in base64_str:
        base64_str = base64_str.split(",")[1]
    
    try:
        img_data = base64.b64decode(base64_str)
        nparr = np.frombuffer(img_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            print("-> Lỗi: Dữ liệu ảnh không hợp lệ.", flush=True)
            return {"success": False, "error": "Invalid image"}
            
        import tempfile
        from wizard.main import extract_qr_data
        
        fd, temp_path = tempfile.mkstemp(suffix=".jpg")
        os.close(fd)
        
        try:
            with open(temp_path, "wb") as f:
                f.write(base64.b64decode(base64_str))
                
            qr_string, engine, err, _, qr_rotated_img = extract_qr_data(temp_path)
            
            if qr_string:
                rotated_base64 = None
                if qr_rotated_img is not None:
                    _, buffer = cv2.imencode('.jpg', qr_rotated_img)
                    rotated_base64 = "data:image/jpeg;base64," + base64.b64encode(buffer).decode('utf-8')
                    
                print(f"-> Quét thành công QR ({engine}): {qr_string}", flush=True)
                return {"success": True, "data": qr_string, "rotatedBase64": rotated_base64}
            else:
                print(f"-> Lỗi QR: {err}", flush=True)
                return {"success": False, "error": err}
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
                
    except Exception as e:
        print(f"-> Lỗi hệ thống khi quét ảnh: {str(e)}", flush=True)
        return {"success": False, "error": f"Exception: {str(e)}"}

class OCRRequest(BaseModel):
    imageBase64: str

@app.post("/api/ocr")
async def extract_ocr(req: OCRRequest):
    print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Nhận yêu cầu trích xuất OCR...", flush=True)
    base64_str = req.imageBase64
    if "," in base64_str:
        base64_str = base64_str.split(",")[1]
        
    try:
        img_data = base64.b64decode(base64_str)
        nparr = np.frombuffer(img_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            return {"success": False, "error": "Invalid image"}
            
        from wizard.main import extract_ocr_data
        ocr_data, ocr_note, rotated_img = extract_ocr_data(img)
        
        if ocr_data.get('CCCD') or ocr_data.get('Họ tên'):
            rotated_base64 = None
            if rotated_img is not None:
                _, buffer = cv2.imencode('.jpg', rotated_img)
                rotated_base64 = "data:image/jpeg;base64," + base64.b64encode(buffer).decode('utf-8')
                
            print(f"-> OCR Thành công: CCCD={ocr_data.get('CCCD')}, Tên={ocr_data.get('Họ tên')} - {ocr_note}")
            # Dù không có CCCD nhưng có tên, hoặc có CCCD, đều trả về success True (hoặc có rotatedBase64 thì ưu tiên trả về)
            return {"success": True, "ocrData": ocr_data, "rotatedBase64": rotated_base64}
        else:
            # Nếu OCR trả về rotated_img nhưng không có tên/CCCD (rất hiếm), ta vẫn có thể muốn giữ lại ảnh đã xoay
            rotated_base64 = None
            if rotated_img is not None:
                _, buffer = cv2.imencode('.jpg', rotated_img)
                rotated_base64 = "data:image/jpeg;base64," + base64.b64encode(buffer).decode('utf-8')
                return {"success": True, "ocrData": ocr_data, "rotatedBase64": rotated_base64, "text": ocr_note}
                
            return {"success": False, "error": ocr_note}
            
    except Exception as e:
        print(f"-> Lỗi OCR: {e}")
        return {"success": False, "error": str(e)}

def _find_cccd_pattern(text: str) -> str:
    match = re.search(r'\b\d{12}\b', text)
    return match.group(0) if match else ""

def format_date(date_str):
    if not date_str or len(date_str) != 8:
        return date_str
    return f"{date_str[0:2]}/{date_str[2:4]}/{date_str[4:8]}"

def get_place_of_issue(qr_data):
    if not qr_data:
        return ""
    fields = qr_data.split('|')
    if len(fields) == 7:
        return "CỤC TRƯỞNG CỤC CẢNH SÁT QUẢN LÝ HÀNH CHÍNH VỀ TRẬT TỰ XÃ HỘI"
    elif len(fields) >= 10:
        return "BỘ CÔNG AN"
    return "CỤC TRƯỞNG CỤC CẢNH SÁT QUẢN LÝ HÀNH CHÍNH VỀ TRẬT TỰ XÃ HỘI"

def get_card_type(qr_data):
    if not qr_data:
        return ""
    fields = qr_data.split('|')
    if len(fields) == 7:
        return "Căn cước công dân"
    elif len(fields) >= 10:
        return "Căn cước"
    return "Không xác định"

def calculate_expiry_date(dob_str):
    if not dob_str or len(dob_str) != 10:
        return ""
    try:
        day, month, year = dob_str.split('/')
        year = int(year)
        current_year = datetime.datetime.now().year
        for age in [14, 25, 40, 60]:
            expiry_year = year + age
            if expiry_year > current_year:
                return f"{day}/{month}/{expiry_year}"
        return "Không thời hạn"
    except:
        return ""

def process_qr_string(qr_str):
    parts = qr_str.split('|')
    import re
    addr_raw = parts[5] if len(parts) > 5 else ''
    addr_clean = re.sub(r',\s*,', ',', addr_raw) # Thay thế ', ,' hoặc ',,' bằng ','
    addr_clean = re.sub(r'\s+', ' ', addr_clean).strip(', ')

    data = {
        'CCCD': parts[0] if len(parts) > 0 else '',
        'CMND': parts[1] if len(parts) > 1 else '',
        'Họ tên': parts[2] if len(parts) > 2 else '',
        'Ngày sinh': format_date(parts[3]) if len(parts) > 3 else '',
        'Giới tính': parts[4] if len(parts) > 4 else '',
        'Nơi thường trú gốc': addr_clean,
        'Ngày cấp CCCD': format_date(parts[6]) if len(parts) > 6 else '',
    }
    notes = []
    if not data['Ngày cấp CCCD']:
        notes.append('Thiếu ngày cấp')
    if not data['Nơi thường trú gốc']:
        notes.append('Thiếu nơi thường trú')
    return data, notes



async def fetch_single_address_async(client, addr):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:151.0) Gecko/20100101 Firefox/151.0',
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json',
        'x-kas': '89232422',
        'Origin': 'https://tienich.vnhub.com',
        'Referer': 'https://tienich.vnhub.com/'
    }
    import json, asyncio, re
    
    # Tiền xử lý: API VNHub rất nhạy cảm với khoảng trắng thừa, đặc biệt là khoảng trắng trước dấu phẩy
    # Ví dụ: "Số Nhà 137B , Trần Hưng Đạo ," -> lỗi data: []
    clean_addr = re.sub(r'\s+,', ',', addr)
    clean_addr = re.sub(r'\s+', ' ', clean_addr).strip()
    
    payload = json.dumps({"address": clean_addr})
    
    # Retry vô hạn với delay 2s cho đến khi thành công hoặc API trả về "không tìm thấy" hợp lệ
    while True:
        try:
            response = await client.post(
                'https://tienich.vnhub.com/api/wards',
                content=payload,
                headers=headers,
                timeout=15.0
            )
            response.raise_for_status()
            res_data = response.json()
            
            if res_data.get('success') and res_data.get('data') and len(res_data['data']) > 0 and res_data['data'][0].get('address'):
                return {
                    "original": addr,
                    "success": True,
                    "converted": res_data['data'][0]['address']
                }
            
            # API thành công nhưng data rỗng = không tìm thấy địa chỉ → dừng
            return {
                "original": addr,
                "success": False,
                "error": "Không tìm thấy địa chỉ tương ứng"
            }
            
        except Exception:
            # Lỗi 500 hoặc mạng → thử lại sau 2s
            await asyncio.sleep(2)
            continue


async def generate_excel_for_items(items: List[ExportItem], room_id: str = None, duplicate_files: List[str] = None):
    import re
    print(f"\n[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Nhận yêu cầu xuất Excel cho {len(items)} bản ghi.", flush=True)
    
    # Store all original filenames for original.zip
    original_filenames = []
    if duplicate_files:
        for dup in duplicate_files:
            original_filenames.append(dup.get("duplicate") if isinstance(dup, dict) else dup)
    
    records = {} # mapping cccd -> record
    
    # First Pass: collect data from items
    for item in items:
        if item.filename and item.filename not in original_filenames:
            original_filenames.append(item.filename)
            
        cccd = None
        if item.qrData:
            extracted, v_notes = process_qr_string(item.qrData)
            cccd = extracted.get('CCCD', '')
        elif item.fromOCR and item.ocrData:
            cccd = item.ocrData.get('CCCD', '')
            
        if not cccd: continue
        
        if cccd not in records:
            records[cccd] = {
                'Họ tên': '', 'CCCD': cccd, 'CMND': '', 'Giới tính': '',
                'Ngày sinh': '', 'Nơi thường trú gốc': '', 'Địa chỉ chuẩn hóa mới': '',
                'Ngày cấp CCCD': '', 'Nơi cấp': '', 'Ngày hết hạn': '', 'Phân loại': '', 'Ghi chú': [], 'QR Raw': '',
                'Ảnh mặt trước CCCD/CC': '',
                'Ảnh mặt sau CCCD/CC': '',
                'Đổi tên Ảnh mặt trước CCCD/CC': '',
                'Đổi tên Ảnh mặt sau CCCD/CC': '',
                'OCR Image Path Front': '',
                'OCR Image Path Back': '',
                'OCR Image Path Unknown': ''
            }
            
        record = records[cccd]
        
        is_qr = bool(item.qrData)
        if is_qr:
            extracted, v_notes = process_qr_string(item.qrData)
            for k in ['Họ tên', 'CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD', 'Nơi cấp', 'Ngày hết hạn', 'Phân loại', 'QR Raw']:
                if extracted.get(k): record[k] = extracted[k]
                
            record['Nơi cấp'] = get_place_of_issue(item.qrData)
            record['Phân loại'] = get_card_type(item.qrData)
            record['Ngày hết hạn'] = calculate_expiry_date(record.get('Ngày sinh', ''))
            record['QR Raw'] = item.qrData
            
            fields = item.qrData.split('|')
            if len(fields) == 7:
                record['Ảnh mặt trước CCCD/CC'] = item.filename or ''
            elif len(fields) >= 10:
                record['Ảnh mặt sau CCCD/CC'] = item.filename or ''
                
            if v_notes:
                record['Ghi chú'].extend(v_notes)
        else: # OCR
            extracted = item.ocrData
            for k in ['Họ tên', 'CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD']:
                if extracted.get(k) and not record.get(k): record[k] = extracted[k]
                
            side = extracted.get('OCR Side')
            if side == 'Front':
                record['OCR Image Path Front'] = item.filename or ''
            elif side == 'Back':
                record['OCR Image Path Back'] = item.filename or ''
            else:
                record['OCR Image Path Unknown'] = item.filename or ''
            
            record['Ghi chú'].append("Lấy bằng OCR")
            
    # ---------- FUZZY MATCH: GỘP MẶT SAU OCR VÀO ĐÚNG BẢN GHI (2/3 TRƯỜNG) ----------
    import unicodedata

    def _norm_for_match(text):
        if not text: return ''
        text = text.upper().strip()
        nfkd = unicodedata.normalize('NFKD', text)
        return ' '.join(''.join(c for c in nfkd if not unicodedata.combining(c)).split())

    orphan_cccds = [
        c for c, rec in records.items()
        if (not rec.get('QR Raw')
            and rec.get('OCR Image Path Back')
            and not rec.get('Ảnh mặt trước CCCD/CC')
            and not rec.get('OCR Image Path Front'))
    ]

    for orphan_cccd in orphan_cccds:
        orphan = records[orphan_cccd]
        o_cccd = orphan.get('CCCD', '')
        o_name = _norm_for_match(orphan.get('Họ tên', ''))
        o_date = orphan.get('Ngày cấp CCCD', '')

        best_cccd = None
        for r_cccd, r_rec in records.items():
            if r_cccd == orphan_cccd:
                continue
            if not (r_rec.get('QR Raw')
                    or r_rec.get('Ảnh mặt trước CCCD/CC')
                    or r_rec.get('OCR Image Path Front')):
                continue
            r_name = _norm_for_match(r_rec.get('Họ tên', ''))
            r_date = r_rec.get('Ngày cấp CCCD', '')

            score = 0
            if o_cccd and o_cccd == r_cccd:              score += 1
            if o_name and r_name and o_name == r_name:   score += 1
            if o_date and r_date and o_date == r_date:   score += 1
            if score >= 2:
                best_cccd = r_cccd
                break

        # Nếu không tìm thấy bằng tiêu chí 2/3 trường, thử tiêu chí linh động hơn:
        # Cần 8 số liên tiếp của CCCD HOẶC trùng khớp Ngày cấp
        # VÀ dòng đích phải ĐANG THIẾU mặt sau
        if not best_cccd:
            for r_cccd, r_rec in records.items():
                if r_cccd == orphan_cccd:
                    continue
                if not (r_rec.get('QR Raw')
                        or r_rec.get('Ảnh mặt trước CCCD/CC')
                        or r_rec.get('OCR Image Path Front')):
                    continue
                # Bắt buộc dòng đích phải thiếu mặt sau mới được ghép bù vào
                if r_rec.get('OCR Image Path Back') or r_rec.get('Ảnh mặt sau CCCD/CC'):
                    continue

                match_8_digits = False
                if len(o_cccd) >= 8 and len(r_cccd) >= 8:
                    for i in range(len(r_cccd) - 7):
                        if r_cccd[i:i+8] in o_cccd:
                            match_8_digits = True
                            break

                r_date = r_rec.get('Ngày cấp CCCD', '')
                match_date = (o_date and r_date and o_date == r_date)

                if match_8_digits or match_date:
                    best_cccd = r_cccd
                    break

        if best_cccd:
            target = records[best_cccd]
            if not target.get('OCR Image Path Back'):
                target['OCR Image Path Back'] = orphan.get('OCR Image Path Back', '')
            for k in ['Họ tên', 'Ngày cấp CCCD']:
                if orphan.get(k) and not target.get(k):
                    target[k] = orphan[k]
            print(f"   [FUZZY MATCH] Ghép mặt sau (CCCD OCR: {orphan_cccd}) vào bản ghi {best_cccd} (khớp 2/3 trường).")
            del records[orphan_cccd]

    # -------------------------------------------------------------------------
    # BƯỚC QUAN TRỌNG: GÁN ẢNH OCR VÀO ĐÚNG MẶT THẺ & MERGE DỮ LIỆU
    # Nếu ảnh quét QR bị mờ/khuyết, nhưng ảnh OCR (không quét được QR) có dữ liệu
    # Hệ thống sẽ dựa vào OCR Side (Mặt Trước/Mặt Sau) để "đắp" vào chỗ trống.
    # -------------------------------------------------------------------------
    for cccd, record in records.items():
        # --- Gán ảnh OCR Mặt Trước ---
        if not record['Ảnh mặt trước CCCD/CC']:
            if record.get('OCR Image Path Front'):
                record['Ảnh mặt trước CCCD/CC'] = record.pop('OCR Image Path Front')
            # Nếu OCR không nhận diện được mặt (Unknown) nhưng đã có ảnh Mặt Sau -> Ảnh Unknown chắc chắn là Mặt Trước
            elif record.get('OCR Image Path Unknown') and record['Ảnh mặt sau CCCD/CC']:
                record['Ảnh mặt trước CCCD/CC'] = record['OCR Image Path Unknown']
                record['OCR Image Path Unknown'] = ''
                
        # --- Gán ảnh OCR Mặt Sau ---
        if not record['Ảnh mặt sau CCCD/CC']:
            if record.get('OCR Image Path Back'):
                record['Ảnh mặt sau CCCD/CC'] = record.pop('OCR Image Path Back')
            # Nếu OCR không nhận diện được mặt (Unknown) nhưng đã có ảnh Mặt Trước -> Ảnh Unknown chắc chắn là Mặt Sau
            elif record.get('OCR Image Path Unknown') and record['Ảnh mặt trước CCCD/CC']:
                record['Ảnh mặt sau CCCD/CC'] = record['OCR Image Path Unknown']
                record['OCR Image Path Unknown'] = ''

        # --- Báo lỗi nếu 1 người thiếu cả 2 mặt ---
        if not record['Ảnh mặt trước CCCD/CC'] and not record['Ảnh mặt sau CCCD/CC']:
            if record.get('OCR Image Path Unknown'):
                record['Ghi chú'].append('Không thể phân biệt được ảnh này là mặt trước hay mặt sau do mờ và không chứa mã QR')

        hoten = record['Họ tên'] or 'KhongTen'
        cmnd = record['CMND']
        hoten_clean = re.sub(r'[\\/*?:"<>|]', '', hoten)
        cmnd_str = f"_{cmnd}" if cmnd else ""
        
        if record['Ảnh mặt trước CCCD/CC']:
            ext = os.path.splitext(record['Ảnh mặt trước CCCD/CC'])[1]
            record['Đổi tên Ảnh mặt trước CCCD/CC'] = f"{hoten_clean}_{cccd}{cmnd_str}_Mặt trước{ext}"
            
        if record['Ảnh mặt sau CCCD/CC']:
            ext = os.path.splitext(record['Ảnh mặt sau CCCD/CC'])[1]
            record['Đổi tên Ảnh mặt sau CCCD/CC'] = f"{hoten_clean}_{cccd}{cmnd_str}_Mặt sau{ext}"

        # Xử lý logic CMND (Yêu cầu mới)
        if not record['CMND']:
            if record.get('QR Raw'):
                record['CMND'] = 'Không có'
            else:
                record['CMND'] = 'Chưa xác định'

        # Tính toán ngày hết hạn dựa trên ngày sinh nếu bị khuyết
        if not record['Ngày hết hạn'] and record.get('Ngày sinh'):
            record['Ngày hết hạn'] = calculate_expiry_date(record['Ngày sinh'])
            record['Phân loại'] = 'Căn cước / CCCD'

        # Lọc trùng lặp ghi chú
        raw_notes = record['Ghi chú']
        if isinstance(raw_notes, str):
            raw_notes = raw_notes.split('; ')
        elif not isinstance(raw_notes, list):
            raw_notes = []
        unique_notes = []
        for note in [n for n in raw_notes if n]:
            for subnote in note.split('; '):
                if subnote and subnote not in unique_notes:
                    unique_notes.append(subnote)
        record['Ghi chú'] = '; '.join(unique_notes)

    processed_data = list(records.values())

    # Thu thập danh sách ảnh "không thuộc dòng nào" (unknown)
    # = ảnh có trong phiên nhưng không có CCCD nào được trích xuất
    all_matched_filenames = set()
    for row in processed_data:
        for field in ['Ảnh mặt trước CCCD/CC', 'Ảnh mặt sau CCCD/CC']:
            f = row.get(field)
            if f:
                all_matched_filenames.add(f)

    unknown_filenames = []
    if room_id:
        room_img_dir = os.path.join(SESSIONS_DIR, room_id, "images")
        if os.path.isdir(room_img_dir):
            for fname in sorted(os.listdir(room_img_dir)):
                if fname not in all_matched_filenames:
                    unknown_filenames.append(fname)

    unique_addresses = list(set([row['Nơi thường trú gốc'] for row in processed_data if row.get('Nơi thường trú gốc')]))
    total_addr = len(unique_addresses)
    print(f"-> Phát hiện {total_addr} địa chỉ độc nhất cần chuẩn hóa qua VNHub.", flush=True)
    
    # Run async API calls - dùng as_completed để in tiến trình từng địa chỉ
    address_map = {}
    if unique_addresses:
        async with httpx.AsyncClient() as client:
            tasks = {asyncio.ensure_future(fetch_single_address_async(client, addr)): addr for addr in unique_addresses}
            done_count = 0
            pending = set(tasks.keys())
            while pending:
                done, pending = await asyncio.wait(pending, return_when=asyncio.FIRST_COMPLETED)
                for fut in done:
                    res = fut.result()
                    done_count += 1
                    address_map[res['original']] = res
                    short = res['original'][:50]
                    if res['success']:
                        print(f"  [{done_count}/{total_addr}] ✓ {short} -> {res['converted']}", flush=True)
                    else:
                        print(f"  [{done_count}/{total_addr}] ✗ {short}: {res.get('error')}", flush=True)
                    if room_id:
                        await manager.broadcast({
                            'type': 'api_progress',
                            'current': done_count,
                            'total': total_addr
                        }, room_id)
            
    # Map back
    for row in processed_data:
        addr = row['Nơi thường trú gốc']
        if addr and addr in address_map:
            result = address_map[addr]
            notes = row['Ghi chú'] if isinstance(row['Ghi chú'], list) else []
            
            if result['success']:
                converted_addr = result.get('converted', '')
                row['Địa chỉ chuẩn hóa mới'] = converted_addr
                if converted_addr.lower().strip() == addr.lower().strip():
                    notes.append("Địa chỉ không đổi")
            else:
                notes.append(result.get('error', 'Lỗi không xác định'))
                
            row['Ghi chú'] = '; '.join(notes)
            
    # Clean up notes and ensure they are strings
    for row in processed_data:
        if isinstance(row['Ghi chú'], list):
            if row.get('QR Raw'):
                row['Ghi chú'] = [n for n in row['Ghi chú'] if 'Lấy bằng OCR' not in n]
                if 'Đọc mã QR' not in row['Ghi chú']:
                    row['Ghi chú'].insert(0, 'Đọc mã QR')
            
            unique_notes = []
            for n in row['Ghi chú']:
                if n not in unique_notes:
                    unique_notes.append(n)
            row['Ghi chú'] = '; '.join(unique_notes)
            
    print(f"-> Hoàn tất xử lý dữ liệu. Đang đóng gói file Excel...", flush=True)

    # Xác định dòng "chưa đầy đủ thông tin" để đưa vào review
    REQUIRED_FIELDS = ['Họ tên', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD',
                       'Ảnh mặt trước CCCD/CC', 'Ảnh mặt sau CCCD/CC']
    review_rows = []
    for row in processed_data:
        missing = [f for f in REQUIRED_FIELDS if not row.get(f)]
        if missing:
            review_rows.append((row, missing))
    
    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    headers = [
        "STT", "Họ tên", "CCCD", "CMND", "Giới tính", "Ngày sinh", 
        "Nơi thường trú gốc", "Địa chỉ chuẩn hóa mới", "Ngày cấp CCCD", "Nơi cấp", "Ngày hết hạn", "Phân loại", "Ghi chú", "QR Raw", 
        "Ảnh mặt trước CCCD/CC", "Ảnh mặt sau CCCD/CC", "Đổi tên Ảnh mặt trước CCCD/CC", "Đổi tên Ảnh mặt sau CCCD/CC"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        
    for row_idx, data in enumerate(processed_data, 2):
        ws.cell(row=row_idx, column=1, value=row_idx-1)
        ws.cell(row=row_idx, column=2, value=data.get('Họ tên', ''))
        c_cell = ws.cell(row=row_idx, column=3, value=data.get('CCCD', ''))
        c_cell.number_format = '@'
        cm_cell = ws.cell(row=row_idx, column=4, value=data.get('CMND', ''))
        cm_cell.number_format = '@'
        ws.cell(row=row_idx, column=5, value=data.get('Giới tính', ''))
        ws.cell(row=row_idx, column=6, value=data.get('Ngày sinh', ''))
        ws.cell(row=row_idx, column=7, value=data.get('Nơi thường trú gốc', ''))
        ws.cell(row=row_idx, column=8, value=data.get('Địa chỉ chuẩn hóa mới', ''))
        ws.cell(row=row_idx, column=9, value=data.get('Ngày cấp CCCD', ''))
        ws.cell(row=row_idx, column=10, value=data.get('Nơi cấp', ''))
        ws.cell(row=row_idx, column=11, value=data.get('Ngày hết hạn', ''))
        ws.cell(row=row_idx, column=12, value=data.get('Phân loại', ''))
        ws.cell(row=row_idx, column=13, value=data.get('Ghi chú', ''))
        ws.cell(row=row_idx, column=14, value=data.get('QR Raw', ''))
        ws.cell(row=row_idx, column=15, value=data.get('Ảnh mặt trước CCCD/CC', ''))
        ws.cell(row=row_idx, column=16, value=data.get('Ảnh mặt sau CCCD/CC', ''))
        ws.cell(row=row_idx, column=17, value=data.get('Đổi tên Ảnh mặt trước CCCD/CC', ''))
        ws.cell(row=row_idx, column=18, value=data.get('Đổi tên Ảnh mặt sau CCCD/CC', ''))
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2), 40)
        ws.column_dimensions[column].width = adjusted_width
        
    ws_dup = wb.create_sheet(title="duplicate")
    ws_dup.append(["STT", "Tên file", "Trùng lặp với"])
    if duplicate_files:
        for i, dup in enumerate(duplicate_files, 1):
            if isinstance(dup, dict):
                ws_dup.append([i, dup.get('duplicate', ''), dup.get('original', '')])
            else:
                ws_dup.append([i, str(dup), ''])
                
    for col in ws_dup.columns:
        ws_dup.column_dimensions[col[0].column_letter].width = 30

    # --- Sheet "Review": dòng dữ liệu chưa đầy đủ ---
    ws_review = wb.create_sheet(title="Review")
    review_headers = ["STT", "CCCD", "Họ tên", "Ảnh mặt trước", "Ảnh mặt sau", "Trường còn thiếu"]
    for col_idx, h in enumerate(review_headers, 1):
        cell = ws_review.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
    for stt, (row, missing) in enumerate(review_rows, 1):
        ws_review.cell(row=stt+1, column=1, value=stt)
        ws_review.cell(row=stt+1, column=2, value=row.get('CCCD', ''))
        ws_review.cell(row=stt+1, column=3, value=row.get('Họ tên', ''))
        ws_review.cell(row=stt+1, column=4, value=row.get('Ảnh mặt trước CCCD/CC', ''))
        ws_review.cell(row=stt+1, column=5, value=row.get('Ảnh mặt sau CCCD/CC', ''))
        ws_review.cell(row=stt+1, column=6, value=', '.join(missing))
    for col in ws_review.columns:
        ws_review.column_dimensions[col[0].column_letter].width = 30

    # --- Sheet "Unknown": ảnh không thuộc dòng nào ---
    ws_unknown = wb.create_sheet(title="Unknown")
    unknown_headers = ["STT", "Tên file gốc"]
    for col_idx, h in enumerate(unknown_headers, 1):
        cell = ws_unknown.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True)
    for stt, fname in enumerate(unknown_filenames, 1):
        ws_unknown.cell(row=stt+1, column=1, value=stt)
        ws_unknown.cell(row=stt+1, column=2, value=fname)
    for col in ws_unknown.columns:
        ws_unknown.column_dimensions[col[0].column_letter].width = 40

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Backup 1 bản lưu vào exports
    try:
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        exports_dir = os.path.join(project_dir, 'webapp', 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%d%m%Y_%H%M%S")
        backup_filepath = os.path.join(exports_dir, f"backup_ket_qua_{timestamp}.xlsx")
        wb.save(backup_filepath)
        print(f"-> Đã tự động backup file Excel tại: {backup_filepath}", flush=True)
    except Exception as e:
        print(f"-> Lỗi khi lưu backup file Excel: {e}", flush=True)
    
    # Create a zip containing the excel and sub zips
    import zipfile
    
    zip_stream = BytesIO()
    with zipfile.ZipFile(zip_stream, 'w') as main_zip:
        main_zip.writestr('ket_qua.xlsx', stream.getvalue())
        
        if room_id:
            room_img_dir = os.path.join(SESSIONS_DIR, room_id, "images")

            # 1. original.zip
            original_zip_stream = BytesIO()
            with zipfile.ZipFile(original_zip_stream, 'w') as sub_zip:
                for fname in original_filenames:
                    img_path = os.path.join(room_img_dir, fname)
                    if os.path.exists(img_path):
                        sub_zip.write(img_path, fname)
            main_zip.writestr('original.zip', original_zip_stream.getvalue())
            
            # 2. rename.zip
            rename_zip_stream = BytesIO()
            with zipfile.ZipFile(rename_zip_stream, 'w') as sub_zip:
                for row in processed_data:
                    folder = "CCCD" if row.get("Phân loại") == "Căn cước công dân" else "CC"
                    
                    front_fname = row.get('Ảnh mặt trước CCCD/CC')
                    front_renamed = row.get('Đổi tên Ảnh mặt trước CCCD/CC')
                    if front_fname and front_renamed:
                        img_path = os.path.join(room_img_dir, front_fname)
                        if os.path.exists(img_path):
                            sub_zip.write(img_path, f"{folder}/{front_renamed}")
                            
                    back_fname = row.get('Ảnh mặt sau CCCD/CC')
                    back_renamed = row.get('Đổi tên Ảnh mặt sau CCCD/CC')
                    if back_fname and back_renamed:
                        img_path = os.path.join(room_img_dir, back_fname)
                        if os.path.exists(img_path):
                            sub_zip.write(img_path, f"{folder}/{back_renamed}")
            main_zip.writestr('rename.zip', rename_zip_stream.getvalue())

            # 3. review.zip: ảnh của các dòng chưa đầy đủ thông tin
            if review_rows:
                review_zip_stream = BytesIO()
                with zipfile.ZipFile(review_zip_stream, 'w') as sub_zip:
                    added = set()
                    for row, _ in review_rows:
                        for field in ['Ảnh mặt trước CCCD/CC', 'Ảnh mặt sau CCCD/CC']:
                            fname = row.get(field)
                            if fname and fname not in added:
                                img_path = os.path.join(room_img_dir, fname)
                                if os.path.exists(img_path):
                                    sub_zip.write(img_path, fname)
                                    added.add(fname)
                main_zip.writestr('review.zip', review_zip_stream.getvalue())

            # 4. unknown.zip: ảnh không thuộc dòng nào
            if unknown_filenames:
                unknown_zip_stream = BytesIO()
                with zipfile.ZipFile(unknown_zip_stream, 'w') as sub_zip:
                    for fname in unknown_filenames:
                        img_path = os.path.join(room_img_dir, fname)
                        if os.path.exists(img_path):
                            sub_zip.write(img_path, fname)
                main_zip.writestr('unknown.zip', unknown_zip_stream.getvalue())
        
    zip_stream.seek(0)
    
    # Return ZIP file
    headers_dict = {
        'Content-Disposition': 'attachment; filename="result.zip"',
        'Access-Control-Expose-Headers': 'Content-Disposition'
    }
    return StreamingResponse(iter([zip_stream.getvalue()]), media_type="application/zip", headers=headers_dict)



class ExportRequest(BaseModel):
    data: List[ExportItem]

@app.post("/api/export")
async def export_excel(req: ExportRequest):
    return await generate_excel_for_items(req.data)

class RoomExportRequest(BaseModel):
    room_id: str

@app.post("/api/room/export")
async def room_export(req: RoomExportRequest):
    room_id = req.room_id
    room = get_room(room_id)
    items = [ExportItem(**i) for i in room["items"]]
    duplicate_files = room.get("duplicate_files", [])
    return await generate_excel_for_items(items, room_id, duplicate_files)

# Mount public directory as root static files
base_dir = os.path.dirname(os.path.abspath(__file__))
public_dir = os.path.join(base_dir, 'public')
app.mount("/", StaticFiles(directory=public_dir, html=True), name="static")
