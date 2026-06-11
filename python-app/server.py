import os
import sys
import json
import base64
import uuid
import datetime
import asyncio
import httpx
from io import BytesIO
import numpy as np

from fastapi import FastAPI, UploadFile, File, Request, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional, Dict, Any

import cv2
from pyzbar.pyzbar import decode, ZBarSymbol
import openpyxl
from openpyxl.styles import Font

app = FastAPI(title="CCCD QR API")

# Setup CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global model variables
detector = None

@app.on_event("startup")
async def startup_event():
    cleanup_old_sessions()
    await load_models()

async def load_models():

    global detector
    model_paths = [
        'models/detect.prototxt', 'models/detect.caffemodel',
        'models/sr.prototxt', 'models/sr.caffemodel'
    ]
    base_dir = os.path.dirname(os.path.abspath(__file__))
    abs_paths = [os.path.join(base_dir, p) for p in model_paths]
    if all(os.path.exists(p) for p in abs_paths):
        try:
            detector = cv2.wechat_qrcode_WeChatQRCode(*abs_paths)
            print("Loaded WeChat QRCode detector.")
        except Exception as e:
            print(f"Error loading detector: {e}")
    else:
        print("Model files not found.")

# ---- Room & WebSocket Sync Manager ----
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = {}

    async def connect(self, websocket: WebSocket, room_id: str):
        await websocket.accept()
        if room_id not in self.active_connections:
            self.active_connections[room_id] = []
        self.active_connections[room_id].append(websocket)

    def disconnect(self, websocket: WebSocket, room_id: str):
        if room_id in self.active_connections:
            if websocket in self.active_connections[room_id]:
                self.active_connections[room_id].remove(websocket)
            if not self.active_connections[room_id]:
                del self.active_connections[room_id]

    async def broadcast(self, message: dict, room_id: str):
        if room_id in self.active_connections:
            for connection in self.active_connections[room_id]:
                try:
                    await connection.send_json(message)
                except Exception:
                    pass

manager = ConnectionManager()
rooms: Dict[str, Dict[str, Any]] = {}
import time
import glob

SESSIONS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sessions")
os.makedirs(SESSIONS_DIR, exist_ok=True)

def cleanup_old_sessions():
    try:
        now = time.time()
        count = 0
        for filepath in glob.glob(os.path.join(SESSIONS_DIR, "*.json")):
            if os.path.isfile(filepath):
                # Check file modified time or read JSON
                # 10 days = 10 * 24 * 60 * 60 seconds = 864000
                if now - os.path.getmtime(filepath) > 864000:
                    os.remove(filepath)
                    count += 1
        if count > 0:
            print(f"Cleaned up {count} old session(s) > 10 days.")
    except Exception as e:
        print(f"Error cleaning up old sessions: {e}")


def save_room(room_id: str):
    try:
        if room_id in rooms:
            filepath = os.path.join(SESSIONS_DIR, f"{room_id}.json")
            data_to_save = {
                "items": rooms[room_id]["items"],
                "seen_cccds": list(rooms[room_id]["seen_cccds"])
            }
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data_to_save, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving session {room_id}: {e}")

ROOMS_BACKUP_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rooms_backup.json")

def cleanup_old_sessions():
    global rooms
    if os.path.exists(ROOMS_BACKUP_FILE):
        try:
            with open(ROOMS_BACKUP_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                for room_id, room_data in data.items():
                    rooms[room_id] = {
                        "items": room_data.get("items", []),
                        "seen_cccds": set(room_data.get("seen_cccds", []))
                    }
            print(f"Loaded {len(rooms)} rooms from backup.")
        except Exception as e:
            print(f"Error loading rooms backup: {e}")

def save_room(room_id):
    try:
        data_to_save = {}
        for room_id, room_data in rooms.items():
            data_to_save[room_id] = {
                "items": room_data["items"],
                "seen_cccds": list(room_data["seen_cccds"])
            }
        with open(ROOMS_BACKUP_FILE, "w", encoding="utf-8") as f:
            json.dump(data_to_save, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving rooms backup: {e}")


@app.websocket("/ws/room/{room_id}")
async def websocket_endpoint(websocket: WebSocket, room_id: str):
    await manager.connect(websocket, room_id)
    try:
        while True:
            _ = await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket, room_id)

class ExportItem(BaseModel):
    filename: Optional[str] = None
    qrData: Optional[str] = None
    error: Optional[str] = None
    fromOCR: Optional[bool] = False
    ocrData: Optional[Dict[str, str]] = None

class RoomAddRequest(BaseModel):
    room_id: str
    item: ExportItem

@app.post("/api/room/add")
async def room_add(req: RoomAddRequest):
    room_id = req.room_id
    item = req.item
    room = get_room(room_id)
    
    cccd_num = None
    if item.qrData:
        parts = item.qrData.split('|')
        cccd_num = parts[0] if parts else None
    elif item.fromOCR and item.ocrData:
        cccd_num = item.ocrData.get('CCCD')
        
    if cccd_num:
        if cccd_num in room["seen_cccds"]:
            # If an OCR is already there and we found a QR, we should replace it.
            # But for simplicity, if it's in seen_cccds, we check if we need to upgrade from OCR to QR.
            # In Python, we can find the index and replace.
            existing_idx = next((i for i, v in enumerate(room["items"]) 
                                 if (v.get("qrData") and v["qrData"].startswith(cccd_num)) or 
                                    (v.get("ocrData") and v["ocrData"].get("CCCD") == cccd_num)), None)
            
            if existing_idx is not None:
                existing_item = room["items"][existing_idx]
                if not existing_item.get("qrData") and item.qrData:
                    # Upgrade OCR to QR
                    room["items"][existing_idx] = item.dict()
                    await manager.broadcast({
                        "type": "update_item",
                        "items": room["items"],
                        "total_count": len(room["items"])
                    }, room_id)
                    save_room(room_id)
                    return {"success": True, "total_count": len(room["items"])}
                else:
                    # Both are OCR or both are QR, check which has more info
                    def count_info(d):
                        if not d: return 0
                        keys = ['Họ tên', 'CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD']
                        return sum(1 for k in keys if d.get(k))
                    
                    old_count = 0
                    new_count = 0
                    
                    # For OCR, compare ocrData
                    if not existing_item.get("qrData") and not item.qrData:
                        old_count = count_info(existing_item.get("ocrData"))
                        new_count = count_info(item.ocrData)
                    
                    if new_count > old_count:
                        room["items"][existing_idx] = item.dict()
                        await manager.broadcast({
                            "type": "update_item",
                            "items": room["items"],
                            "total_count": len(room["items"])
                        }, room_id)
                        save_room(room_id)
                        return {"success": True, "total_count": len(room["items"])}
                    else:
                        return {"success": False, "error": "Duplicate CCCD"}
        else:
            room["seen_cccds"].add(cccd_num)
            
    item_dict = item.dict()
    room["items"].append(item_dict)
    
    await manager.broadcast({
        "type": "new_item",
        "item": item_dict,
        "total_count": len(room["items"])
    }, room_id)
    save_room(room_id)
    return {"success": True, "total_count": len(room["items"])}

@app.get("/api/room/state/{room_id}")
async def room_state(room_id: str):
    room = get_room(room_id)
    return {"success": True, "items": room["items"], "total_count": len(room["items"])}

class RoomClearRequest(BaseModel):
    room_id: str

@app.post("/api/room/clear")
async def room_clear(req: RoomClearRequest):
    room_id = req.room_id
    if room_id in rooms:
        rooms[room_id] = {"items": [], "seen_cccds": set()}
        save_room(room_id)
        await manager.broadcast({
            "type": "clear",
            "total_count": 0
        }, room_id)
    return {"success": True}

# ---- End Room Manager ----

class ScanQRRequest(BaseModel):
    imageBase64: str
    filename: Optional[str] = None

class LogCameraRequest(BaseModel):
    message: str

@app.post("/api/log_camera")
async def log_camera(req: LogCameraRequest):
    print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {req.message}", flush=True)
    return {"success": True}

@app.post("/api/scan_qr")
async def scan_qr(req: ScanQRRequest):
    fname = f"'{req.filename}'" if req.filename else "từ Live Camera"
    print(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Nhận yêu cầu quét ảnh {fname}...", flush=True)
    if not detector:
        print("-> Lỗi: Mô hình WeChat QRCode chưa được tải.", flush=True)
        return {"success": False, "error": "Model not loaded"}
    
    # decode base64
    base64_str = req.imageBase64
    if "," in base64_str:
        base64_str = base64_str.split(",")[1]
    
    try:
        img_data = base64.b64decode(base64_str)
        nparr = np.frombuffer(img_data, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)
        
        if img is None:
            print("-> Lỗi: Dữ liệu ảnh không hợp lệ.", flush=True)
            return {"success": False, "error": "Invalid image"}
            
        # 1. Try decoding directly with pyzbar
        decoded_objects = decode(img, symbols=[ZBarSymbol.QRCODE])
        
        # 2. Try grayscale and thresholding for blurry/dark images
        if not decoded_objects:
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            decoded_objects = decode(gray, symbols=[ZBarSymbol.QRCODE])
            
            if not decoded_objects:
                # Apply adaptive thresholding
                thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
                decoded_objects = decode(thresh, symbols=[ZBarSymbol.QRCODE])
                
        if decoded_objects:
            qr_data = decoded_objects[0].data.decode('utf-8')
            if '|' in qr_data and len(qr_data.split('|')) >= 6:
                print(f"-> Quét thành công QR (pyzbar): {qr_data[:50]}...", flush=True)
                return {"success": True, "data": qr_data}
            else:
                print(f"-> pyzbar phát hiện QR rác (không phải CCCD), chuyển sang WeChat...", flush=True)
            
        # 3. Fallback to WeChat QRCode CNN detector
        if detector:
            res, _ = detector.detectAndDecode(img)
            if res and len(res) > 0:
                print(f"-> Quét thành công QR (WeChat): {res[0][:50]}...", flush=True)
                return {"success": True, "data": res[0]}
            
        print("-> Không tìm thấy mã QR trong ảnh.", flush=True)
        return {"success": False, "error": "QR not found"}
    except Exception as e:
        print(f"-> Lỗi hệ thống khi quét ảnh: {str(e)}", flush=True)
        return {"success": False, "error": str(e)}

def format_date(date_str):
    if not date_str or len(date_str) != 8:
        return date_str
    return f"{date_str[0:2]}/{date_str[2:4]}/{date_str[4:8]}"

def process_qr_string(qr_str):
    parts = qr_str.split('|')
    data = {
        'CCCD': parts[0] if len(parts) > 0 else '',
        'CMND': parts[1] if len(parts) > 1 else '',
        'Họ tên': parts[2] if len(parts) > 2 else '',
        'Ngày sinh': format_date(parts[3]) if len(parts) > 3 else '',
        'Giới tính': parts[4] if len(parts) > 4 else '',
        'Nơi thường trú gốc': parts[5] if len(parts) > 5 else '',
        'Ngày cấp CCCD': format_date(parts[6]) if len(parts) > 6 else '',
    }
    notes = []
    if not data['Ngày cấp CCCD']:
        notes.append('Thiếu ngày cấp')
    if not data['Nơi thường trú gốc']:
        notes.append('Thiếu nơi thường trú')
    return data, notes

async def fetch_single_address_async(client, addr):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:151.0) Gecko/20100101 Firefox/151.0',
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json',
        'x-kas': '89232422',
        'Origin': 'https://tienich.vnhub.com',
        'Referer': 'https://tienich.vnhub.com/'
    }
    try:
        response = await client.post(
            'https://tienich.vnhub.com/api/wards',
            json={"address": addr},
            headers=headers,
            timeout=15.0
        )
        response.raise_for_status()
        res_data = response.json()
        if res_data.get('success') and res_data.get('data') and len(res_data['data']) > 0 and res_data['data'][0].get('address'):
            return {
                "original": addr,
                "success": True,
                "converted": res_data['data'][0]['address']
            }
        else:
            err_msg = "Không tìm thấy địa chỉ tương ứng"
            if res_data.get('success') is False and res_data.get('error'):
                err_msg = f"API bị lỗi: {res_data.get('error')}"
            return {
                "original": addr,
                "success": False,
                "error": err_msg
            }
    except Exception as e:
        return {
            "original": addr,
            "success": False,
            "error": f"Lỗi API: {str(e)}"
        }


async def generate_excel_for_items(items: List[ExportItem]):
    # Sort items: QR data first, then OCR, then errors.
    items.sort(key=lambda x: 0 if x.qrData else (1 if x.fromOCR else 2))
    
    print(f"\n[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] Nhận yêu cầu xuất Excel cho {len(items)} bản ghi.", flush=True)
    processed_data = []
    seen_cccds = set()
    
    for item in items:
        row = {
            'Họ tên': '', 'CCCD': '', 'CMND': '', 'Giới tính': '',
            'Ngày sinh': '', 'Nơi thường trú gốc': '', 'Địa chỉ chuẩn hóa mới': '',
            'Ngày cấp CCCD': '', 'Ghi chú': '', 'QR Raw': ''
        }
        notes = []
        
        if item.error:
            notes.append(item.error)
        elif item.qrData:
            row['QR Raw'] = item.qrData
            extracted, v_notes = process_qr_string(item.qrData)
            
            cccd_num = extracted.get('CCCD', '')
            if cccd_num:
                if cccd_num in seen_cccds:
                    continue
                seen_cccds.add(cccd_num)
                
            row.update(extracted)
            notes.extend(v_notes)
        elif item.fromOCR and item.ocrData:
            extracted = item.ocrData
            notes.append("Lấy bằng OCR")
            
            cccd_num = extracted.get('CCCD', '')
            if cccd_num:
                if cccd_num in seen_cccds:
                    continue
                seen_cccds.add(cccd_num)
                
            row.update(extracted)
            
        row['Ghi chú'] = '; '.join(notes)
        processed_data.append(row)
        
    unique_addresses = list(set([row['Nơi thường trú gốc'] for row in processed_data if row.get('Nơi thường trú gốc')]))
    print(f"-> Phát hiện {len(unique_addresses)} địa chỉ độc nhất cần chuẩn hóa qua VNHub.", flush=True)
    
    # Run async API calls
    address_map = {}
    async with httpx.AsyncClient() as client:
        tasks = [fetch_single_address_async(client, addr) for addr in unique_addresses]
        results = await asyncio.gather(*tasks)
        for res in results:
            address_map[res['original']] = res
            if res['success']:
                print(f"  + Thành công: '{res['original']}' -> '{res['converted']}'", flush=True)
            else:
                print(f"  + Thất bại ({res['original']}): {res.get('error')}", flush=True)
            
    # Map back
    for row in processed_data:
        addr = row['Nơi thường trú gốc']
        if addr and addr in address_map:
            result = address_map[addr]
            notes = [row['Ghi chú']] if row['Ghi chú'] else []
            
            if result['success']:
                row['Địa chỉ chuẩn hóa mới'] = result.get('converted', '')
            else:
                notes.append(result.get('error', 'Lỗi không xác định'))
                
            row['Ghi chú'] = '; '.join(notes)
            
    print(f"-> Hoàn tất xử lý dữ liệu. Đang đóng gói file Excel...", flush=True)
    
    # Generate Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    
    headers = [
        "STT", "Họ tên", "CCCD", "CMND", "Giới tính", "Ngày sinh", 
        "Nơi thường trú gốc", "Địa chỉ chuẩn hóa mới", "Ngày cấp CCCD", "Ghi chú", "QR Raw"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        
    for row_idx, data in enumerate(processed_data, 2):
        ws.cell(row=row_idx, column=1, value=row_idx-1)
        ws.cell(row=row_idx, column=2, value=data.get('Họ tên', ''))
        # Using string to keep leading zeros
        c_cell = ws.cell(row=row_idx, column=3, value=data.get('CCCD', ''))
        c_cell.number_format = '@'
        cm_cell = ws.cell(row=row_idx, column=4, value=data.get('CMND', ''))
        cm_cell.number_format = '@'
        
        ws.cell(row=row_idx, column=5, value=data.get('Giới tính', ''))
        ws.cell(row=row_idx, column=6, value=data.get('Ngày sinh', ''))
        ws.cell(row=row_idx, column=7, value=data.get('Nơi thường trú gốc', ''))
        ws.cell(row=row_idx, column=8, value=data.get('Địa chỉ chuẩn hóa mới', ''))
        ws.cell(row=row_idx, column=9, value=data.get('Ngày cấp CCCD', ''))
        ws.cell(row=row_idx, column=10, value=data.get('Ghi chú', ''))
        ws.cell(row=row_idx, column=11, value=data.get('QR Raw', ''))
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    # Backup 1 bản lưu vào web-app/exports
    try:
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        exports_dir = os.path.join(project_dir, 'web-app', 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filepath = os.path.join(exports_dir, f"backup_ket_qua_{timestamp}.xlsx")
        wb.save(backup_filepath)
        print(f"-> Đã tự động backup file Excel tại: {backup_filepath}", flush=True)
    except Exception as e:
        print(f"-> Lỗi khi lưu backup file Excel: {e}", flush=True)
    
    # Return Excel file
    headers_dict = {
        'Content-Disposition': 'attachment; filename="ket_qua.xlsx"',
        'Access-Control-Expose-Headers': 'Content-Disposition'
    }
    return StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_dict)


class ExportRequest(BaseModel):
    data: List[ExportItem]

@app.post("/api/export")
async def export_excel(req: ExportRequest):
    return await generate_excel_for_items(req.data)

class RoomExportRequest(BaseModel):
    room_id: str

@app.post("/api/room/export")
async def room_export(req: RoomExportRequest):
    room_id = req.room_id
    room = get_room(room_id)
    items = [ExportItem(**i) for i in room["items"]]
    return await generate_excel_for_items(items)

# Mount web-app directory as root static files
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
web_app_dir = os.path.join(base_dir, 'web-app')
app.mount("/", StaticFiles(directory=web_app_dir, html=True), name="static")
