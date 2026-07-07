import os
import glob
import pandas as pd
import unicodedata
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime

def remove_accents(input_str):
    if pd.isna(input_str) or input_str is None:
        return ""
    s = str(input_str).strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = s.replace('đ', 'd')
    return s

def main():
    current_time_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    print("--- CHƯƠNG TRÌNH XỬ LÝ DỮ LIỆU QUÉT QR CCCD ---")
    
    target_file = input("1. Kéo thả hoặc dán đường dẫn file Excel danh sách gốc vào đây:\n>> ").strip().strip('"').strip("'").strip()
    while not os.path.isfile(target_file):
        print("Lỗi: Không tìm thấy file danh sách gốc. Vui lòng kiểm tra lại đường dẫn.")
        target_file = input(">> ").strip().strip('"').strip("'").strip()
        
    export_path = input("\n2. Kéo thả hoặc dán đường dẫn file Excel export (hoặc thư mục chứa các file export) vào đây:\n>> ").strip().strip('"').strip("'").strip()
    while not os.path.exists(export_path):
        print("Lỗi: Không tìm thấy đường dẫn export. Vui lòng kiểm tra lại.")
        export_path = input(">> ").strip().strip('"').strip("'").strip()
        
    print("\nBắt đầu xử lý...")
    
    # 1. Đọc và gộp dữ liệu từ thư mục exports
    if os.path.isdir(export_path):
        export_files = glob.glob(os.path.join(export_path, '*.xlsx'))
    else:
        export_files = [export_path]
        
    if not export_files:
        print(f"Không tìm thấy file excel nào tại '{export_path}'")
        return
    
    exports_df_list = []
    for f in export_files:
        try:
            df = pd.read_excel(f)
            exports_df_list.append(df)
            print(f" - Đã đọc file: {f}")
        except Exception as e:
            print(f"Lỗi đọc file {f}: {e}")
            
    if not exports_df_list:
        return
        
    exports_df = pd.concat(exports_df_list, ignore_index=True)
    
    if 'Họ tên' not in exports_df.columns or 'Ghi chú' not in exports_df.columns:
        print("Lỗi: Các file trong exports cần có cột 'Họ tên' và 'Ghi chú'")
        return
        
    # Tạo mapping name -> data
    name_to_data = {}
    for _, row in exports_df.iterrows():
        name_val = row['Họ tên']
        note_val = row.get('Ghi chú')
        cccd_val = row.get('CCCD')
        cmnd_val = row.get('CMND')
        
        name = remove_accents(name_val)
        note = str(note_val).strip() if pd.notna(note_val) else ""
        note_unaccented = remove_accents(note)
        cccd = str(cccd_val).strip() if pd.notna(cccd_val) else ""
        cmnd = str(cmnd_val).strip() if pd.notna(cmnd_val) else ""
        
        if not name:
            continue
            
        current_data = name_to_data.get(name, {})
        current_status = current_data.get("status")
        
        new_status = current_status
        # Ưu tiên "Đọc mã QR" nếu có nhiều kết quả trùng tên
        if "doc ma qr" in note_unaccented:
            new_status = "Đọc mã QR"
        elif "qr khong doc duoc" in note_unaccented:
            if current_status != "Đọc mã QR":
                new_status = "QR không đọc được"
        else:
            # Nếu có tên trong file export nhưng không có 2 ghi chú trên, vẫn đánh dấu là Chưa đầy đủ
            if not current_status:
                new_status = "QR không đọc được"
                
        # Cập nhật CCCD, CMND (Ưu tiên lấy nếu chưa có, hoặc nếu mới chuyển lên trạng thái "Đọc mã QR")
        new_cccd = current_data.get("cccd", "")
        if cccd and cccd.lower() not in ['nan', 'none']:
            if not new_cccd or new_status == "Đọc mã QR":
                new_cccd = cccd
                
        new_cmnd = current_data.get("cmnd", "")
        if cmnd and cmnd.lower() not in ['nan', 'none']:
            if not new_cmnd or new_status == "Đọc mã QR":
                new_cmnd = cmnd
                
        name_to_data[name] = {
            "original_name": name_val,
            "status": new_status,
            "cccd": new_cccd,
            "cmnd": new_cmnd
        }
                
    print(f"Đã tải thông tin {len(name_to_data)} người duy nhất từ thư mục exports.")
    
    # 2. Xử lý file đích bằng openpyxl để giữ nguyên định dạng
    if not os.path.exists(target_file):
        print(f"Lỗi: Không tìm thấy file đích '{target_file}'")
        return
        
    try:
        print(f"Đang đọc file đích: {target_file}")
        wb = openpyxl.load_workbook(target_file)
        ws = wb.active
        ws.title = "Tổng quan"
    except Exception as e:
        print(f"Lỗi khi đọc file '{target_file}': {e}")
        return
        
    # Tìm dòng chứa tiêu đề
    header_row = None
    headers = {}
    for r in range(1, min(20, ws.max_row + 1)):
        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[r], start=1) if cell.value}
        if 'HỌ VÀ TÊN' in headers:
            header_row = r
            break
            
    if not header_row:
        print("Lỗi: Không tìm thấy dòng tiêu đề có chứa 'HỌ VÀ TÊN' trong 20 dòng đầu")
        return
    
    col_name = headers.get('HỌ VÀ TÊN')
    col_name_no_accent = headers.get('HỌ VÀ TÊN KHÔNG DẤU')
    col_cccd = headers.get('CCCD')
    col_cmnd = headers.get('CMND')
    col_status = headers.get('TRẠNG THÁI')
    
    # Tự động chèn các cột còn thiếu
    cols_to_insert = []
    if not col_name_no_accent: cols_to_insert.append('HỌ VÀ TÊN KHÔNG DẤU')
    if not col_cccd: cols_to_insert.append('CCCD')
    if not col_cmnd: cols_to_insert.append('CMND')
    if not col_status: cols_to_insert.append('TRẠNG THÁI')
    
    if cols_to_insert:
        print(f"Tự động bổ sung các cột {cols_to_insert} vào file Excel vì chưa tìm thấy.")
        ws.insert_cols(col_name + 1, len(cols_to_insert))
        for idx, c_name in enumerate(cols_to_insert):
            cell = ws.cell(row=header_row, column=col_name + 1 + idx)
            cell.value = c_name
            cell.font = Font(bold=True)
            
        # Cập nhật lại headers vì đã chèn cột
        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[header_row], start=1) if cell.value}
        
        col_name_no_accent = headers.get('HỌ VÀ TÊN KHÔNG DẤU')
        col_cccd = headers.get('CCCD')
        col_cmnd = headers.get('CMND')
        col_status = headers.get('TRẠNG THÁI')
    
    if not col_name:
        print("Lỗi: File đích không có cột 'HỌ VÀ TÊN'")
        return
        
    print("Đang kiểm tra và bổ sung tên mới từ exports...")
    existing_names = set()
    col_stt = headers.get('STT')
    
    real_max_row = header_row
    last_stt_val = 0
    
    for row in range(header_row + 1, ws.max_row + 1):
        target_name_cell = ws.cell(row=row, column=col_name).value
        if target_name_cell:
            existing_names.add(remove_accents(target_name_cell))
            real_max_row = row
            if col_stt:
                stt_val = ws.cell(row=row, column=col_stt).value
                try:
                    if stt_val is not None:
                        last_stt_val = int(stt_val)
                except ValueError:
                    pass
            
    red_font = Font(color="FF0000")
    added_count = 0
    added_names = set()
    current_append_row = real_max_row + 1
    for name, data in name_to_data.items():
        if name not in existing_names:
            added_names.add(name)
            cell = ws.cell(row=current_append_row, column=col_name)
            cell.value = data.get('original_name', name)
            
            if col_stt:
                last_stt_val += 1
                ws.cell(row=current_append_row, column=col_stt).value = last_stt_val
                
            current_append_row += 1
            added_count += 1
            
    if added_count > 0:
        print(f"Đã bổ sung {added_count} người mới vào danh sách gốc (chữ đỏ).")

    # Khởi tạo danh sách chứa dữ liệu
    list_full = []
    list_partial = []
    list_none = []
    
    # Khởi tạo các màu
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    
    last_col = col_status if col_status else (max(headers.values()) if headers else ws.max_column)
    
    print("Đang đối chiếu dữ liệu...")
    matched_names = set()
    for row in range(header_row + 1, ws.max_row + 1):
        target_name_cell = ws.cell(row=row, column=col_name).value
        if not target_name_cell:
            continue
            
        unaccented_target = remove_accents(target_name_cell)
        
        # Bổ sung Họ tên không dấu (In hoa)
        if col_name_no_accent:
            ws.cell(row=row, column=col_name_no_accent).value = unaccented_target.upper()
            
        data = name_to_data.get(unaccented_target, {})
        if data:
            matched_names.add(unaccented_target)
            
        status = data.get("status")
        
        c_cccd = data.get("cccd", "")
        c_cmnd = data.get("cmnd", "")
        
        # Bổ sung CCCD, CMND nếu tìm thấy
        if col_cccd and c_cccd:
            ws.cell(row=row, column=col_cccd).value = c_cccd
        if col_cmnd and c_cmnd:
            ws.cell(row=row, column=col_cmnd).value = c_cmnd
            
        is_added = unaccented_target in added_names
        person_data = [
            target_name_cell,
            unaccented_target.upper(),
            c_cccd,
            c_cmnd,
            is_added
        ]
        
        if status == "Đọc mã QR":
            ws.cell(row=row, column=col_status).value = 'Thông tin đầy đủ'
            list_full.append(person_data)
            # Xóa màu của dòng
            for col_idx in range(1, last_col + 1):
                ws.cell(row=row, column=col_idx).fill = no_fill
        elif status == "QR không đọc được":
            ws.cell(row=row, column=col_status).value = 'Thông tin chưa đầy đủ'
            list_partial.append(person_data)
            # Tô màu vàng cho dòng
            for col_idx in range(1, last_col + 1):
                ws.cell(row=row, column=col_idx).fill = yellow_fill
        else:
            ws.cell(row=row, column=col_status).value = 'Chưa có thông tin'
            list_none.append(person_data)
            # Tô màu hồng cho dòng
            for col_idx in range(1, last_col + 1):
                ws.cell(row=row, column=col_idx).fill = pink_fill
                
        # Bôi đỏ tất cả các cột nếu là dòng mới thêm
        if is_added:
            for col_idx in range(1, last_col + 1):
                ws.cell(row=row, column=col_idx).font = red_font
                
    # Xoá các sheet cũ nếu có để tránh lỗi
    for s_name in ['Thống kê', 'Thông tin đầy đủ', 'Thông tin chưa đầy đủ', 'Chưa có thông tin']:
        if s_name in wb.sheetnames:
            del wb[s_name]
            
    # Tạo sheet thống kê
    print("Đang tạo các sheet phân loại...")
    ws_stat = wb.create_sheet('Thống kê')
    
    ws_stat.append([f'BẢNG THỐNG KÊ CHI TIẾT ({current_time_str})'])
    ws_stat.append([])
    
    total_target = len(list_full) + len(list_partial) + len(list_none)
    total_exports = len(name_to_data)
    unmatched_exports = total_exports - len(matched_names)
    
    ws_stat.append(['CHỈ TIÊU THỐNG KÊ', 'SỐ LƯỢNG', 'TỶ LỆ (%)'])
    ws_stat.append(['Tổng số người trong file danh sách', total_target, '100%'])
    
    pct_full = f"{(len(list_full)/total_target*100):.1f}%" if total_target > 0 else "0%"
    pct_partial = f"{(len(list_partial)/total_target*100):.1f}%" if total_target > 0 else "0%"
    pct_none = f"{(len(list_none)/total_target*100):.1f}%" if total_target > 0 else "0%"
    
    ws_stat.append(['- Đã có thông tin (Đầy đủ)', len(list_full), pct_full])
    ws_stat.append(['- Đã có thông tin (Chưa đầy đủ - OCR)', len(list_partial), pct_partial])
    ws_stat.append(['- Chưa có thông tin', len(list_none), pct_none])
    
    ws_stat.append([])
    ws_stat.append(['THỐNG KÊ ĐỐI SOÁT', 'SỐ LƯỢNG', 'GHI CHÚ'])
    ws_stat.append(['Tổng số dữ liệu thu thập được (Từ exports)', total_exports, 'Số người quét được thực tế'])
    ws_stat.append(['- Đã khớp với file danh sách', len(matched_names), ''])
    ws_stat.append(['- Dữ liệu dư (Khách vãng lai/Không có tên)', unmatched_exports, 'Có quét nhưng không có tên trong file danh sách'])
    
    ws_stat.append([])
    ws_stat.append(['GHI CHÚ MÀU SẮC TRONG TẤT CẢ CÁC SHEET', 'Ý NGHĨA', ''])
    ws_stat.append(['Chữ màu đỏ', 'Người chưa có tên trong danh sách gốc, được tự động chèn thêm từ dữ liệu máy quét', ''])
    ws_stat.append(['Nền màu trắng (Không tô màu)', 'Đã quét mã QR thành công (Thông tin đầy đủ)', ''])
    ws_stat.append(['Nền tô màu vàng', 'Đã quét nhưng không đọc được mã QR (Thông tin chưa đầy đủ)', ''])
    ws_stat.append(['Nền tô màu hồng', 'Chưa có thông tin quét (Chưa đến hoặc chưa quét được)', ''])
    
    ws_stat.append([])
    ws_stat.append(['GHI CHÚ Ý NGHĨA CÁC SHEET (TRANG TÍNH)', 'Ý NGHĨA', ''])
    ws_stat.append(['Sheet "Tổng quan"', 'Danh sách tổng hợp toàn bộ dữ liệu gốc và dữ liệu đã quét', ''])
    ws_stat.append(['Sheet "Thông tin đầy đủ"', 'Danh sách rút trích những người đã quét mã QR thành công (Nền trắng)', ''])
    ws_stat.append(['Sheet "Thông tin chưa đầy đủ"', 'Danh sách rút trích những người quét lỗi mã QR hoặc cần kiểm tra lại (Nền vàng)', ''])
    ws_stat.append(['Sheet "Chưa có thông tin"', 'Danh sách những người có trong danh sách gốc nhưng chưa có thông tin quét (Nền hồng)', ''])
    
    # Định dạng sheet thống kê
    from openpyxl.styles import Alignment
    
    ws_stat.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    ws_stat.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_stat.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in ws_stat[3]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for cell in ws_stat[9]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for cell in ws_stat[14]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for cell in ws_stat[20]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    ws_stat.column_dimensions['A'].width = 50
    ws_stat.column_dimensions['B'].width = 15
    ws_stat.column_dimensions['C'].width = 30
        
    def write_list_sheet(sheet_name, data_list):
        ws_list = wb.create_sheet(sheet_name)
        
        ws_list.append([f'DANH SÁCH {sheet_name.upper()} ({current_time_str})'])
        ws_list.append([])
        
        headers_list = ['STT', 'HỌ VÀ TÊN', 'HỌ VÀ TÊN KHÔNG DẤU', 'CCCD', 'CMND']
        ws_list.append(headers_list)
        for cell in ws_list[3]:
            cell.font = Font(bold=True)
            
        red_font = Font(color="FF0000")
        for idx, row_data in enumerate(data_list, start=1):
            is_added = row_data[-1]
            actual_data = row_data[:-1]
            ws_list.append([idx] + actual_data)
            
            if is_added:
                for col_idx in range(1, len(actual_data) + 2):
                    ws_list.cell(row=idx + 3, column=col_idx).font = red_font
                    
        ws_list.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws_list.cell(row=1, column=1).font = Font(bold=True, size=14)
        from openpyxl.styles import Alignment
        ws_list.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
        # Tự động chỉnh độ rộng cột
        for col in ws_list.columns:
            max_length = 0
            # Dùng col[-1] vì dòng đầu tiên bị merge nên col[0] có thể là MergedCell không có column_letter
            column = col[-1].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_list.column_dimensions[column].width = adjusted_width
            
    write_list_sheet('Thông tin đầy đủ', list_full)
    write_list_sheet('Thông tin chưa đầy đủ', list_partial)
    write_list_sheet('Chưa có thông tin', list_none)
            
    # 3. Lưu file
    try:
        print("Đang lưu file đích...")
        wb.save(target_file)
        print(f"Hoàn thành! Đã lưu kết quả vào '{target_file}'.")
        print(f"\n--- THỐNG KÊ ---")
        print(f"Tổng số dòng đã xử lý: {len(list_full) + len(list_partial) + len(list_none)}")
        print(f" - Thông tin đầy đủ: {len(list_full)}")
        print(f" - Thông tin chưa đầy đủ: {len(list_partial)}")
        print(f" - Chưa có thông tin: {len(list_none)}")
    except Exception as e:
        print(f"Lỗi khi lưu file '{target_file}'. Vui lòng đảm bảo file không bị mở ở chương trình khác. Lỗi: {e}")

if __name__ == "__main__":
    main()
