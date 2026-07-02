import os
import warnings

# Giới hạn số luồng của OpenMP/MKL để tránh thread explosion khi chạy ThreadPoolExecutor
# PHẢI ĐẶT TRƯỚC KHI IMPORT TORCH, CV2
os.environ["OMP_NUM_THREADS"] = "1"
os.environ["OPENBLAS_NUM_THREADS"] = "1"
os.environ["MKL_NUM_THREADS"] = "1"
os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
os.environ["NUMEXPR_NUM_THREADS"] = "1"

warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

import PIL.Image
if not hasattr(PIL.Image, 'ANTIALIAS'):
    PIL.Image.ANTIALIAS = PIL.Image.LANCZOS
import sys
import logging
logging.basicConfig(filename='debug_flow.log', level=logging.INFO, 
                    format='%(asctime)s - [%(threadName)s] - %(message)s')
def LOG(msg):
    logging.info(msg)

import glob
import tempfile
import uuid
import cv2
cv2.setNumThreads(1)

from pyzbar.pyzbar import decode
from pyzbar.pyzbar import ZBarSymbol
import zxingcpp
import pytesseract
import openpyxl
from openpyxl.styles import Font, PatternFill
import requests
import datetime
import pillow_heif
import numpy as np
from PIL import Image

import torch
if torch.cuda.is_available():
    try:
        torch.cuda.set_per_process_memory_fraction(0.6)
    except Exception: pass

IN_COLAB = 'COLAB_RELEASE_TAG' in os.environ
REFRESH_RATE = 0.00833 if IN_COLAB else 10

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from vietocr_engine import extract_text_from_image
import re
import json
import warnings
import unicodedata
import concurrent.futures
import zipfile
import threading
import difflib

# Caches cho Rule-based & Fuzzy Matching
OCR_RULES_CACHE = {}
OCR_RULES_MTIME = 0
VALID_LOCATIONS_CACHE = []

def load_ocr_rules():
    global OCR_RULES_CACHE, OCR_RULES_MTIME
    rules_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'ocr_rules.json')
    try:
        if os.path.exists(rules_path):
            mtime = os.path.getmtime(rules_path)
            if mtime > OCR_RULES_MTIME:
                with open(rules_path, 'r', encoding='utf-8') as f:
                    OCR_RULES_CACHE = json.load(f)
                OCR_RULES_MTIME = mtime
    except Exception as e:
        print(f"[Warning] Failed to load ocr_rules.json (giữ cache cũ): {e}")
    return OCR_RULES_CACHE

def load_valid_locations():
    global VALID_LOCATIONS_CACHE
    if VALID_LOCATIONS_CACHE:
        return VALID_LOCATIONS_CACHE
    data_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'data.json')
    try:
        if os.path.exists(data_path):
            with open(data_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                locations = set()
                for level1 in data:
                    locations.add(level1.get('name', ''))
                    clean_l1 = re.sub(r'^(Tỉnh|Thành phố)\s+', '', level1.get('name', ''))
                    locations.add(clean_l1)
                    for level2 in level1.get('level2s', []):
                        locations.add(level2.get('name', ''))
                        clean_l2 = re.sub(r'^(Quận|Huyện|Thị xã|Thành phố)\s+', '', level2.get('name', ''))
                        locations.add(clean_l2)
                        for level3 in level2.get('level3s', []):
                            locations.add(level3.get('name', ''))
                            clean_l3 = re.sub(r'^(Phường|Xã|Thị trấn)\s+', '', level3.get('name', ''))
                            locations.add(clean_l3)
                VALID_LOCATIONS_CACHE = [x for x in locations if x]
    except Exception as e:
        print(f"[Warning] Failed to load valid locations from data.json: {e}")
    return VALID_LOCATIONS_CACHE

# Tắt cảnh báo chia cho 0 của numpy bên trong thư viện VietOCR
warnings.filterwarnings("ignore", category=RuntimeWarning, message="invalid value encountered in divide")

DEBUG_MODE = '--debug' in sys.argv

# Danh sách 300 họ phổ biến tiếng Việt (không dấu, lowercase)
_VN_SURNAMES = {
    'a', 'an', 'au', 'ba', 'bach', 'ban', 'bang', 'banh', 'bao', 'be',
    'bien', 'bo', 'bui', 'ca', 'cai', 'cam', 'can', 'cao', 'cap', 'cat',
    'chao', 'chau', 'che', 'chi', 'chiem', 'chiu', 'chu', 'chuc', 'chung', 'chuong',
    'co', 'cong', 'cu', 'cung', 'cut', 'dai', 'dam', 'dan', 'dang', 'danh',
    'dao', 'dau', 'deo', 'diec', 'diem', 'dien', 'diep', 'dieu', 'dinh', 'do',
    'doan', 'doi', 'dong', 'du', 'duong', 'duy', 'gian', 'giang', 'giap', 'ha',
    'hac', 'han', 'hang', 'hau', 'ho', 'hoa', 'hoang', 'hong', 'hua', 'hung',
    'huong', 'huynh', 'ka', 'kha', 'khau', 'khieu', 'khong', 'khuat', 'khuc', 'khuong',
    'khuu', 'kien', 'kieu', 'kim', 'la', 'lac', 'lai', 'lam', 'lang', 'lanh',
    'lao', 'lau', 'le', 'leng', 'leu', 'lien', 'lieng', 'lieu', 'linh', 'lo',
    'loc', 'loi', 'long', 'lu', 'luan', 'luc', 'luong', 'luu', 'luyen', 'ly',
    'ma', 'mac', 'mach', 'mai', 'man', 'mang', 'manh', 'mau', 'me', 'mo',
    'mong', 'moong', 'mua', 'nay', 'ngac', 'ngan', 'nghiem', 'ngo', 'ngoc', 'ngon',
    'ngu', 'nguy', 'nguyen', 'nham', 'nhan', 'nhao', 'nhu', 'nie', 'ninh', 'nong', 'o',
    'on', 'ong', 'pham', 'phan', 'phi', 'pho', 'phong', 'phu', 'phung', 'phuong', 'ngoi',
    'quach', 'quan', 'quang', 'que', 'quyen', 'ro', 'sa', 'sai', 'sam', 'san',
    'son', 'su', 'sung', 'ta', 'tac', 'tan', 'tang', 'tao', 'tat', 'thach',
    'thai', 'tham', 'than', 'thang', 'thanh', 'thao', 'thi', 'thieu', 'tho', 'thoi',
    'thuong', 'thuy', 'tien', 'tiet', 'tieu', 'to', 'toan', 'ton', 'tong', 'tonnu',
    'tonthat', 'tra', 'trac', 'tram', 'tran', 'trang', 'tri', 'trieu', 'trinh', 'trung',
    'truong', 'tu', 'tuong', 'ung', 'uong', 'va', 'van', 'vang', 'vi', 'vien',
    'vinh', 'vo', 'vong', 'vu', 'vuong', 'vuu', 'vy', 'xa', 'xong', 'y',
    'yen',
}

_VN_NAME_SYLLABLES = set("AN ANH BA BAC BACH BAN BANG BAO BE BEN BICH BINH BO BON BUI CA CAM CAN CANH CAO CAT CHAU CHI CHIEN CHINH CHU CHUC CHUAN CHUNG CHUYEN CON CU CUC CUONG DA DAI DAN DANG DAO DAT DAU DE DIEM DIEN DIEP DIEU DINH DO DOAN DOANH DONG DU DUC DUNG DUONG DUY DUYEN EM GIA GIAC GIANG GIAO GIAP HA HAI HAN HANG HANH HAO HE HIEN HIEP HIEU HINH HO HOA HOAI HOAN HOANG HOI HONG HOP HUNG HUONG HUU HUY HUYEN HUYNH ICH KHA KHAI KHANG KHANH KHAO KHE KHOA KHOI KHONG KHUAT KHUE KHUU KHUYEN KIEN KIEU KIM KY LA LAI LAM LAN LANG LANH LAP LE LIEN LIEU LINH LO LOAN LOC LOI LONG LU LUA LUAN LUC LUONG LUU LY MAC MACH MAI MAN MANG MANH MAO MAU MINH MOC MONG MUOI MY NAM NGA NGAN NGANH NGHI NGHIA NGHIEM NGO NGOC NGOI NGON NGU NGUYEN NGUYET NHA NHAN NHAT NHI NHIEN NHO NHU NHUAN NHUNG NIEN NINH NOAN NONG NU NUOI NUONG OA OANH ONG PHA PHAI PHAM PHAN PHANG PHAT PHI PHIEN PHONG PHU PHUC PHUNG PHUONG QUACH QUAN QUANG QUE QUOC QUY QUYEN QUYET QUYNH RAO SA SAM SAN SANG SAU SEN SINH SOA SON SONG SUONG SY TA TAI TAM TAN TANG TANH TAO TAY THA THACH THAI THAM THAN THANG THANH THAO THAT THAY THE THI THIEN THIET THIEU THINH THOA THOAI THOM THU THUAN THUC THUONG THUY THUYEN THY TIEN TIEP TIN TINH TO TOA TOAI TOAN TON TONG TRA TRAM TRAN TRANG TRANH TRAO TRI TRIEU TRINH TRONG TRU TRUC TRUNG TRUONG TRUYEN TU TUAN TUAT TUE TUI TUNG TUY TUYEN TUYET UNG UYEN VAN VANG VI VIEN VIET VINH VO VONG VU VUONG VY XINH XUA XUAN Y YEN".lower().split()).union(_VN_SURNAMES)

_VN_PROVINCE_CODES = {
    '001', '002', '004', '006', '008', '010', '011', '012', '014', '015',
    '017', '019', '020', '022', '024', '025', '026', '027', '030', '031',
    '033', '034', '035', '036', '037', '038', '040', '042', '044', '045',
    '046', '048', '049', '051', '052', '054', '056', '058', '060', '062',
    '064', '066', '067', '068', '070', '072', '074', '075', '077', '079',
    '080', '082', '083', '084', '086', '087', '089', '091', '092', '093',
    '094', '095', '096'
}

def _is_valid_name(s):
    """Tên hợp lệ: 2-5 từ, bắt đầu bằng họ VN phổ biến, không số."""
    if not s or re.search(r'\d', s):
        return False
    words = s.strip().split()
    if not (2 <= len(words) <= 5):
        return False
    # Chuẩn hóa họ: xóa dấu để so sánh
    first = unicodedata.normalize('NFD', words[0].lower())
    first_ascii = ''.join(c for c in first if unicodedata.category(c) != 'Mn')
    first_ascii = first_ascii.replace('đ', 'd')
    return first_ascii in _VN_SURNAMES

# Danh sách 63 tỉnh/thành phố VN (không dấu, lowercase)
_VN_PROVINCES = {
    'an giang', 'ba ria-vung tau','ba ria - vung tau','ba ria', 'vung tau', 'bac giang', 'bac kan', 'bac lieu', 'bac ninh', 'ben tre', 'binh dinh', 'binh duong', 'binh phuoc',
    'binh thuan', 'ca mau', 'can tho', 'cao bang', 'da nang', 'dak lak', 'dak nong', 'dien bien', 'dong nai', 'dong thap',
    'gia lai', 'ha giang', 'ha nam', 'ha noi', 'ha tinh', 'hai duong', 'hai phong', 'hau giang', 'hcm', 'ho chi minh',
    'hoa binh', 'hung yen', 'khanh hoa', 'kien giang', 'kon tum', 'lai chau', 'lam dong', 'lang son', 'lao cai', 'long an',
    'nam dinh', 'nghe an', 'ninh binh', 'ninh thuan', 'phu tho', 'phu yen', 'quang binh', 'quang nam', 'quang ngai', 'quang ninh',
    'quang tri', 'soc trang', 'son la', 'tay ninh', 'thai binh', 'thai nguyen', 'thanh hoa', 'thua thien hue', 'tien giang', 'tra vinh',
    'tuyen quang', 'vinh long', 'vinh phuc', 'vung tau', 'yen bai',
}

# Global locks
ocr_lock = threading.Lock()
qr_lock = threading.Lock()

# Global AI Models
detector = None

def init_models():
    global detector
    model_paths = [
        'models/detect.prototxt', 'models/detect.caffemodel',
        'models/sr.prototxt', 'models/sr.caffemodel'
    ]
    base_dir = os.path.dirname(os.path.abspath(__file__))
    abs_paths = [os.path.join(base_dir, p) for p in model_paths]
    
    if all(os.path.exists(p) for p in abs_paths):
        try:
            detector = cv2.wechat_qrcode_WeChatQRCode(*abs_paths)
            print("Đã tải thành công mô hình WeChat QRCode siêu cấp.")
        except Exception as e:
            print(f"Lỗi khởi tạo WeChat QRCode: {e}")
    else:
        print("Cảnh báo: Không tìm thấy file model WeChat QRCode. Khả năng quét ảnh mờ sẽ bị giảm.")
def clean_path(path_str):
    if not path_str:
        return path_str
    path_str = path_str.strip().strip('\'"')
    if os.name != 'nt':
        # Replace backslash-escaped characters commonly produced by terminal drag-and-drop on macOS/Linux
        path_str = path_str.replace('\\ ', ' ')
        # path_str = path_str.replace('\\(', '(')
        # path_str = path_str.replace('\\)', ')')
        # path_str = path_str.replace('\\&', '&')
        # path_str = path_str.replace('\\;', ';')
    return path_str

def format_date(date_str):
    if not date_str or len(date_str) != 8:
        return date_str
    return f"{date_str[0:2]}/{date_str[2:4]}/{date_str[4:8]}"

def get_place_of_issue(qr_data):
    if not qr_data:
        return ""
    fields = qr_data.split('|')
    if len(fields) == 7:
        return "CỤC TRƯỞNG CỤC CẢNH SÁT QUẢN LÝ HÀNH CHÍNH VỀ TRẬT TỰ XÃ HỘI"
    elif len(fields) >= 10:
        return "BỘ CÔNG AN"
    return "CỤC TRƯỞNG CỤC CẢNH SÁT QUẢN LÝ HÀNH CHÍNH VỀ TRẬT TỰ XÃ HỘI"

def get_card_type(qr_data):
    if not qr_data:
        return ""
    fields = qr_data.split('|')
    if len(fields) == 7:
        return "Căn cước công dân"
    elif len(fields) >= 10:
        return "Căn cước"
    return "Không xác định"

def calculate_expiry_date(dob_str):
    if not dob_str or len(dob_str) != 10:
        return ""
    try:
        day, month, year = dob_str.split('/')
        year = int(year)
        current_year = datetime.datetime.now().year
        for age in [14, 25, 40, 60]:
            expiry_year = year + age
            if expiry_year > current_year:
                return f"{day}/{month}/{expiry_year}"
        return "Không thời hạn"
    except:
        return ""

def extract_qr_data(image_path):
    try:
        # Read the image
        if image_path.lower().endswith('.heic'):
            heif_file = pillow_heif.read_heif(image_path)
            img_pil = Image.frombytes(
                heif_file.mode, 
                heif_file.size, 
                heif_file.data,
                "raw",
                heif_file.mode,
                heif_file.stride,
            )
            # Convert to numpy array and BGR for OpenCV
            img = cv2.cvtColor(np.array(img_pil), cv2.COLOR_RGB2BGR)
        else:
            img = cv2.imread(image_path)
            
        if img is None:
            return None, None, "Lỗi đọc file ảnh", None

        def _try_scan(original_scan_img):
            # Cắt giảm kích thước nếu ảnh quá lớn để tránh treo (hang) vô thời hạn ở thư viện C++
            h, w = original_scan_img.shape[:2]
            max_dim = 1500.0
            scale = 1.0
            if max(h, w) > max_dim:
                scale = max_dim / max(h, w)
                scan_img = cv2.resize(original_scan_img, (int(w * scale), int(h * scale)), interpolation=cv2.INTER_AREA)
            else:
                scan_img = original_scan_img

            # 1. zxingcpp
            try:
                res = zxingcpp.read_barcode(scan_img)
                if res and res.text:
                    if not re.search(r'[^\x00-\x7FÀ-ỹ\s\|\:\-/\.]', res.text):
                        pos = res.position
                        cx = (pos.top_left.x + pos.bottom_right.x) / 2
                        cy = (pos.top_left.y + pos.bottom_right.y) / 2
                        return res.text, "zxing-cpp", (cx / scale, cy / scale)
            except Exception:
                pass

            # 2. pyzbar
            decoded_objects = decode(scan_img, symbols=[ZBarSymbol.QRCODE])
            if not decoded_objects:
                gray = cv2.cvtColor(scan_img, cv2.COLOR_BGR2GRAY)
                decoded_objects = decode(gray, symbols=[ZBarSymbol.QRCODE])
                if not decoded_objects:
                    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
                    decoded_objects = decode(thresh, symbols=[ZBarSymbol.QRCODE])

            if decoded_objects:
                obj = decoded_objects[0]
                txt = obj.data.decode('utf-8')
                if '|' in txt and len(txt.split('|')) >= 6:
                    if not re.search(r'[^\x00-\x7FÀ-ỹ\s\|\:\-/\.]', txt):
                        rect = obj.rect
                        cx = rect.left + rect.width / 2
                        cy = rect.top + rect.height / 2
                        return txt, "pyzbar", (cx / scale, cy / scale)

            # 3. wechat_qrcode
            global detector
            if detector:
                try:
                    res, points = detector.detectAndDecode(scan_img)
                    if res and len(res) > 0:
                        txt = res[0]
                        if '|' in txt and len(txt.split('|')) >= 6:
                            if not re.search(r'[^\x00-\x7FÀ-ỹ\s\|\:\-/\.]', txt):
                                pts = points[0]
                                cx = sum(p[0] for p in pts) / 4
                                cy = sum(p[1] for p in pts) / 4
                                return txt, "WeChat QRCode", (cx / scale, cy / scale)
                except Exception:
                    pass
            
            return None, None, None

        # 1. Quét toàn bộ ảnh
        with qr_lock:
            qr_data, engine, center = _try_scan(img)
        
        # 2. Quét góc phần tư phía trên bên phải (CCCD) nếu toàn bộ ảnh thất bại
        if not qr_data:
            h, w = img.shape[:2]
            crop = img[0:int(h/2), int(w/2):w]
            with qr_lock:
                qr_data, engine, crop_center = _try_scan(crop)
            if qr_data and crop_center:
                center = (crop_center[0] + w/2, crop_center[1])

        rotated_img = None
        if qr_data and center:
            cx, cy = center
            h, w = img.shape[:2]
            
            # Lấy vector từ tâm ảnh đến QR
            dx = cx - w/2
            dy = cy - h/2
            
            # Chỉ xoay nếu QR nằm lệch rõ ràng khỏi tâm (tránh nhiễu)
            if max(abs(dx), abs(dy)) > min(w, h) * 0.05:
                if abs(dx) > abs(dy):
                    # QR nằm thiên về hai bên trái/phải -> Thẻ đang nằm ngang
                    if dx < 0:
                        # QR bên trái -> Thẻ bị lộn ngược 180 độ
                        rotated_img = cv2.rotate(img, cv2.ROTATE_180)
                    # Nếu dx > 0: QR bên phải -> Thẻ đúng chiều, không cần xoay
                else:
                    # QR nằm thiên về trên/dưới -> Thẻ đang nằm dọc
                    if dy < 0:
                        # QR ở trên -> Xoay 90 độ cùng chiều kim đồng hồ
                        rotated_img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
                    else:
                        # QR ở dưới -> Xoay 90 độ ngược chiều kim đồng hồ
                        rotated_img = cv2.rotate(img, cv2.ROTATE_90_COUNTERCLOCKWISE)

        if qr_data:
            return qr_data, engine, None, img, rotated_img
        else:
            return None, None, "QR không đọc được", img, None

    except Exception as e:
        return None, None, f"Lỗi xử lý ảnh: {str(e)}", None, None


def parse_ocr_text(text):
                rules = load_ocr_rules()
                data = {
                    'CCCD': '', 'CMND': '', 'Họ tên': '', 'Ngày sinh': '',
                    'Giới tính': '', 'Nơi thường trú gốc': '', 'Ngày cấp CCCD': '',
                    'OCR Side': '', 'Raw Text Upper': text.upper() if text else '',
                    'Raw Text': text if text else ''
                }
                layer1_gender = ''
                layer2_gender = ''

                if not text.strip():
                    return data

                text_upper = text.upper()
                # Sửa lỗi kinh điển OCR trên CCCD: IDVNM0 bị đọc thành ADDNN1, I0VNM, 1DVNM
                text_upper = re.sub(r'\bADDNN1', 'IDVNM0', text_upper)
                text_upper = re.sub(r'\b[I1L]D?VNM[O0]?', 'IDVNM0', text_upper)

                # ---------------------------------------------------------
                # 1. NHẬN DIỆN MẶT THẺ (FRONT / BACK)
                if "<<" in text_upper or "IDVNM" in text_upper or "VNM" in text_upper or "ĐẶC ĐIỂM NHẬN DẠNG" in text_upper or "NGÓN TRỎ" in text_upper or "CỤC TRƯỞNG" in text_upper or "BỘ CÔNG AN" in text_upper:
                    data['OCR Side'] = 'Back'
                # Các từ khóa đặc trưng của Mặt Trước
                elif "CĂN CƯỚC" in text_upper or "CẦN CƯỚC" in text_upper or "CÔNG DÂN" in text_upper or "ĐỘC LẬP" in text_upper or "TỰ DO" in text_upper or "HỌ VÀ TÊN" in text_upper:
                    data['OCR Side'] = 'Front'

                # ---------------------------------------------------------
                # 2. TRÍCH XUẤT SỐ CCCD
                # ---------------------------------------------------------
                text_mrz = text_upper.replace('O', '0') 
                
                # Quy tắc vàng: 12 chữ số liên tiếp nằm ngay liền trước dấu '<'
                # Duyệt qua TẤT CẢ vị trí match, ưu tiên lấy cái bắt đầu bằng 0
                for mrz_m in re.finditer(r'(\d{12})<', text_mrz):
                    candidate = mrz_m.group(1)
                    if candidate.startswith('0'):
                        data['CCCD'] = candidate
                        data['OCR Side'] = 'Back'
                        break

                # Nếu không tìm thấy pattern <, thử lắp ráp từ VNM chuẩn ICAO
                if not data['CCCD'] and ("IDVNM" in text_upper or "VNM" in text_upper):
                    mrz_match = re.search(r'VNM(\d{9})\d(\d{3})', text_mrz)
                    if mrz_match:
                        assembled = mrz_match.group(2) + mrz_match.group(1)
                        if assembled.startswith('0') and len(assembled) == 12:
                            data['CCCD'] = assembled
                            data['OCR Side'] = 'Back'

                # Nếu vẫn chưa có: tìm dòng/khối IDVNM, gom ký tự số + '<' từ khối đó
                # Bỏ 3 ký tự cuối (<<X), lấy 12 ký tự cuối còn lại = CCCD
                # Lưu ý: OCR có thể vỡ dòng MRZ ra nhiều dòng nhỏ → gom cả khối 5 dòng tiếp theo
                # Lưu ý: MRZ không có chữ O → mọi 'O' đều là số 0 bị OCR đọc sai
                if not data['CCCD']:
                    mrz_lines = text_mrz.split('\n')  # text_mrz đã replace O→0 rồi
                    for i, line in enumerate(mrz_lines):
                        line_stripped = line.strip()
                        if line_stripped.startswith('IDVN') or line_stripped.startswith('VNM'):
                            # Gom: dòng IDVNN + tối đa 5 dòng tiếp theo (MRZ có thể wrap)
                            block_lines = mrz_lines[i:i+6]
                            block = ' '.join(block_lines)
                            # Bỏ tiền tố IDVNM/IDVNN/VNM
                            after_prefix = re.sub(r'^(I?D?VNM|I?D?VNN)', '', block.strip())
                            # Chỉ giữ lại chữ số và dấu '<' (MRZ không có chữ O)
                            cleaned = re.sub(r'[^0-9<]', '', after_prefix)
                            
                            # Cách 1: Bố cục thẻ gắn chip (ICAO Doc 9303 TD3)
                            # IDVNM + 9 ký tự số giấy tờ + 1 số checksum + 12 số CCCD
                            # Để chống dịch chuyển index (do OCR đọc thiếu/thừa số đầu dòng), quét chuỗi số của dòng đầu tiên
                            # và tìm chuỗi 12 số bắt đầu bằng 0 và có mã tỉnh Việt Nam hợp lệ ở khu vực Optional Data (sau index 10).
                            first_line_clean = re.sub(r'^(I?D?VNM|I?D?VNN)', '', line_stripped)
                            first_line_digits = re.sub(r'[^0-9<]', '', first_line_clean)
                            for offset in range(10, len(first_line_digits) - 11):
                                candidate = first_line_digits[offset:offset+12]
                                if len(candidate) == 12 and candidate.startswith('0') and candidate[:3] in _VN_PROVINCE_CODES:
                                    data['CCCD'] = candidate
                                    data['OCR Side'] = 'Back'
                                    break
                            
                            # Cách 2: Bố cục fallback của thẻ cũ hoặc bị thiếu ký tự (như mất dấu < ở cuối)
                            if not data['CCCD'] and len(cleaned) >= 15:
                                candidate = cleaned[:-3][-12:]  # bỏ 3 cuối, lấy 12 cuối
                                if len(candidate) == 12 and candidate.startswith('0'):
                                    data['CCCD'] = candidate
                                    data['OCR Side'] = 'Back'
                            break

                # ------- MẶT TRƯỚC (Front): không có MRZ, quét số trực tiếp -------
                # Chỉ chạy nếu chưa có CCCD từ MRZ VÀ thẻ không phải mặt sau
                # Tránh tình trạng mặt sau bị OCR rác ra ngẫu nhiên 12 chữ số
                if not data['CCCD'] and data['OCR Side'] != 'Back':
                    text_numbers = text_upper.replace('O', '0')
                    # Tìm tất cả cụm số (có thể chứa khoảng trắng) dài 12-25 ký tự
                    # Regex này lấy luôn các số bị dính vào chữ cái ở đầu/cuối (như TTBB0088195038174)
                    cccd_matches = re.findall(r'(?<!\d)([\d \t\xa0]{12,25})(?!\d)', text_numbers)
                    valid_cccds = []
                    for match_str in cccd_matches:
                        val = re.sub(r'\s+', '', match_str)  # loại mọi whitespace
                        # Có trường hợp OCR đọc dư 1 vài số 0 ở đầu (do chữ 'số' bị đọc thành 's0' -> 0)
                        if 12 <= len(val) <= 15:
                            candidate = val[-12:] # Lấy đúng 12 số cuối cùng
                            if candidate.startswith('0'):
                                valid_cccds.append(candidate)
                    
                    if valid_cccds:
                        # Ưu tiên cụm nằm trên dòng có chứa từ khóa 'SỐ'/'CƯỚC' (nhãn số thẻ)
                        data['CCCD'] = valid_cccds[0]
                        data['OCR Side'] = 'Front'
                        for line in text_upper.split('\n'):
                            if any(kw in line for kw in rules.get("cccd_keywords", [])):
                                m = re.search(r'(?<!\d)([\d \t\xa0]{12,25})(?!\d)', line.replace('O','0'))
                                if m:
                                    val = re.sub(r'\s+', '', m.group(1))
                                    if 12 <= len(val) <= 15:
                                        candidate = val[-12:]
                                        if candidate.startswith('0'):
                                            data['CCCD'] = candidate
                                            data['OCR Side'] = 'Front'
                                            break
                all_dates = re.findall(r'(?<!\d)\d{2}/\d{2}/\d{4}(?!\d)', text)
    
                # ---------------------------------------------------------
                # 3. TRÍCH XUẤT GIỚI TÍNH (Layer 2 - Global Fallback)
                # Lưu ý: Chỉ chạy cho mặt trước. Ưu tiên regex từ độc lập, có word boundary.
                # ---------------------------------------------------------
                text_lower = text.lower()
                if data['OCR Side'] != 'Back':
                    nam_count = len(re.findall(r'\bnam\b', text_lower))
                    vietnam_count = len(re.findall(r'việt nam|viet nam|hà nam|quảng nam|hải nam|đông nam|đồng nam|phương nam|tây nam|nam định', text_lower))
                    if nam_count > vietnam_count:
                        layer2_gender = 'Nam'
                    elif re.search(r'\bn[uưứữ][\s]*\b', text_lower) or re.search(r'\bnữ\b', text_lower):
                        layer2_gender = 'Nữ'
        
                # ---------------------------------------------------------
                # 3b. TRÍCH XUẤT TÊN TỪ MRZ LINE 3 (BACK SIDE)
                # MRZ line 3 có format: SURNAME<<GIVEN<NAMES<<<
                # OCR đọc nhầm: '<' → 'C'/'K', '<<' → 'CK'/'CC', padding → 'ES'/'CECCES'...
                # Ví dụ: TRAN<LE<THAO<<NGUYEN<<< → TRANCKLECTHAOCNGUYENCECCES
                # ---------------------------------------------------------
                if data['OCR Side'] == 'Back' and not data['Họ tên']:
                    def _extract_mrz_name_words(s):
                        """Tách các từ tên thực sự từ chuỗi MRZ bằng thuật toán Greedy Match với từ điển."""
                        # Sửa các lỗi OCR kinh điển làm biến dạng từ
                        ocr_fixes = rules.get("mrz_name_ocr_fixes", {})
                        for bad, good in ocr_fixes.items():
                            s = s.replace(bad, good)
                        # Loại bỏ re.sub(r'[CKEAS<]{3,}$', '', s) vì nó cắt nhầm chữ cuối của tên (VD: DUC -> DU, GIAC -> GIA)
                        # Biểu thức regex findall bên dưới đã tự động bỏ qua các chuỗi rác không có trong từ điển.
                        # Danh sách âm tiết tên Tiếng Việt phổ biến (không dấu)
                        names_list = sorted(list(_VN_NAME_SYLLABLES), key=len, reverse=True)
                        pattern = r'(?i)(' + '|'.join(names_list) + r')'
                        
                        matches = re.findall(pattern, s)
                        
                        # Loại bỏ ảo giác OCR (padding <<<< bị đọc thành từ lặp lại nhiều lần)
                        for i in range(len(matches) - 2):
                            if matches[i] == matches[i+1] == matches[i+2]:
                                matches = matches[:i]
                                break
                        
                        # Nếu vẫn còn lặp 2 lần ở cuối chuỗi (do ảo giác ngắn), cắt bỏ nếu tên đã đủ dài
                        while len(matches) > 3 and matches[-1] == matches[-2]:
                            matches = matches[:-2]
                            
                        return matches

                    for line in text_upper.split('\n'):
                        # OCR đôi khi sinh ra khoảng trắng giữa các chữ trong MRZ -> Xóa toàn bộ khoảng trắng
                        ls = line.replace(' ', '').strip()
                        
                        # Xóa các ký tự rác (số, dấu câu) ở cuối đuôi do OCR đọc nhầm viền
                        ls = re.sub(r'[^A-Z<]+$', '', ls)
                        
                        # Xóa các ký tự số, dấu câu rác ở đầu chuỗi (ví dụ: "36PHAN..." -> "PHAN...")
                        ls = re.sub(r'^[^A-Z]+', '', ls)
                        
                        # Chặn các dòng tiếng Anh hoặc tiếng Việt không dấu bị nhận diện nhầm
                        if any(bad in ls for bad in rules.get("mrz_ignore_words", [])):
                            continue
                            
                        # MRZ Line 3 chứa tên, KHÔNG có số, KHÔNG có dấu tiếng Việt
                        # OCR đọc dấu << thành CK, CC, CE, KCK, KS... và có thể ảo giác lặp lại nên nới lỏng len <= 65
                        if (15 <= len(ls) <= 65
                            and re.match(r'^[A-Z<]+$', ls)  # Chỉ gồm chữ cái A-Z và < (không có số, không dấu)
                            and re.search(r'[<CKES]{2}', ls) # Chắc chắn có ít nhất 1 cụm 2 ký tự độn (CK, CC, CE, KS...) thay cho <<
                            and not ls.startswith('IDVN')
                            and not ls.startswith('VNM')
                        ):
                            # Dọn dẹp padding thực tế
                            ls_clean = re.sub(r'[<]+$', '', ls)
                            
                            # Áp dụng ocr_fixes trước khi split để tránh split nhầm vào trong từ lỗi (VD: NGOCK bị split tại CK)
                            ocr_fixes = rules.get("mrz_name_ocr_fixes", {})
                            for bad, good in ocr_fixes.items():
                                ls_clean = ls_clean.replace(bad, good)
                            
                            # Tách Họ và Đệm+Tên dựa trên 2 dấu << liên tiếp (OCR -> CC, CK, CE, CEC, KK, KS...)
                            # Bỏ EK ra khỏi danh sách cắt vì EK dễ cắt nhầm tên (VD: LEKK -> L và EK)
                            split_parts = re.split(r'CK|CEC|KCK|CC|CE|CS|KK|KS', ls_clean, maxsplit=1)
                            if len(split_parts) == 2:
                                surname_words = _extract_mrz_name_words(split_parts[0])
                                given_words   = _extract_mrz_name_words(split_parts[1])
                                words = surname_words + given_words
                                
                            else:
                                words = _extract_mrz_name_words(split_parts[0])
                            
                            # Xóa các cụm 1 chữ cái bị nhận diện nhầm ở đuôi (thường do C, K, E rớt lại tạo thành CA, KY)
                            # Nhưng giữ lại chữ Y (Thị Y, A Y)
                            while len(words) > 2 and len(words[-1]) == 2 and words[-1] in ['CA', 'KY', 'KE', 'CE', 'CI', 'CO']:
                                words.pop()
                                
                            if len(words) >= 2:
                                data['Họ tên'] = ' '.join(words)
                                break

                # ---------------------------------------------------------
                # 3.5. TRÍCH XUẤT HỌ TÊN VƯỢT DÒNG (Dành cho CCCD đứt nét hoặc SMS screenshot)
                # ---------------------------------------------------------
                raw_name = ""
                name_block_matches = re.findall(r'(?i)(?:h[oọ]\s*(?:v[aà]\s*)?t[eê]n|t[eê]n\s*khai\s*sinh|full\s*name|fui\s*nam|kho\s*v[aà]\s*t[eê]n)\s*[:\s]+(.{1,60}?)[,.]?\s*(?:ng[aà]y\s*sinh|date\s*of\s*birth|ng[aà]y,\s*th[aá]ng|dob|sinh\s*ng[aà]y|\bsinh\b\s*:|gi[oớ]i\s*t[ií]nh|qu[oố]c\s*t[iị]ch|n[oơ]i\s*th[uư][oờ]ng\s*tr[uú]|qu[eê]\s*qu[aá]n)', text, re.DOTALL)
                
                # Nếu không tìm thấy, thử tìm theo cấu trúc Single-line (VD: ten: Lam My Linh, Ngay)
                if not name_block_matches:
                    text_oneline = text.replace('\n', ' ')
                    name_block_matches = re.findall(r'(?i)(?:h[oọ]\s*(?:v[aà]\s*)?t[eê]n|t[eê]n\s*khai\s*sinh|full\s*name|fui\s*nam|kho\s*v[aà]\s*t[eê]n|\bt[eê]n\b)\s*[:\s]+([^\n,.]+)', text_oneline)

                # Fallback cuối: Tìm các từ viết hoa chuẩn (Title Case) dính liền với chữ Ngày/Sinh (Bị đứt khúc chữ "Họ tên" ở xa)
                if not name_block_matches:
                    text_oneline = text.replace('\n', ' ')
                    name_block_matches = re.findall(r'\b([A-Z][a-zA-ZÀ-Ỹà-ỹ]*(?:\s+[A-Z][a-zA-ZÀ-Ỹà-ỹ]*){1,6})\s*[,.]?\s*(?:Ng[aà]y|Sinh|ng[aà]y|sinh)\b', text_oneline)

                if name_block_matches:
                    # Lọc ra các raw_name không chứa số (ưu tiên cao nhất vì tên không chứa số)
                    no_digit_names = [n for n in name_block_matches if not re.search(r'\d', n)]
                    if no_digit_names:
                        raw_name = sorted(no_digit_names, key=len, reverse=True)[0]
                    else:
                        raw_name = sorted(name_block_matches, key=len, reverse=True)[0]
                    # Nếu tên bị dính chữ NGAY ở cuối (VD: THI LIEU.NGAY sinh:) thì cắt bỏ
                    raw_name = re.sub(r'(?i)\.?\s*ng[aà]y\s*$', '', raw_name)
                    raw_name = raw_name.replace('.', ' ').replace(',', ' ')
                    name_words = []
                    has_upper = any(w.isupper() for w in raw_name.split() if len(w) > 1)
                    for w in raw_name.split():
                        cw = re.sub(r'[^a-zA-Z\xC0-\u024F\u1E00-\u1EFF]', '', w)
                        if len(cw) >= 1:
                            cw_ascii = unicodedata.normalize('NFD', cw.lower())
                            cw_ascii = ''.join(c for c in cw_ascii if unicodedata.category(c) != 'Mn').replace('đ', 'd')
                            is_valid_vn_syllable = cw_ascii in _VN_NAME_SYLLABLES
                            
                            # Lọc rác OCR xen kẽ (VD: LE NGOC cormona THUY LIEU -> bỏ cormona)
                            if has_upper and cw.islower() and len(cw) >= 4 and not is_valid_vn_syllable:
                                continue
                            name_words.append(cw)
                    clean_name = " ".join(name_words).upper()
                    
                    # Fix một số lỗi OCR dính chữ kinh điển TRƯỚC KHI cắt từ rác (để LEDO -> LE DO không bị chém mất)
                    clean_name = clean_name.replace('BICHINHIEN', 'BICH NHIEN')
                    clean_name = re.sub(r'^TRANH\b', 'TRAN', clean_name)
                    clean_name = re.sub(r'^(LE|TRAN|NGUYEN|PHAM|VU|VO|DANG|BUI|DO|HO|PHAN|LY|HUYNH|HOANG|NGO)(THI|VAN|DO|NGOC|XUAN|HUU|MINH|DUY|QUOC|BAO|TRUNG)\b', r'\1 \2', clean_name)
                    
                    # Loại bỏ các từ rác ở đầu chuỗi do OCR ảo giác (vd: KILL Vu Thuc Uyen -> Vu Thuc Uyen)
                    words_upper = clean_name.split()
                    while len(words_upper) > 2:
                        first = unicodedata.normalize('NFD', words_upper[0].lower())
                        first_ascii = ''.join(c for c in first if unicodedata.category(c) != 'Mn').replace('đ', 'd')
                        if first_ascii in _VN_SURNAMES:
                            break
                        words_upper.pop(0)
                    clean_name = " ".join(words_upper)
                    
                    # Dọn dẹp từ rác dính ở cuối tên (VD: NGUYEN THI NGOC TRANG NGAY)
                    clean_name = re.sub(r'\b(?:NGAY|DATE|SINH|BIRTH)\b.*$', '', clean_name).strip()
                    
                    if len(clean_name) > 3 and _is_valid_name(clean_name):
                        data['Họ tên'] = clean_name

                lines = [line.strip() for line in text.split('\n') if line.strip()]

                # ---------------------------------------------------------
                # 4. DUYỆT TỪNG DÒNG (VÉT THÔNG TIN: TÊN, ĐỊA CHỈ, NGÀY THÁNG)
                # OCR đọc ảnh từ trên xuống dưới, nên ta duyệt từng dòng để bắt từ khóa
                # ---------------------------------------------------------
                for i, line in enumerate(lines):
                    line_lower = line.lower()
    
                    # 1. Name
                    if not data['Họ tên'] and any(kw in line_lower for kw in rules.get("name_keywords", [])):
                        if ":" in line:
                            name_part = line.split(":", 1)[1].strip()
                            name_part = name_part.rstrip('.')
                            # Cắt tại dấu phẩy đầu tiên để loại bỏ phần dính sau (VD: "Nguyen Thi Diem Truc, Ngay sinh:")
                            name_part = name_part.split(',')[0].strip()
                            if (name_part.isupper() or name_part.istitle()) and len(name_part) > 3 and _is_valid_name(name_part):
                                data['Họ tên'] = name_part
                            # Nếu có name_part nhưng quá ngắn (VD: chỉ có mỗi chữ NGUYEN), thử ghép với dòng tiếp theo
                            elif len(name_part) > 0 and i + 1 < len(lines):
                                next_line = lines[i+1].replace('|', '').strip()
                                combined_name = f"{name_part} {next_line}".strip()
                                if _is_valid_name(combined_name) and _is_valid_name_starts_with_surname(combined_name):
                                    data['Họ tên'] = combined_name
                        
                        # Nếu không có dấu 2 chấm, thử lấy dòng tiếp theo
                        if not data['Họ tên'] and i + 1 < len(lines):
                            next_line = lines[i+1].replace('|', '').strip()
                            # Phải viết hoa (ALLCAPS hoặc Title Case) VÀ bắt đầu bằng họ VN hợp lệ
                            if ((next_line.isupper() or next_line.istitle())
                                    and len(next_line) > 3
                                    and not re.search(r'\d', next_line)
                                    and _is_valid_name(next_line)
                                    and not any(kw in next_line.lower() for kw in ['ngày', 'sinh', 'quốc', 'tịch', 'giới', 'tính'])):
                                data['Họ tên'] = next_line
                
                    # 2. DOB
                    if "sinh" in line_lower or "birth" in line_lower:
                        for j in range(i, min(i+2, len(lines))):
                            m = re.search(r'\b\d{2}/\d{2}/\d{4}\b', lines[j])
                            if m:
                                data['Ngày sinh'] = m.group(0)
                                break
                
                    # --- BƯỚC 4.3: TRÍCH XUẤT ĐỊA CHỈ (NƠI THƯỜNG TRÚ/CƯ TRÚ) ---
                    addr_kws = rules.get("address_keywords", [])
                    if any(kw in line_lower for kw in addr_kws):
                        addr_parts = []
                        if ":" in line:
                            val = line.split(":", 1)[1].strip()
                        else:
                            addr_pattern = r'(?i)(' + '|'.join(addr_kws) + r')[^a-z0-9]*'
                            m = re.split(addr_pattern, line)
                            val = m[-1].strip() if len(m) > 2 else ""
                            
                        # Loại bỏ các chuỗi nhiễu có thể bám ngay cùng dòng
                        for reg in rules.get("address_clean_regexes_1", []):
                            val = re.sub(reg, '', val).strip()
                        if len(val) >= 2:
                            addr_parts.append(val)
        
                        # Quét các dòng tiếp theo để nối đuôi địa chỉ do địa chỉ thường rất dài và bị rớt dòng.
                        for j in range(i + 1, min(i + 7, len(lines))):
                            next_line = lines[j].replace('|', '').strip()
                            next_lower = next_line.lower()
                            
                            current_addr = ", ".join(filter(bool, addr_parts))
                            commas_count = current_addr.count(',')
                            
                            # CÁC TỪ KHOÁ NGẮT (BREAK) - Rác ngoài thẻ hoặc Mặt sau
                            hard_stops = rules.get("hard_stops", [])
                            soft_stops = rules.get("soft_stops", [])
                            
                            if any(stop_word in next_lower for stop_word in hard_stops):
                                break
                            
                            if any(stop_word in next_lower for stop_word in soft_stops):
                                # Nếu địa chỉ chưa đủ dài (dưới 4 dấu phẩy, địa chỉ VN thường 2-5 dấu phẩy),
                                # có thể đang quét thẻ 2 cột (layout "giá trị đến" nằm song song với địa chỉ).
                                # Bỏ qua break để cho phép regex bên dưới xóa nhãn này và lấy phần địa chỉ còn lại.
                                if commas_count < 4:
                                    pass
                                else:
                                    break
                                
                            # Bỏ qua dòng aho giac OCR: một từ lặp lại >= 3 lần liên tiếp (VD: "điển điển điển điển...")
                            _words_in_line = next_lower.split()
                            if len(_words_in_line) >= 4:
                                _word_counts = {}
                                for _w in _words_in_line:
                                    _word_counts[_w] = _word_counts.get(_w, 0) + 1
                                _max_repeat = max(_word_counts.values())
                                if _max_repeat >= 3:
                                    continue

                            clean_line = next_line
                            for reg in rules.get("address_clean_regexes_2", []):
                                clean_line = re.sub(reg, '', clean_line).strip()
                            
                            # Loại bỏ chữ 'Có' rớt lại do cắt cụm 'Có giá trị đến' bị thiếu
                            # Xử lý các dạng: 'Có :', 'Có', 'Có ,' đứng 1 mình hoặc kẹp ở đầu/cuối chuỗi
                            clean_line = re.sub(r'(?i)^(c[oó]|co\u0301)\s*[:.,]*\s*', '', clean_line).strip()
                            clean_line = re.sub(r'(?i)\s+(c[oó]|co\u0301)\s*[:.,]*$', '', clean_line).strip()
                            
                            # Cắt bỏ rác là số CCCD (12 số) hoặc số điện thoại lọt vào địa chỉ, và ngày tháng bị dính (vd 1010/2037)
                            clean_line = re.sub(r'\b\d{9,15}\b', '', clean_line).strip()
                            clean_line = re.sub(r'\b\d{4}/\d{4}\b', '', clean_line).strip()
                            
                            # Không lấy vào địa chỉ nếu dòng này lọt Họ Tên vào
                            if data.get('Họ tên') and clean_line == data['Họ tên']:
                                continue

                            if len(clean_line) >= 2:
                                # Chống lặp: chỉ skip nếu giống HỆT dòng đã có (case-insensitive)
                                # KHÔNG dùng substring check — sẽ bỏ sót dòng địa chỉ dài hơn
                                is_dup = any(
                                    clean_line.lower() == p.lower()
                                    for p in addr_parts
                                )
                                if not is_dup:
                                    addr_parts.append(clean_line)
                                    
                                    # Kiểm tra xem dòng vừa thêm có phải điểm cuối địa chỉ (chứa tên tỉnh thành) không
                                    # Chuẩn hóa về không dấu lowercase để so sánh
                                    cl_lower = clean_line.lower().replace("đ", "d").replace("-", " ")
                                    cl_nfd = unicodedata.normalize('NFD', cl_lower)
                                    cl_ascii = "".join(c for c in cl_nfd if unicodedata.category(c) != 'Mn')
                                    
                                    is_end = False
                                    for p in _VN_PROVINCES:
                                        # Tìm bằng regex word boundary để tránh match một phần (VD: 'an giang' trong chuỗi khác)
                                        matches = list(re.finditer(r'\b' + re.escape(p) + r'\b', cl_ascii))
                                        if matches:
                                            # Lấy match cuối cùng trên dòng
                                            last_match = matches[-1]
                                            prefix_str = cl_ascii[:last_match.start()].strip().strip(',')
                                            
                                            # Kiểm tra xem ngay trước tỉnh/thành có phải là tiền tố cấp quận/huyện/thành phố không
                                            # VD: "tp soc trang", "thanh pho ben tre", "tx", "thi xa"
                                            is_sub_admin = False
                                            for prefix in ['tp', 'thanh pho', 'tx', 'thi xa', 'huyen', 'quan']:
                                                if prefix_str.endswith(prefix):
                                                    is_sub_admin = True
                                                    break
                                            
                                            if not is_sub_admin:
                                                is_end = True
                                                break
                                                
                                    if is_end:
                                        break
                
                        addr = ", ".join(filter(bool, addr_parts))
                        
                        addr = ", ".join(filter(bool, addr_parts))
                        addr = re.sub(r'(?i)^[a-z]\s+ngày\s*sinh', '', addr).strip()
                        addr = re.sub(r',\s*,', ',', addr)
                        addr = re.sub(r'\s{2,}', ' ', addr)
                        
                        # Sửa các lỗi OCR địa danh kinh điển
                        addr = re.sub(r'(?i)thành phố cần thơn\b', 'Thành Phố Cần Thơ', addr)
                        addr = re.sub(r'(?i)cần thơn\b', 'Cần Thơ', addr)
                        addr = re.sub(r'(?i)lào cần thơ\b', 'Lão, Cần Thơ', addr)
                        # Xóa rác xuất hiện ở cuối địa chỉ ("ngl", "hận", các từ rác OCR đơn lẻ)
                        addr = re.sub(r'(?i),?\s*\b(ngl|hận|ngi)\b\s*,?', ',', addr)
                        addr = re.sub(r',\s*,', ',', addr).strip(',').strip()
                            
                        # Tẩy sạch dấu phẩy thừa do nối chuỗi
                        addr = re.sub(r',\s*,', ',', addr).lstrip(', ').rstrip('., ')
                        
                        addr_cleaned = clean_address_string(addr)
                        data['Nơi thường trú gốc'] = addr_cleaned
        
                    # --- BƯỚC 4.4: TRÍCH XUẤT GIỚI TÍNH (Layer 1 - Label-based line scan) ---
                    if "giới tính" in line_lower or "sex" in line_lower or "gioi tinh" in line_lower:
                        if "nữ" in line_lower or "nư" in line_lower or "nu" in line_lower:
                            layer1_gender = 'Nữ'
                        elif "nam" in line_lower:
                            layer1_gender = 'Nam'
        
                    # --- BƯỚC 4.5: TRÍCH XUẤT NGÀY CẤP ---
                    if ("ngày, tháng, năm" in line_lower or "date, month, year" in line_lower or "date of issue" in line_lower or "cấp" in line_lower or "ngay cap" in line_lower or "cap:" in line_lower or "cap :" in line_lower) and "sinh" not in line_lower and "hết hạn" not in line_lower and "expiry" not in line_lower and "birth" not in line_lower:
                        for j in range(i, min(i+3, len(lines))):
                            m = re.search(r'\b\d{2}/\d{2}/\d{4}\b', lines[j])
                            if m:
                                data['Ngày cấp CCCD'] = m.group(0)
                                break
                
                # ---------------------------------------------------------
                # 5. FALLBACK CHO NGÀY SINH (Nếu thuật toán từ khóa thất bại)
                # Lấy mốc thời gian < 2020 để làm ngày sinh dự phòng
                # ---------------------------------------------------------
                if not data['Ngày sinh'] and all_dates:
                    first_date = all_dates[0]
                    try:
                        if int(first_date.split('/')[-1]) < 2020:
                            data['Ngày sinh'] = first_date
                    except: pass

                # Fallback cho Ngày cấp (Mặt sau chỉ có 1 ngày duy nhất là Ngày cấp)
                if not data['Ngày cấp CCCD'] and data['OCR Side'] == 'Back' and all_dates:
                    # Lấy ngày cuối cùng tìm thấy (tránh nhầm với ngày sinh nếu ảnh ghép 2 mặt)
                    last_date = all_dates[-1]
                    try:
                        if int(last_date.split('/')[-1]) >= 2016:
                            data['Ngày cấp CCCD'] = last_date
                    except: pass

                # Hậu xử lý (Post-processing) làm sạch rác do OCR đọc lem viền
                if data.get('Họ tên'):
                    # Xoá các phụ âm đứng trơ trọi ở cuối tên do vết xước (VD: TRẦN NGỌC MUỘI T -> TRẦN NGỌC MUỘI)
                    data['Họ tên'] = re.sub(r'\s+[BCDGHKLMNPQRSTVX]$', '', data['Họ tên'], flags=re.IGNORECASE).strip()

                # ---------------------------------------------------------
                # 6. TỔNG HỢP GIỚI TÍNH TỪ 3 LAYERS
                # Layer 0 > Layer 1 > Layer 2
                # ---------------------------------------------------------
                layer0_gender = ''
                if data['CCCD'] and len(data['CCCD']) >= 12:
                    gender_digit = data['CCCD'][3]
                    if gender_digit in '02468':
                        layer0_gender = 'Nam'
                    elif gender_digit in '13579':
                        layer0_gender = 'Nữ'

                data['Giới tính'] = layer0_gender or layer1_gender or layer2_gender

                return data

def order_points(pts):
    import numpy as np
    rect = np.zeros((4, 2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)]
    rect[2] = pts[np.argmax(s)]
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)]
    rect[3] = pts[np.argmax(diff)]
    return rect

USE_OPENCV_ALIGN_FIRST = True

def _align_card_opencv(img, target_width=1000, target_height=630):
    import cv2
    import numpy as np
    
    # Resize to a smaller, standard size for robust edge detection
    h, w = img.shape[:2]
    ratio = 800.0 / h
    small_img = cv2.resize(img, (int(w * ratio), 800))
    
    gray = cv2.cvtColor(small_img, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    
    # Enhance contrast
    gray = cv2.equalizeHist(gray)
    
    # Edge detection
    edged = cv2.Canny(gray, 30, 150)
    
    # Morphology to close gaps in the edge of the card
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
    edged = cv2.morphologyEx(edged, cv2.MORPH_CLOSE, kernel)
    
    cnts, _ = cv2.findContours(edged.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if not cnts:
        return None, None
        
    cnts = sorted(cnts, key=cv2.contourArea, reverse=True)[:5]
    
    img_area = small_img.shape[0] * small_img.shape[1]
    
    screenCnt = None
    for c in cnts:
        # Check area: ID card should be relatively large (at least 30% of the image)
        area = cv2.contourArea(c)
        if area < img_area * 0.30:
            continue
            
        peri = cv2.arcLength(c, True)
        
        # Try multiple epsilons to find a 4-point contour
        for eps in [0.02, 0.03, 0.04, 0.05]:
            approx = cv2.approxPolyDP(c, eps * peri, True)
            if len(approx) == 4 and cv2.isContourConvex(approx):
                pts = approx.reshape(4, 2)
                rect_pts = order_points(pts.astype("float32"))
                w_side = (np.linalg.norm(rect_pts[1] - rect_pts[0]) + np.linalg.norm(rect_pts[2] - rect_pts[3])) / 2
                h_side = (np.linalg.norm(rect_pts[3] - rect_pts[0]) + np.linalg.norm(rect_pts[2] - rect_pts[1])) / 2
                if h_side < 10: continue
                aspect_ratio = max(w_side, h_side) / min(w_side, h_side)
                
                # CCCD aspect ratio is approx 1.58 (85.6mm / 53.98mm)
                if 1.3 <= aspect_ratio <= 1.8:
                    screenCnt = approx
                    break
        if screenCnt is not None:
            break
            
    if screenCnt is not None:
        # Scale points back to original image size
        screenCnt = (screenCnt / ratio).astype("int")
        pts = screenCnt.reshape(4, 2)
        rect = order_points(pts.astype("float32"))
        return rect, "opencv_contour"
        
    return None, None

def _align_card_onnx(img):
    import cv2
    import numpy as np
    
    scale = 0.5
    h, w = img.shape[:2]
    small_img = cv2.resize(img, (int(w * scale), int(h * scale)))
    
    from vietocr_engine import get_ocr_engine
    ocr = get_ocr_engine()
    bxs = ocr(small_img, 0)
    
    if bxs:
        min_x = float('inf')
        min_y = float('inf')
        max_x = 0
        max_y = 0
        
        for box, text in bxs:
            for pt in box:
                x, y = pt
                min_x = min(min_x, x)
                min_y = min(min_y, y)
                max_x = max(max_x, x)
                max_y = max(max_y, y)
                
        min_x = int(min_x / scale)
        min_y = int(min_y / scale)
        max_x = int(max_x / scale)
        max_y = int(max_y / scale)
        
        margin_x = int((max_x - min_x) * 0.05)
        margin_y = int((max_y - min_y) * 0.05)
        
        min_x = max(0, min_x - margin_x)
        min_y = max(0, min_y - margin_y)
        max_x = min(w, max_x + margin_x)
        max_y = min(h, max_y + margin_y)
        
        rect = np.array([
            [min_x, min_y],
            [max_x, min_y],
            [max_x, max_y],
            [min_x, max_y]
        ], dtype="float32")
        
        return rect, "onnx_cluster"
        
    return None, None

def align_card(img, target_width=1000, target_height=630, use_opencv_align=None):
    import cv2
    import numpy as np
    original = img.copy()
    debug_img = original.copy()
    
    rect = None
    method = ""
    
    use_opencv = USE_OPENCV_ALIGN_FIRST if use_opencv_align is None else use_opencv_align
    if use_opencv:
        rect, method = _align_card_opencv(img, target_width, target_height)
        
    if rect is None:
        # Fallback to ONNX clustering
        rect, method = _align_card_onnx(img)
        
    if rect is None:
        # Fallback cuối cùng bằng minAreaRect trên mask
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 15))
        morph = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
        cnts, _ = cv2.findContours(morph.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if len(cnts) > 0:
            c = max(cnts, key=cv2.contourArea)
            min_rect = cv2.minAreaRect(c)
            box = cv2.boxPoints(min_rect)
            box = np.intp(box)
            rect = order_points(box.astype("float32"))
            method = "minAreaRect"
        else:
            return original, debug_img, "failed"
            
    # Draw debug points
    if method == "opencv_contour" or method == "minAreaRect":
        cv2.drawContours(debug_img, [rect.astype("int")], -1, (0, 255, 0) if method == "opencv_contour" else (0, 0, 255), 5)
    elif method == "onnx_cluster":
        min_x, min_y = int(rect[0][0]), int(rect[0][1])
        max_x, max_y = int(rect[2][0]), int(rect[2][1])
        cv2.rectangle(debug_img, (min_x, min_y), (max_x, max_y), (255, 0, 0), 5)

    # Xử lý Warp Perspective
    w1 = np.linalg.norm(rect[2] - rect[3])
    w2 = np.linalg.norm(rect[1] - rect[0])
    h1 = np.linalg.norm(rect[1] - rect[2])
    h2 = np.linalg.norm(rect[0] - rect[3])
    max_width = max(int(w1), int(w2))
    max_height = max(int(h1), int(h2))
    
    if max_width < max_height:
        dst = np.array([
            [0, 0],
            [target_height - 1, 0],
            [target_height - 1, target_width - 1],
            [0, target_width - 1]], dtype="float32")
        out_shape = (target_height, target_width)
    else:
        dst = np.array([
            [0, 0],
            [target_width - 1, 0],
            [target_width - 1, target_height - 1],
            [0, target_height - 1]], dtype="float32")
        out_shape = (target_width, target_height)
        
    M = cv2.getPerspectiveTransform(rect, dst)
    warped = cv2.warpPerspective(original, M, out_shape)
    
    return warped, debug_img, method

def extract_ocr_data(image_path_or_cv2img, use_opencv_align=None):
    LOG('Started extract_ocr_data')
    """
    Hàm xử lý OCR (Trích xuất văn bản từ ảnh) bằng AI.
    Hỗ trợ chia 2 luồng độc lập cho mặt sau CCCD (Ngày cấp và MRZ).
    """
    try:
        if isinstance(image_path_or_cv2img, str):
            import cv2
            img_to_ocr = cv2.imread(image_path_or_cv2img)
            if img_to_ocr is None:
                return {"CCCD": "", "CMND": "", "Họ tên": "", "Ngày sinh": "", "Giới tính": "", "Nơi thường trú gốc": "", "Ngày cấp CCCD": "", "OCR Side": ""}, "Không thể đọc file ảnh", None
        else:
            img_to_ocr = image_path_or_cv2img
            
    except Exception as e:
        return {}, f"Lỗi thư viện OCR: {str(e)}", None
        
    try:
        import cv2
        import numpy as np
        import warnings
        import os
        import time
        
        # Benchmarking data
        use_opencv = USE_OPENCV_ALIGN_FIRST if use_opencv_align is None else use_opencv_align
        _last_timing = {}
        t_start_all = time.time()
        
        has_glare_warning = False
        
        def safe_extract_text(img, fast_mode=False):
            nonlocal has_glare_warning
            with warnings.catch_warnings(record=True) as w:
                warnings.simplefilter("always")
                result = extract_text_from_image(img, fast_mode=fast_mode)
                for warning in w:
                    if issubclass(warning.category, RuntimeWarning) and "invalid value encountered in divide" in str(warning.message):
                        has_glare_warning = True
                return result
 
        # --- BƯỚC 1: XÁC ĐỊNH BIÊN VÀ LÀM PHẲNG THẺ ---
        t0 = time.time()
        LOG('Calling align_card')
        card_img, debug_detect_img, align_method = align_card(img_to_ocr, use_opencv_align=use_opencv)
        LOG(f'Finished align_card, method={align_method}')
        _last_timing['detect_and_align_card'] = time.time() - t0
        _last_timing['align_method'] = align_method
        
        # Tối ưu tốc độ: Thu nhỏ ảnh nếu quá lớn (giữ độ nét nhưng giảm khối lượng tính toán cho AI)
        card_h, card_w = card_img.shape[:2]
        max_dim = 1500
        if max(card_w, card_h) > max_dim:
            scale = max_dim / max(card_w, card_h)
            card_img = cv2.resize(card_img, (int(card_w * scale), int(card_h * scale)), interpolation=cv2.INTER_AREA)
        
        debug_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "debug")
        os.makedirs(debug_dir, exist_ok=True)
        cv2.imwrite(os.path.join(debug_dir, "original_image.jpg"), img_to_ocr)
        
        if align_method == "contour":
            cv2.imwrite(os.path.join(debug_dir, "detected_card_contour.jpg"), debug_detect_img)
        else:
            cv2.imwrite(os.path.join(debug_dir, "detected_card_fallback_box.jpg"), debug_detect_img)
            
        cv2.imwrite(os.path.join(debug_dir, "warped_card_normalized.jpg"), card_img)

        # --- BƯỚC 2: TÌM HƯỚNG XOAY CHUẨN DỰA TRÊN WARPED CARD ---
        if align_method == "opencv_contour":
            rotations = [
                (None, "Không xoay"),
                (cv2.ROTATE_180, "Xoay 180 độ")
            ]
        else:
            rotations = [
                (None, "Không xoay"),
                (cv2.ROTATE_90_COUNTERCLOCKWISE, "Xoay trái 90 độ"),
                (cv2.ROTATE_90_CLOCKWISE, "Xoay phải 90 độ"),
                (cv2.ROTATE_180, "Xoay 180 độ")
            ]
        
        is_back_side = False
        best_rot_score = -1
        best_back_rotated_img = None
        best_back_rot_name = ""
        best_raw_mrz_text = ""
        best_mrz_lines = []
        best_thresh_bottom_img = None
        
        t0_rot = time.time()
        t_ocr_total = 0
        for rot_code, rot_name in rotations:
            rotated = card_img if rot_code is None else cv2.rotate(card_img, rot_code)
            hr, wr = rotated.shape[:2]
            
            # Crop bottom 35% từ ảnh thẻ ĐÃ WARP
            bottom_crop = rotated[int(hr * 0.65):hr, :]
            
            # Lấy text bằng Tesseract thay vì VietOCR để tăng tốc X10
            t_ocr_start = time.time()
            import pytesseract
            
            gray_bottom = cv2.cvtColor(bottom_crop, cv2.COLOR_BGR2GRAY)
            hr_b, wr_b = gray_bottom.shape[:2]
            resized_b = cv2.resize(gray_bottom, (wr_b*2, hr_b*2), interpolation=cv2.INTER_CUBIC)
            bordered = cv2.copyMakeBorder(resized_b, 40, 40, 40, 40, cv2.BORDER_CONSTANT, value=[255])
            thresh_bottom = cv2.threshold(bordered, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]
            
            custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789<'
            text_bottom = pytesseract.image_to_string(thresh_bottom, config=custom_config)
            t_ocr_total += (time.time() - t_ocr_start)
            
            # Đánh giá điểm (Score)
            upper_text = text_bottom.upper()
            score = 0
            if 'IDVNM' in upper_text:
                score += 500
            if 'VNM' in upper_text:
                score += 200
            
            # Khôi phục một phần dấu < bị nhận diện sai thành K hoặc khoảng trắng (CHỈ cho mrz_lines xuất ra)
            text_bottom_fixed = upper_text.replace('K', '<').replace(' ', '<')
            clean_mrz_text = re.sub(r'[^A-Z0-9<\n]', '', text_bottom_fixed)
            mrz_lines = [l.strip() for l in clean_mrz_text.split('\n') if len(l.strip()) >= 20]
            
            if score > best_rot_score and score > 50:
                is_back_side = True
                best_rot_score = score
                best_back_rotated_img = rotated
                best_back_rot_name = rot_name
                best_raw_mrz_text = text_bottom
                best_mrz_lines = mrz_lines
                best_thresh_bottom_img = thresh_bottom
            
            # Dừng sớm nếu đã nhận ra mặt sau với độ tin cậy cao (có cả IDVNM lẫn VNM)
            if best_rot_score >= 700:
                break
        
        _last_timing['orientation_detection'] = (time.time() - t0_rot) - t_ocr_total
        _last_timing['ocr_full_card'] = t_ocr_total
                
        # --- BƯỚC 3: XỬ LÝ THEO MẶT THẺ ---
        best_data = {'CCCD': '', 'Họ tên': '', 'Ngày sinh': '', 'OCR Side': '', 'Raw Text Upper': ''}
        best_note = "Ảnh mờ hoặc không thể nhận diện được"
        rotated_return = None
        
        if is_back_side:
            # === PIPELINE MẶT SAU (CHIA 2 LUỒNG) ===
            best_data['OCR Side'] = 'Back'
            rotated_return = best_back_rotated_img
            best_note = f"Lấy bằng OCR 2 luồng độc lập ({best_back_rot_name})"
            hr, wr = best_back_rotated_img.shape[:2]
            
            # Ghi ảnh rotated_best_back
            cv2.imwrite(os.path.join(debug_dir, "rotated_best_back.jpg"), best_back_rotated_img)
            cv2.imwrite(os.path.join(debug_dir, "thresholded_mrz.jpg"), best_thresh_bottom_img)
            
            # LUỒNG 1: MRZ
            t_mrz = time.time()
            bottom_crop = best_back_rotated_img[int(hr * 0.65):hr, :]
            cv2.imwrite(os.path.join(debug_dir, "bottom_region_mrz.jpg"), bottom_crop)
            
            # Cố gắng dùng Tesseract với preprocessing đặc biệt để lấy định dạng MRZ chuẩn xác (có dấu <)
            try:
                gray_bottom = cv2.cvtColor(bottom_crop, cv2.COLOR_BGR2GRAY)
                hr_b, wr_b = gray_bottom.shape[:2]
                resized_b = cv2.resize(gray_bottom, (wr_b*2, hr_b*2), interpolation=cv2.INTER_CUBIC)
                bordered = cv2.copyMakeBorder(resized_b, 40, 40, 40, 40, cv2.BORDER_CONSTANT, value=[255])
                
                import pytesseract
                custom_config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789<'
                
                # 3 mức threshold: tự động (Otsu), nhẹ (120), mạnh (160)
                # Thử từ thấp đến cao, chọn mức nào cho nhiều dòng MRZ hợp lệ nhất
                thresh_levels = [
                    ("otsu",  lambda img: cv2.threshold(img, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)[1]),
                    ("light", lambda img: cv2.threshold(img, 120, 255, cv2.THRESH_BINARY)[1]),
                    ("strong",lambda img: cv2.threshold(img, 160, 255, cv2.THRESH_BINARY)[1]),
                ]
                
                best_tess_lines = []
                best_thresh_img = None
                
                for level_name, thresh_fn in thresh_levels:
                    thresh_tess = thresh_fn(bordered)
                    tess_text = pytesseract.image_to_string(thresh_tess, config=custom_config)
                    
                    candidate_lines = []
                    for line in tess_text.split('\n'):
                        line = line.strip()
                        if not line: continue
                        if len(line) > 30 and ' ' in line[:3]:
                            line = line.split(' ', 1)[-1]
                        line_clean = re.sub(r'[^A-Z0-9<]', '', line)
                        if len(line_clean) >= 20:
                            candidate_lines.append(line_clean)
                    
                    # Sửa lỗi phổ biến
                    if candidate_lines and candidate_lines[0].startswith('1DVNM'):
                        candidate_lines[0] = 'IDVNM' + candidate_lines[0][5:]
                    
                    # Chọn mức nào có >= 3 dòng MRZ HOẶC nhiều ký tự nhất
                    total_chars = sum(len(l) for l in candidate_lines)
                    best_chars = sum(len(l) for l in best_tess_lines)
                    has_idvnm = any('IDVNM' in l for l in candidate_lines)
                    
                    if (len(candidate_lines) >= 3 and has_idvnm) or total_chars > best_chars:
                        best_tess_lines = candidate_lines
                        best_thresh_img = thresh_tess
                    
                    # Dừng sớm nếu đã có đủ 3 dòng và có IDVNM
                    if len(best_tess_lines) >= 3 and any('IDVNM' in l for l in best_tess_lines):
                        break
                
                if best_thresh_img is not None:
                    cv2.imwrite(os.path.join(debug_dir, "thresholded_mrz.jpg"), best_thresh_img)
                    
                if len(best_tess_lines) >= 3:
                    best_mrz_lines = best_tess_lines
            except Exception:
                pass
            
            if best_mrz_lines:
                temp_data = parse_ocr_text("\n".join(best_mrz_lines))
                
                # FUSE: Thử lấy tên từ VietOCR (vì VietOCR đọc chữ cái tốt hơn Tesseract)
                name_vietocr = ""
                if 'best_raw_mrz_text' in locals() and best_raw_mrz_text:
                    temp_data_vietocr = parse_ocr_text(best_raw_mrz_text.upper())
                    name_vietocr = temp_data_vietocr.get('Họ tên', '')
                    
                mrz_name = name_vietocr if name_vietocr else temp_data.get('Họ tên', '')
                mrz_cccd = temp_data.get('CCCD', '') or (temp_data_vietocr.get('CCCD', '') if 'temp_data_vietocr' in locals() else '')
                
                # CHỐNG LỖI MÀN HÌNH CHAT ZALO/SMS CHỨA THẺ CŨ Ở DƯỚI:
                # Ảnh mặt sau thẻ thật KHÔNG BAO GIỜ có tên ở phần chữ in (chỉ có ở phần MRZ).
                # Nếu trước khi xử lý MRZ mà `best_data` ĐÃ CÓ TÊN, thì 100% tên đó đến từ nội dung tin nhắn SMS!
                # Nội dung tin nhắn SMS là Golden Truth, KHÔNG ĐƯỢC để MRZ (có thể là của thẻ cũ nằm dưới) ghi đè!
                
                if best_data.get('Họ tên') and mrz_name and best_data['Họ tên'] != mrz_name:
                    best_note += " | CẢNH BÁO: Phát hiện MRZ của thẻ cũ trong lịch sử chat, ưu tiên tin nhắn gốc."
                else:
                    if mrz_name: best_data['Họ tên'] = mrz_name
                    
                if best_data.get('CCCD') and mrz_cccd and best_data['CCCD'] != mrz_cccd:
                    pass # Giữ CCCD gốc từ SMS
                else:
                    if mrz_cccd: best_data['CCCD'] = mrz_cccd
                    
                mrz_gender = temp_data.get('Giới tính', '') or (temp_data_vietocr.get('Giới tính', '') if 'temp_data_vietocr' in locals() else '')
                if mrz_gender and not best_data.get('Giới tính'):
                    best_data['Giới tính'] = mrz_gender
            _last_timing['mrz_extraction'] = time.time() - t_mrz
            
            t_ocr_back = time.time()
            # LUỒNG 2: NGÀY CẤP (Top 70%)
            top_crop = best_back_rotated_img[0:int(hr * 0.70), :]
            cv2.imwrite(os.path.join(debug_dir, "top_region_issue_date.jpg"), top_crop)
            
            # Tiền xử lý nhẹ: Contrast
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            lab = cv2.cvtColor(top_crop, cv2.COLOR_BGR2LAB)
            l, a, b = cv2.split(lab)
            l2 = clahe.apply(l)
            img_contrast = cv2.cvtColor(cv2.merge((l2,a,b)), cv2.COLOR_LAB2BGR)
            
            raw_issue_text = safe_extract_text(img_contrast, fast_mode=True)
            
            issue_date = ""
            lines = raw_issue_text.split('\n')
            for i, line in enumerate(lines):
                ll = line.lower()
                if "ngày" in ll or "tháng" in ll or "date" in ll or "năm" in ll or "cấp" in ll:
                    for j in range(i, min(i+3, len(lines))):
                        m = re.search(r'(?<!\d)\d{2}/\d{2}/\d{4}(?!\d)', lines[j])
                        if m:
                            issue_date = m.group(0)
                            break
                if issue_date: break
            
            best_data['Ngày cấp CCCD'] = issue_date
            
            best_data['back_side_raw'] = {
                "issue_date": issue_date,
                "mrz_lines": best_mrz_lines,
                "raw_issue_text": raw_issue_text,
                "raw_mrz_text": best_raw_mrz_text,
                "confidence": {"issue_date": 1.0 if issue_date else 0.0, "mrz": 1.0 if best_mrz_lines else 0.0}
            }
            
            best_data['Raw Text'] = f"--- TOP TEXT (Ngày cấp) ---\n{raw_issue_text}\n\n--- BOTTOM TEXT (MRZ) ---\n{best_raw_mrz_text}"
            _last_timing['ocr_từng_field'] = time.time() - t_ocr_back

            # --- TÍCH HỢP ADDRESS V1.5 (ZONAL OCR MẶT SAU THẺ MỚI) ---
            try:
                import debug_address_v15
                if best_back_rotated_img is not None:
                    _, addr_v15, status_v15, _ = debug_address_v15.extract_address_v15(best_back_rotated_img)
                    if status_v15 == "Pass" and addr_v15:
                        addr_v15 = re.sub(r'(?i)thành phố cần thơn\b', 'Thành Phố Cần Thơ', addr_v15)
                        addr_v15 = re.sub(r'(?i)cần thơn\b', 'Cần Thơ', addr_v15)
                        addr_v15 = re.sub(r'(?i)lào cần thơ\b', 'Lão, Cần Thơ', addr_v15)
                        best_data['Nơi thường trú gốc'] = clean_address_string(addr_v15)
                        best_note += " [Zonal Address V1.5 (Back)]"
            except Exception as e:
                LOG(f"Address V1.5 error on back side: {e}")
            
            return best_data, best_note, rotated_return
            
        else:
            # === PIPELINE MẶT TRƯỚC (SỬ DỤNG ẢNH THẺ ĐÃ WARP CHUẨN HOÁ) ===
            best_data = {'CCCD': '', 'Họ tên': '', 'Ngày sinh': '', 'OCR Side': 'Front', 'Raw Text Upper': ''}
            best_note = "Ảnh mờ hoặc không thể nhận diện được"
            rotated_return = None
            
            # Cần thử hướng xoay vì align_card có thể trả về ảnh bị xoay
            # Do thuật toán OpenCV luôn canh lề chuẩn hình chữ nhật nằm ngang (landscape),
            # nên thẻ chỉ có thể nằm đúng chiều (0 độ) hoặc lật ngược (180 độ).
            if align_method == "opencv_contour":
                front_rotations = [
                    (None, "Không xoay"),
                    (cv2.ROTATE_180, "Xoay 180 độ")
                ]
            else:
                front_rotations = [
                    (None, "Không xoay"),
                    (cv2.ROTATE_90_COUNTERCLOCKWISE, "Xoay trái 90 độ"),
                    (cv2.ROTATE_90_CLOCKWISE, "Xoay phải 90 độ"),
                    (cv2.ROTATE_180, "Xoay 180 độ")
                ]
            
            keywords = ["CỘNG HÒA", "ĐỘC LẬP", "CĂN CƯỚC", "SỐ / NO", "HỌ VÀ TÊN"]
            
            best_front_score = -1
            best_front_data = None
            best_front_img = None
            best_front_note = ""
            
            t_front_orient = time.time()
            t_front_ocr = 0
            t_front_parse = 0
            
            for rot_code, rot_name in front_rotations:
                rotated = card_img if rot_code is None else cv2.rotate(card_img, rot_code)
                
                # Tăng contrast nhẹ cho mặt trước
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                lab = cv2.cvtColor(rotated, cv2.COLOR_BGR2LAB)
                l, a, b = cv2.split(lab)
                l2 = clahe.apply(l)
                front_contrast = cv2.cvtColor(cv2.merge((l2,a,b)), cv2.COLOR_LAB2BGR)
                
                t_o = time.time()
                text_rot = safe_extract_text(front_contrast, fast_mode=True)
                t_front_ocr += (time.time() - t_o)
                
                t_p = time.time()
                data_rot = parse_ocr_text(text_rot)
                t_front_parse += (time.time() - t_p)
                
                # Chấm điểm độ tin cậy của hướng xoay
                score = 0
                if data_rot.get('CCCD'): score += 100
                if data_rot.get('Họ tên'): score += 50
                if data_rot.get('Ngày sinh'): score += 30
                if data_rot.get('OCR Side') == 'Front': score += 200
                
                upper_text = text_rot.upper()
                for kw in keywords:
                    if kw in upper_text:
                        score += 20
                
                is_sms = any(kw in upper_text for kw in ["THUÊ BAO","TTTB","THUE BAO","TB", "MOBIFONE", "VINAPHONE", "VIETTEL", "TRẢ TRƯỚC", "TRA TRUOC", "TÀI KHOẢN", "TAI KHOAN", "GÓI CƯỚC", "GOI CUOC", "MẬT KHẨU", "QUÝ KHÁCH"])
                
                # Phân tích trật tự từ trên xuống dưới để phạt ảnh bị lộn ngược (180 độ)
                if data_rot.get('OCR Side') == 'Back':
                    idx_ngay = min([upper_text.find(x) for x in ["NGÀY", "CỤC", "THÁNG"] if x in upper_text] + [99999])
                    idx_mrz = min([upper_text.find(x) for x in ["VNM", "IDVNM"] if x in upper_text] + [99999])
                    if idx_ngay != 99999 and idx_mrz != 99999:
                        if idx_ngay < idx_mrz: score += 100  # Đúng chiều (Ngày cấp ở trên MRZ)
                        else: score -= 100                   # Lộn ngược
                else:
                    idx_top = min([upper_text.find(x) for x in ["CỘNG", "CĂN", "CHỨNG"] if x in upper_text] + [99999])
                    idx_bot = min([upper_text.find(x) for x in ["CẤP", "THƯỜNG TRÚ", "ĐỊA CHỈ", "QUỐC TỊCH"] if x in upper_text] + [99999])
                    if idx_top != 99999 and idx_bot != 99999:
                        if idx_top < idx_bot: score += 100   # Đúng chiều
                        else: score -= 100                   # Lộn ngược
                    elif idx_top == 99999 and idx_bot == 99999 and data_rot.get('CCCD'):
                        # Nếu không có từ khóa mặt trước nhưng CÓ đủ 3 trường quan trọng -> Khả năng cao là SMS Screenshot
                        if data_rot.get('Họ tên') and data_rot.get('Ngày sinh'):
                            score += 100 # Thưởng thêm điểm cho SMS hợp lệ
                        else:
                            # Tìm thấy dãy 12 số nhưng KHÔNG hề có từ khóa nào của thẻ CCCD mặt trước và thiếu tên/ngày sinh
                            # Khả năng cực cao là AI bị ảo giác (đọc chữ lộn ngược của ảnh screenshot sinh ra số rác)
                            score -= 300
                        
                if score > best_front_score:
                    best_front_score = score
                    best_front_data = data_rot
                    best_front_img = rotated
                    best_front_note = f"Lấy bằng OCR ({rot_name})" + (" [SMS]" if is_sms else "")
                
                # Dừng sớm nếu điểm đã đủ cao (có CCCD + tên + nhận diện đúng mặt trước) HOẶC là SMS có đủ 3 trường
                if best_front_score >= 250 or (data_rot.get('CCCD') and data_rot.get('Họ tên') and data_rot.get('Ngày sinh')) or is_sms:
                    break
            
            _last_timing['orientation_detection'] = (time.time() - t_front_orient) - t_front_ocr - t_front_parse
            _last_timing['ocr_full_card'] = t_front_ocr
            _last_timing['parse_fields'] = t_front_parse
            
            # Nếu điểm quá thấp, có thể align_card crop sai hoặc CLAHE làm hỏng text (thường gặp với ảnh chụp màn hình)
            # Fallback về OCR ảnh gốc toàn phần (không làm nét)
            # Tối ưu: Nếu là opencv_contour thì sẽ được retry với ONNX sau, không cần tốn thời gian fallback gốc
            if best_front_score < 50 and align_method != "opencv_contour":
                # Resize ảnh gốc nếu quá lớn để tránh nghẽn CPU (DBNet trên ảnh 12MP tốn >20s)
                h_orig, w_orig = img_to_ocr.shape[:2]
                img_fallback = img_to_ocr.copy()
                if max(h_orig, w_orig) > 1500:
                    sc = 1500 / max(h_orig, w_orig)
                    img_fallback = cv2.resize(img_fallback, (int(w_orig * sc), int(h_orig * sc)), interpolation=cv2.INTER_AREA)

                t_fallback_orient = time.time()
                for rot_code, rot_name in front_rotations:
                    rotated = img_fallback if rot_code is None else cv2.rotate(img_fallback, rot_code)
                    
                    t_o = time.time()
                    LOG(f'Calling safe_extract_text fast_mode=True rot={rot_name}')
                    text_rot = safe_extract_text(rotated, fast_mode=True)
                    LOG('Finished safe_extract_text')
                    _last_timing['ocr_full_card'] += (time.time() - t_o)
                    
                    t_p = time.time()
                    data_rot = parse_ocr_text(text_rot)
                    _last_timing['parse_fields'] += (time.time() - t_p)
                    
                    score = 0
                    if data_rot.get('CCCD'): score += 100
                    if data_rot.get('Họ tên'): score += 50
                    if data_rot.get('Ngày sinh'): score += 30
                    if data_rot.get('OCR Side') == 'Front': score += 200
                    
                    upper_text = text_rot.upper()
                    for kw in keywords:
                        if kw in upper_text:
                            score += 20
                            
                    is_sms = any(kw in upper_text for kw in ["THUÊ BAO","TTTB","THUE BAO","TB", "MOBIFONE", "VINAPHONE", "VIETTEL", "TRẢ TRƯỚC", "TRA TRUOC", "TÀI KHOẢN", "TAI KHOAN", "GÓI CƯỚC", "GOI CUOC", "MẬT KHẨU", "QUÝ KHÁCH"])
                
                # Phân tích trật tự từ trên xuống dưới để phạt ảnh bị lộn ngược (180 độ)
                    if data_rot.get('OCR Side') == 'Back':
                        idx_ngay = min([upper_text.find(x) for x in ["NGÀY", "CỤC", "THÁNG"] if x in upper_text] + [99999])
                        idx_mrz = min([upper_text.find(x) for x in ["VNM", "IDVNM"] if x in upper_text] + [99999])
                        if idx_ngay != 99999 and idx_mrz != 99999:
                            if idx_ngay < idx_mrz: score += 100
                            else: score -= 100
                    else:
                        idx_top = min([upper_text.find(x) for x in ["CỘNG", "CĂN", "CHỨNG"] if x in upper_text] + [99999])
                        idx_bot = min([upper_text.find(x) for x in ["CẤP", "THƯỜNG TRÚ", "ĐỊA CHỈ"] if x in upper_text] + [99999])
                        if idx_top != 99999 and idx_bot != 99999:
                            if idx_top < idx_bot: score += 100
                            else: score -= 100
                        elif idx_top == 99999 and idx_bot == 99999 and data_rot.get('CCCD'):
                            if data_rot.get('Họ tên') and data_rot.get('Ngày sinh'):
                                score += 100
                            else:
                                score -= 300
                            
                    if score > best_front_score:
                        best_front_score = score
                        best_front_data = data_rot
                        best_front_img = rotated
                        best_front_note = f"Lấy bằng OCR toàn phần ({rot_name})" + (" [SMS]" if is_sms else "")
                    
                    # Dừng sớm nếu đủ điểm HOẶC là SMS
                    if best_front_score >= 250 or (data_rot.get('CCCD') and data_rot.get('Họ tên') and data_rot.get('Ngày sinh')) or is_sms:
                        break
                    
            best_img = None
            if best_front_data and best_front_score > 0:
                best_data = best_front_data
                best_img = best_front_img
                rotated_return = best_front_img
                best_note = best_front_note
            else:
                best_img = card_img
                rotated_return = card_img
                    
            # (Phần xử lý tương phản và làm nét cho mặt trước giữ nguyên)
            def missing_critical(d):
                # Nếu là ảnh SMS (không có các mốc CỘNG HÒA của thẻ) thì bỏ qua không cần cố tìm địa chỉ bằng mọi giá
                if best_front_score > 50 and best_front_score < 290 and d.get('CCCD') and d.get('Họ tên'):
                    return False
                return not d.get('CCCD') or not d.get('Họ tên') or not d.get('Ngày sinh') or not d.get('Nơi thường trú gốc')
                
            is_sms_detected = "[SMS]" in best_note
            if missing_critical(best_data) and best_img is not None and not is_sms_detected:
                clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
                lab = cv2.cvtColor(best_img, cv2.COLOR_BGR2LAB)
                l, a, b = cv2.split(lab)
                l2 = clahe.apply(l)
                img_contrast = cv2.cvtColor(cv2.merge((l2,a,b)), cv2.COLOR_LAB2BGR)
                
                LOG('Calling safe_extract_text fast_mode=False (contrast)')
                text_p2 = safe_extract_text(img_contrast)
                LOG('Finished safe_extract_text')
                data_p2 = parse_ocr_text(text_p2)
                
                merged = False
                for k, v in data_p2.items():
                    # fast_mode có thể trích xuất thiếu chữ (VD: tên thiếu chữ cuối do box nhỏ bị loại bỏ)
                    # Nên OCR toàn phần phải được ưu tiên ghi đè nếu nó tìm thấy dữ liệu
                    if v:
                        best_data[k] = v
                        merged = True
                        
                if merged:
                    best_note += " + Lọc Tương phản"
                    
                if missing_critical(best_data):
                    kernel = np.array([[-1,-1,-1], [-1,9,-1], [-1,-1,-1]])
                    img_sharpen = cv2.filter2D(best_img, -1, kernel)
                    
                    LOG('Calling safe_extract_text fast_mode=False (sharpen)')
                    text_p3 = safe_extract_text(img_sharpen)
                    LOG('Finished safe_extract_text')
                    data_p3 = parse_ocr_text(text_p3)
                    
                    merged_p3 = False
                    for k, v in data_p3.items():
                        if v and not best_data.get(k):
                            best_data[k] = v
                            merged_p3 = True
                    
                    if merged_p3:
                        best_note += " + Làm nét"

            # --- TÍCH HỢP ADDRESS V1.5 (ZONAL OCR) ---
            try:
                import debug_address_v15
                if best_img is not None:
                    _, addr_v15, status_v15, _ = debug_address_v15.extract_address_v15(best_img)
                    if status_v15 == "Pass" and addr_v15:
                        addr_v15 = re.sub(r'(?i)thành phố cần thơn\b', 'Thành Phố Cần Thơ', addr_v15)
                        addr_v15 = re.sub(r'(?i)cần thơn\b', 'Cần Thơ', addr_v15)
                        addr_v15 = re.sub(r'(?i)lào cần thơ\b', 'Lão, Cần Thơ', addr_v15)
                        best_data['Nơi thường trú gốc'] = clean_address_string(addr_v15)
                        best_note += " [Zonal Address V1.5]"
            except Exception as e:
                LOG(f"Address V1.5 error: {e}")

            max_score = 0
            if 'best_front_score' in locals() and best_front_score > 0:
                max_score = max(max_score, best_front_score)
            if 'best_rot_score' in locals() and best_rot_score > 0:
                max_score = max(max_score, best_rot_score)
                
            is_garbage = (not best_data['CCCD'] and not best_data['Họ tên'] and not best_data['Ngày sinh']) or (max_score < 50)
            
            is_suspicious_crop = False
            if best_data['OCR Side'] == 'Front':
                if not best_data['Họ tên'] or not best_data['CCCD'] or len(best_data['CCCD']) < 12:
                    is_suspicious_crop = True
            elif best_data['OCR Side'] == 'Back':
                # Require either Issue Date or MRZ
                if not best_data.get('Ngày cấp CCCD') and not best_data.get('CCCD'):
                    is_suspicious_crop = True
            
            if is_garbage or is_suspicious_crop:
                # Nếu crop OpenCV fail dẫn đến không đọc được chữ, retry bằng ONNX
                if use_opencv and align_method == "opencv_contour":
                    res_data, res_note, res_img = extract_ocr_data(img_to_ocr, use_opencv_align=False)
                    if res_note:
                        res_note = res_note + " (OpenCV crop fail, fallback to ONNX)"
                    return res_data, res_note, res_img
                    
                return best_data, "Ảnh mờ hoặc không thể nhận diện được", rotated_return
                
            return best_data, best_note, rotated_return
            
    except Exception as e:
        return {}, f"Lỗi OCR: {str(e)}", None

def process_qr_string(qr_string):
    parts = qr_string.split('|')
    addr_raw = parts[5] if len(parts) > 5 else ''
    addr_clean = re.sub(r',\s*-\s*', ' ', addr_raw) # Xóa ', -' (vd: KDC, - Hưng Phú 1 -> KDC Hưng Phú 1)
    addr_clean = re.sub(r',\s*,', ',', addr_clean) # Thay thế ', ,' hoặc ',,' bằng ','
    addr_clean = re.sub(r'\s+', ' ', addr_clean).strip(', ')

    data = {
        'CCCD': parts[0] if len(parts) > 0 else '',
        'CMND': parts[1] if len(parts) > 1 else '',
        'Họ tên': parts[2] if len(parts) > 2 else '',
        'Ngày sinh': format_date(parts[3]) if len(parts) > 3 else '',
        'Giới tính': parts[4] if len(parts) > 4 else '',
        'Nơi thường trú gốc': addr_clean,
        'Ngày cấp CCCD': format_date(parts[6]) if len(parts) > 6 else '',
    }
    
    notes = []
    if not data['Ngày cấp CCCD']:
        notes.append('Thiếu ngày cấp')
    if not data['Nơi thường trú gốc']:
        notes.append('Thiếu nơi thường trú')
        
    return data, notes

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# Global session with connection pooling to make multi-threading actually fast
_address_session = requests.Session()
_adapter = HTTPAdapter(pool_connections=10, pool_maxsize=10)
_address_session.mount('http://', _adapter)
_address_session.mount('https://', _adapter)

def fetch_single_address(addr):
    LOG(f'Started fetch_single_address for {addr}')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:151.0) Gecko/20100101 Firefox/151.0',
        'Accept': 'application/json, text/plain, */*',
        'Content-Type': 'application/json',
        'x-kas': '89232422',
        'Origin': 'https://tienich.vnhub.com',
        'Referer': 'https://tienich.vnhub.com/'
    }
    import json, time, re
    t0 = time.time()
    
    # Tiền xử lý: API VNHub rất nhạy cảm với khoảng trắng thừa, đặc biệt là khoảng trắng trước dấu phẩy
    # Ví dụ: "Số Nhà 137B , Trần Hưng Đạo ," -> lỗi data: []
    clean_addr = re.sub(r'\s+,', ',', addr)
    clean_addr = re.sub(r'\s+', ' ', clean_addr).strip()
    
    payload = json.dumps({"address": clean_addr})
    
    # Retry tối đa 100 lần đối với lỗi mạng/500
    # Nếu API trả về thành công nhưng data = [] thì chuyển ngay sang backup
    last_exception = None
    
    for attempt in range(30):
        try:
            response = _address_session.post(
                'https://tienich.vnhub.com/api/wards',
                data=payload,
                headers=headers,
                timeout=5
            )
            response.raise_for_status()
            res_data = response.json()
            
            if res_data.get('success') and res_data.get('data') and len(res_data['data']) > 0 and res_data['data'][0].get('address'):
                converted_addr = res_data['data'][0]['address']
                
                # Tiêu diệt rác API VNHub (VD: trả về chữ "Substates")
                converted_addr = re.sub(r'(?i)\bsubstates\b', '', converted_addr)
                converted_addr = re.sub(r',\s*,', ',', converted_addr) # Xóa dấu phẩy thừa do xóa chữ
                converted_addr = re.sub(r'\s+', ' ', converted_addr).strip(', ')
                
                return {
                    "original": addr,
                    "success": True,
                    "converted": converted_addr,
                    "_processing_time": time.time() - t0
                }
            
            # Nếu API trả về data rỗng, có thể do Rate Limit ngầm từ VNHub
            # Ngủ 2s rồi thử lại
            if attempt < 29:
                time.sleep(2)
                continue
            break
            
        except Exception as e:
            last_exception = e
            # Lỗi 500 hoặc mạng → thử lại sau 1s
            if attempt < 29:
                time.sleep(1)
                continue
            break  # hết 30 lần → dùng backup

    # -------------------------------------------------------
    # BACKUP: Geovina.io.vn — chỉ chạy khi VNHub thất bại
    # Response field cần lấy: data.full_new_address
    # -------------------------------------------------------
    import os as _os
    while True:
        geovina_token = _os.environ.get(
            'GEOVINA_DEMO_TOKEN',
            '1782095814853.ab6fc10225b874be936bd6fe9a020c6e0a5418e03a1215c0463d5628c91083e7'
        )
        geovina_headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:152.0) Gecko/20100101 Firefox/152.0',
            'Accept': 'application/json',
            'Accept-Language': 'en-US,vi;q=0.9,en;q=0.8',
            'Content-Type': 'application/json',
            'X-Demo-Token': geovina_token,
            'X-Api-Key': 'gvn_5740dceda5cb2424b787f1153da3802a721ae3f6',
            'Referer': 'https://www.geovina.io.vn/',
            'Origin': 'https://www.geovina.io.vn',
            'Connection': 'keep-alive',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'Priority': 'u=0',
            'TE': 'trailers'
        }
        
        try:
            # Retry nội bộ cho Geovina (tránh 429 Too Many Requests)
            for geo_attempt in range(4):
                geo_resp = _address_session.post(
                    'https://www.geovina.io.vn/parse',
                    data=payload,
                    headers=geovina_headers,
                    timeout=15
                )
                if geo_resp.status_code == 429:
                    time.sleep(2)
                    continue
                geo_resp.raise_for_status()
                break
                
            geo_data = geo_resp.json()
            
            is_token_error = False
            if geo_resp.status_code in (401, 403):
                is_token_error = True
            elif not geo_data.get('success'):
                err_text = geo_data.get('error', '').lower()
                if "token" in err_text or "hết hạn" in err_text or "không hợp lệ" in err_text or "unauthorized" in err_text:
                    is_token_error = True
                    
            if is_token_error:
                updated = check_and_prompt_geovina_token(current_failed_token=geovina_token)
                if updated:
                    continue # Thử lại với token mới
                else:
                    return {"original": addr, "success": False, "error": geo_data.get('error', 'Token Geovina đã hết hạn và người dùng đã bỏ qua.'), "_processing_time": time.time() - t0}

            full_new = geo_data.get('data', {}).get('full_new_address', '')
            if geo_data.get('success') and full_new:
                converted_addr = re.sub(r'\s+', ' ', full_new).strip(', ')
                return {"original": addr, "success": True, "converted": converted_addr, "source": "geovina_backup", "_processing_time": time.time() - t0}

            return {"original": addr, "success": False, "error": geo_data.get('error', "Không tìm thấy địa chỉ tương ứng (VNHub + Geovina đều thất bại)"), "_processing_time": time.time() - t0}

        except Exception as geo_err:
            err_primary = str(last_exception) if last_exception else "data rỗng sau nhiều lần thử"
            return {
                "original": addr,
                "success": False,
                "error": f"Lỗi kết nối API sau nhiều lần thử VNHub ({err_primary}); Geovina backup cũng lỗi ({str(geo_err)})",
                "_processing_time": time.time() - t0
            }
def call_address_api(address_list, max_workers=4):
    if not address_list:
        return []
        
    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(fetch_single_address, addr): addr for addr in address_list}
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())
            
    return results

from rich.console import Console
from rich.panel import Panel
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn, TimeRemainingColumn, ProgressColumn
from rich.prompt import Prompt, Confirm

class CountColumn(ProgressColumn):
    def render(self, task):
        from rich.text import Text
        completed = int(task.completed)
        total = int(task.total) if task.total else 0
        remaining = total - completed
        return Text(f"[Xong: {completed}/{total} | Còn: {remaining}/{total}]", style="green")

class ETAColumn(ProgressColumn):
    def render(self, task):
        from rich.text import Text
        import datetime
        completed = task.completed
        if completed == 0 or task.elapsed is None:
            return Text("-:--:--", style="cyan")
        speed = completed / task.elapsed
        remaining = task.total - completed
        if speed > 0:
            eta_seconds = remaining / speed
            eta_str = str(datetime.timedelta(seconds=int(eta_seconds)))
            return Text(eta_str, style="cyan")
        return Text("-:--:--", style="cyan")

class SpeedColumn(ProgressColumn):
    def render(self, task):
        from rich.text import Text
        completed = task.completed
        if completed == 0 or task.elapsed is None or task.elapsed == 0:
            return Text("[⚡ --.-s/ảnh]", style="yellow")
        speed = task.elapsed / completed
        return Text(f"[⚡ {speed:.1f}s/ảnh]", style="yellow")

console = Console()

def get_unique_images(image_paths):
    import hashlib
    from PIL import Image
    unique_paths = []
    seen_md5 = set()
    seen_dhash = []
    duplicates_count = 0
    
    # Sắp xếp danh sách file để ưu tiên giữ lại:
    # 1. Các file đã được đánh số (ưu tiên số nhỏ hơn)
    # 2. Các file cũ hơn (được thêm vào trước)
    def sort_key(p):
        base = os.path.splitext(os.path.basename(p))[0]
        try:
            mtime = os.path.getmtime(p)
        except Exception:
            mtime = float('inf')
            
        if base.isdigit():
            return (0, int(base), mtime)
        else:
            return (1, float('inf'), mtime)
            
    image_paths.sort(key=sort_key)
    
    def get_dhash(img_path):
        try:
            with Image.open(img_path) as img:
                img = img.convert('L').resize((9, 8), Image.Resampling.LANCZOS)
                pixels = list(img.getdata())
                diff = [pixels[r * 9 + c] > pixels[r * 9 + c + 1] for r in range(8) for c in range(8)]
                return sum([2 ** i for (i, v) in enumerate(diff) if v])
        except Exception:
            return None

    def hamming_distance(h1, h2):
        return bin(h1 ^ h2).count('1')
    
    with Progress(
        SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
        BarColumn(), TaskProgressColumn(), TimeElapsedColumn(), 
        TextColumn("⏳ ETA:"), TimeRemainingColumn(), console=console,
    ) as progress:
        task = progress.add_task("[cyan]Đang quét dHash & MD5 để diệt ảnh trùng lặp...", total=len(image_paths))
        for path in image_paths:
            try:
                with open(path, 'rb') as f:
                    file_md5 = hashlib.md5(f.read()).hexdigest()
                    
                if file_md5 in seen_md5:
                    is_duplicate = True
                else:
                    is_duplicate = False
                    file_dhash = get_dhash(path)
                    if file_dhash is not None:
                        for old_dhash in seen_dhash:
                            if hamming_distance(file_dhash, old_dhash) <= 2:
                                is_duplicate = True
                                break
                                
                if is_duplicate:
                    duplicates_count += 1
                    try:
                        os.remove(path)
                    except Exception:
                        pass
                else:
                    seen_md5.add(file_md5)
                    if file_dhash is not None:
                        seen_dhash.append(file_dhash)
                    unique_paths.append(path)
            except Exception:
                unique_paths.append(path)
            progress.advance(task)
            
    if duplicates_count > 0:
        console.print(f"[bold yellow]⚠️ Đã lọc bỏ và XÓA {duplicates_count} ảnh trùng lặp hoàn toàn về nội dung khỏi ổ cứng![/bold yellow]")
        
    return unique_paths

def run_wizard(input_dir, normalize_address=True):
    from rich.text import Text
    import time
    start_time = time.time()
    file_logs = []
    
    # Xóa dấu nháy đơn/kép nếu người dùng kéo thả thư mục vào terminal có sinh ra
    input_dir = clean_path(input_dir)

    if not input_dir or not os.path.isdir(input_dir):
        console.print(f"\n[bold red]❌ Lỗi:[/bold red] Thư mục '{input_dir}' không tồn tại hoặc đường dẫn không đúng.")
        Prompt.ask("[dim]Nhấn Enter để thử lại...[/dim]")
        return

    # Search for common image formats
    image_paths = []
    for ext in ('*.jpg', '*.jpeg', '*.png', '*.heic', '*.webp', '*.JPG', '*.JPEG', '*.PNG', '*.HEIC', '*.WEBP'):
        image_paths.extend(glob.glob(os.path.join(input_dir, ext)))

    if not image_paths:
        console.print("\n[bold red]❌ Thật tiếc, không tìm thấy file ảnh nào (.jpg, .png) trong thư mục này.[/bold red]")
        Prompt.ask("[dim]Nhấn Enter để thử lại...[/dim]")
        return

    total_raw_images = len(image_paths)
    image_paths = get_unique_images(image_paths)
    duplicates_count = total_raw_images - len(image_paths)

    console.print(f"\n[bold green]✅ Đã quét thư mục và chuẩn bị xử lý {len(image_paths)} file ảnh.[/bold green]")
    all_original_image_paths = image_paths.copy()
    
    # --- INCREMENTAL SCAN LOGIC ---
    incremental_scan = False
    old_records = []
    processed_images_set = set()
    max_renamed_idx = 0
    
    clean_input_dir = os.path.normpath(input_dir)
    possible_exports = [
        clean_input_dir + "_exports",
        os.path.join(clean_input_dir, "exports")
    ]
    
    all_old_excels = []
    for edir in possible_exports:
        if os.path.isdir(edir):
            excels = glob.glob(os.path.join(edir, "*.xlsx"))
            all_old_excels.extend([f for f in excels if not os.path.basename(f).startswith('~$')])
            
    latest_excel = None
    if all_old_excels:
        latest_excel = max(all_old_excels, key=os.path.getmtime)
        
    if latest_excel:
        if IN_COLAB:
            incremental_scan = True
        else:
            incremental_scan = Confirm.ask(f"\n[bold yellow]Phát hiện thư mục này đã từng được xử lý (có file {os.path.basename(latest_excel)}). Bạn muốn QUÉT NỐI TIẾP (chỉ quét ảnh mới ném vào) không? (Chọn No để quét lại từ đầu)[/bold yellow]", default=True)
    elif os.path.exists(os.path.join(input_dir, "original.zip")):
        if IN_COLAB:
            user_input = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        else:
            user_input = clean_path(Prompt.ask("\n[bold yellow]Phát hiện thư mục này đã từng được xử lý (có file original.zip) nhưng không tự động tìm thấy thư mục kết quả cũ.\n👉 Nếu bạn muốn QUÉT NỐI TIẾP, vui lòng kéo thả THƯ MỤC EXPORT cũ hoặc FILE EXCEL cũ vào đây (Nhấn Enter để tìm trong thư mục export mặc định của app, hoặc gõ 'n' để quét lại từ đầu)[/bold yellow]"))
            if not user_input or user_input.lower() == 'y':
                user_input = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
            
        if user_input and user_input.lower() != 'n':
            if os.path.isfile(user_input) and user_input.endswith('.xlsx'):
                latest_excel = user_input
                incremental_scan = True
            elif os.path.isdir(user_input):
                excels = glob.glob(os.path.join(user_input, "*.xlsx"))
                excels = [f for f in excels if not os.path.basename(f).startswith('~$')]
                if excels:
                    if len(excels) == 1:
                        latest_excel = excels[0]
                        incremental_scan = True
                    else:
                        excels.sort(key=os.path.getmtime, reverse=True)
                        console.print("\n[bold cyan]Tìm thấy nhiều file Excel, vui lòng chọn file bạn muốn dùng để làm mốc Quét nối tiếp:[/bold cyan]")
                        for i, f in enumerate(excels, 1):
                            mtime = datetime.datetime.fromtimestamp(os.path.getmtime(f)).strftime('%Y-%m-%d %H:%M:%S')
                            console.print(f"  [bold yellow][{i}][/bold yellow]. {os.path.basename(f)} [dim]({mtime})[/dim]")
                            
                        choice = Prompt.ask("\n[bold cyan]Nhập số thứ tự file Excel[/bold cyan] (Enter để chọn file mới nhất [1])", default="1").strip()
                        try:
                            choice_idx = int(choice) - 1
                            if 0 <= choice_idx < len(excels):
                                latest_excel = excels[choice_idx]
                                incremental_scan = True
                            else:
                                console.print("[red]❌ Lựa chọn không hợp lệ, hủy Quét Nối Tiếp.[/red]")
                        except ValueError:
                            console.print("[red]❌ Lựa chọn không hợp lệ, hủy Quét Nối Tiếp.[/red]")
                else:
                    console.print("[red]❌ Không tìm thấy file Excel nào trong thư mục bạn vừa nhập.[/red]")

    if incremental_scan and latest_excel:
        console.print(f"[cyan]Đang đọc file Excel cũ ({os.path.basename(latest_excel)}) để lọc ra các ảnh mới...[/cyan]")
        try:
            wb = openpyxl.load_workbook(latest_excel)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            col_idx = {name: i for i, name in enumerate(headers)}
            
            img_front_col = col_idx.get('Ảnh mặt trước CCCD/CC')
            img_back_col = col_idx.get('Ảnh mặt sau CCCD/CC')
            renamed_front_col = col_idx.get('Đổi tên Ảnh mặt trước CCCD/CC')
            renamed_back_col = col_idx.get('Đổi tên Ảnh mặt sau CCCD/CC')
            img_other_col = col_idx.get('Ảnh khác (SMS/Chụp màn hình/...)')
            renamed_other_col = col_idx.get('Đổi tên Ảnh khác')
            
            if img_front_col is None or img_back_col is None:
                console.print("[red]❌ File Excel cũ không có cột tên ảnh, không thể quét nối tiếp.[/red]")
                incremental_scan = False
            else:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    old_records.append({
                        'Họ tên': row[col_idx.get('Họ tên')] if 'Họ tên' in col_idx else '',
                        'CCCD': row[col_idx.get('CCCD')] if 'CCCD' in col_idx else '',
                        'CMND': row[col_idx.get('CMND')] if 'CMND' in col_idx else '',
                        'Giới tính': row[col_idx.get('Giới tính')] if 'Giới tính' in col_idx else '',
                        'Ngày sinh': row[col_idx.get('Ngày sinh')] if 'Ngày sinh' in col_idx else '',
                        'Nơi thường trú gốc': row[col_idx.get('Nơi thường trú gốc')] if 'Nơi thường trú gốc' in col_idx else '',
                        'Địa chỉ chuẩn hóa mới': row[col_idx.get('Địa chỉ chuẩn hóa mới')] if 'Địa chỉ chuẩn hóa mới' in col_idx else '',
                        'Ngày cấp CCCD': row[col_idx.get('Ngày cấp CCCD')] if 'Ngày cấp CCCD' in col_idx else '',
                        'Nơi cấp': row[col_idx.get('Nơi cấp')] if 'Nơi cấp' in col_idx else '',
                        'Ngày hết hạn': row[col_idx.get('Ngày hết hạn')] if 'Ngày hết hạn' in col_idx else '',
                        'Phân loại': row[col_idx.get('Phân loại')] if 'Phân loại' in col_idx else '',
                        'Ghi chú': row[col_idx.get('Ghi chú')] if 'Ghi chú' in col_idx else '',
                        'QR Raw': row[col_idx.get('QR Raw')] if 'QR Raw' in col_idx else '',
                        'Ảnh mặt trước CCCD/CC': row[img_front_col],
                        'Ảnh mặt sau CCCD/CC': row[img_back_col],
                        'Đổi tên Ảnh mặt trước CCCD/CC': row[renamed_front_col] if renamed_front_col is not None else '',
                        'Đổi tên Ảnh mặt sau CCCD/CC': row[renamed_back_col] if renamed_back_col is not None else '',
                        'Ảnh khác (SMS/Chụp màn hình/...)': row[img_other_col] if img_other_col is not None else '',
                        'Đổi tên Ảnh khác': row[renamed_other_col] if renamed_other_col is not None else ''
                    })
                    front = row[img_front_col]
                    back = row[img_back_col]
                    renamed_front = row[renamed_front_col] if renamed_front_col is not None else ''
                    renamed_back = row[renamed_back_col] if renamed_back_col is not None else ''
                    other = row[img_other_col] if img_other_col is not None else ''
                    renamed_other = row[renamed_other_col] if renamed_other_col is not None else ''
                    
                    if front: processed_images_set.add(str(front))
                    if back: processed_images_set.add(str(back))
                    if renamed_front: processed_images_set.add(str(renamed_front))
                    if renamed_back: processed_images_set.add(str(renamed_back))
                    if other:
                        for p in str(other).split(', '):
                            if p: processed_images_set.add(p)
                    if renamed_other:
                        for p in str(renamed_other).split(', '):
                            if p: processed_images_set.add(p)
                            
                # Đọc thêm các sheet phụ (nếu có) để không quét lại ảnh lỗi/trùng lặp
                if 'duplicate' in wb.sheetnames:
                    ws_dup = wb['duplicate']
                    for row in ws_dup.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 2 and row[1]:
                            processed_images_set.add(str(row[1]))
                            
                if 'Unknown' in wb.sheetnames:
                    ws_unk = wb['Unknown']
                    for row in ws_unk.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 2 and row[1]:
                            processed_images_set.add(str(row[1]))
                            
                if 'QR_scanned' in wb.sheetnames:
                    ws_qr = wb['QR_scanned']
                    for row in ws_qr.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 2 and row[1]:
                            processed_images_set.add(str(row[1]))
                            
                if 'OCR_scanned' in wb.sheetnames:
                    ws_ocr = wb['OCR_scanned']
                    for row in ws_ocr.iter_rows(min_row=2, values_only=True):
                        if len(row) >= 2 and row[1]:
                            processed_images_set.add(str(row[1]))
        except Exception as e:
            console.print(f"[red]❌ Lỗi đọc file Excel cũ: {e}[/red]")
            incremental_scan = False
                    
    realtime_csv = os.path.join(input_dir, "ket_qua_scan_tam_thoi.csv")
    realtime_log = os.path.join(input_dir, "log_scan_tam_thoi.txt")
    crashed_items = []
    
    if os.path.exists(realtime_csv):
        try:
            import csv
            with open(realtime_csv, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    crashed_items.append(row)
                    if row.get('Image Path'):
                        processed_images_set.add(row['Image Path'])
            if crashed_items:
                console.print(f"[bold green]✅ Đã phục hồi {len(crashed_items)} kết quả scan bị gián đoạn từ {os.path.basename(realtime_csv)}.[/bold green]")
                incremental_scan = True
                # Khôi phục file log tạm
                if os.path.exists(realtime_log):
                    with open(realtime_log, 'r', encoding='utf-8') as flog:
                        file_logs.extend(flog.read().splitlines())
        except Exception as e:
            pass

    if incremental_scan:
        new_image_paths = [p for p in image_paths if os.path.basename(p) not in processed_images_set]
        
        # Chỉ lấy max_renamed_idx từ danh sách ảnh ĐÃ XỬ LÝ thành công trong file Excel cũ
        for p in processed_images_set:
            base = os.path.splitext(os.path.basename(p))[0]
            if base.isdigit() and int(base) > max_renamed_idx:
                max_renamed_idx = int(base)
                
        if not new_image_paths:
            if crashed_items:
                console.print("\n[bold green]✅ Không tìm thấy ảnh mới nào. Đang tiến hành xuất dữ liệu đã phục hồi ra Excel...[/bold green]")
                image_paths = []
            else:
                console.print("\n[bold green]✅ Không tìm thấy ảnh mới nào được chép thêm vào. Kết thúc quá trình quét nối tiếp![/bold green]")
                return
        else:
            console.print(f"[bold green]✅ Đã tự động lọc ra [yellow]{len(new_image_paths)}[/yellow] ảnh mới cần xử lý.[/bold green]")
            image_paths = new_image_paths

    # --- AUTO BACKUP AND RENAME LOGIC ---
    zip_path = os.path.join(input_dir, "original.zip")
    
    if not os.path.exists(zip_path) or incremental_scan:
        action_word = "bổ sung" if incremental_scan else "gốc"
        mode = 'a' if incremental_scan else 'w'
        
        files_to_process = []
        for p in image_paths:
            base = os.path.splitext(os.path.basename(p))[0]
            if base.isdigit():
                pass # Tuyệt đối không bao giờ đổi tên lại những file đã được đánh số (dù là do crash hay từ lần quét cũ)
            else:
                files_to_process.append(p)

        if files_to_process:
            start_idx = max_renamed_idx + 1 if incremental_scan else 1
            
            console.print(f"[cyan]📦 Đang nén {action_word} {len(files_to_process)} file ảnh vào original.zip...[/cyan]")
            try:
                with zipfile.ZipFile(zip_path, mode, zipfile.ZIP_DEFLATED) as zipf:
                    existing_in_zip = set(zipf.namelist()) if mode == 'a' else set()
                    for file_path in files_to_process:
                        bname = os.path.basename(file_path)
                        if bname not in existing_in_zip:
                            zipf.write(file_path, bname)
                            existing_in_zip.add(bname)
                
                if incremental_scan:
                    console.print(f"[cyan]🔄 Đang đổi tên các file ảnh mới nối tiếp (từ số {start_idx})...[/cyan]")
                else:
                    console.print("[cyan]🔄 Đang đổi tên các file ảnh theo số thứ tự...[/cyan]")
                    
                # Bước 1: Đổi tên thành tên tạm (để tránh ghi đè ngẫu nhiên)
                temp_paths = []
                for file_path in files_to_process:
                    ext = os.path.splitext(file_path)[1]
                    temp_name = f"temp_{uuid.uuid4().hex[:8]}{ext}"
                    temp_path = os.path.join(input_dir, temp_name)
                    os.rename(file_path, temp_path)
                    temp_paths.append((temp_path, ext))
                    
                # Bước 2: Đổi tên thành số thứ tự
                new_image_paths = []
                for i, (temp_path, ext) in enumerate(temp_paths, start_idx):
                    new_name = f"{i}{ext}"
                    new_path = os.path.join(input_dir, new_name)
                    os.rename(temp_path, new_path)
                    new_image_paths.append(new_path)
                    
                # Cập nhật lại list image_paths
                final_image_paths = []
                renamed_idx = 0
                for p in image_paths:
                    if p in files_to_process:
                        final_image_paths.append(new_image_paths[renamed_idx])
                        renamed_idx += 1
                    else:
                        final_image_paths.append(p)
                image_paths = final_image_paths
                console.print("[bold green]✅ Đã sao lưu và đổi tên thành công![/bold green]")
            except Exception as e:
                console.print(f"[bold red]❌ Lỗi trong quá trình sao lưu/đổi tên: {e}[/bold red]")
                return
        else:
            console.print("[yellow]⚠️ Các ảnh mới đều đã được đánh số ở lần chạy trước, bỏ qua bước sao lưu và đổi tên.[/yellow]")
    else:
        console.print("[yellow]⚠️ Bỏ qua bước sao lưu và đổi tên do file original.zip đã tồn tại trong thư mục này.[/yellow]")
    # ------------------------------------
    
    # Cấu hình luồng xử lý
    num_threads_input = Prompt.ask("\n[cyan]Nhập số luồng xử lý ảnh song song[/cyan] (Enter để mặc định là 4)", default="4").strip()
    try:
        num_threads = int(num_threads_input) if num_threads_input else 4
    except ValueError:
        console.print("[yellow]⚠️ Giá trị không hợp lệ, sử dụng mặc định: 4 luồng.[/yellow]")
        num_threads = 4

    api_threads = 4
    if normalize_address:
        api_threads_input = Prompt.ask("[cyan]Nhập số luồng gọi API địa chỉ song song[/cyan] (Enter để mặc định là 4)", default="4").strip()
        try:
            api_threads = int(api_threads_input) if api_threads_input else 4
        except ValueError:
            console.print("[yellow]Số luồng không hợp lệ, dùng mặc định 4[/yellow]")
            api_threads = 4

    # Wizard confirmation
    if IN_COLAB:
        confirm = True
    else:
        confirm = Confirm.ask("\n[bold yellow]Bạn có muốn bắt đầu xử lý ngay bây giờ không?[/bold yellow]", default=True)
    if not confirm:
        console.print("[yellow]Đã hủy quá trình.[/yellow]")
        return

    console.print("\n")
    console.print(Panel(f"[bold cyan]🚀 BẮT ĐẦU XỬ LÝ {len(image_paths)} ẢNH VỚI {num_threads} LUỒNG...[/bold cyan]", border_style="green"))

    # Đặt lại start_time để loại bỏ thời gian chờ người dùng nhập lệnh
    start_time = time.time()

    temp_rotated_dir = os.path.join(tempfile.gettempdir(), f"cccd_exports_{uuid.uuid4().hex[:8]}")
    os.makedirs(temp_rotated_dir, exist_ok=True)

    processed_data = []
    seen_cccds = set()
    
    def process_single_image(img_path):
        import time
        if not IN_COLAB:
            console.print(f"[dim]⏳ Bắt đầu đưa vào AI: {os.path.basename(img_path)}...[/dim]")
        t0 = time.time()
        
        qr_string, engine, err, img, qr_rotated_img = extract_qr_data(img_path)
        
        log_msgs = []
        if qr_string:
            log_msgs.append(f"[green]✅ [Đã quét mã QR bằng {engine}]:[/green] {qr_string}")
            
        row_data = {
            'Họ tên': '', 'CCCD': '', 'CMND': '', 'Giới tính': '',
            'Ngày sinh': '', 'Nơi thường trú gốc': '', 'Địa chỉ chuẩn hóa mới': '',
            'Ngày cấp CCCD': '', 'Nơi cấp': '', 'Ngày hết hạn': '', 'Phân loại': '', 'Ghi chú': '', 'QR Raw': '',
            'Image Path': os.path.basename(img_path),
            'Full Image Path': img_path,
            'Scan Type': 'error'
        }
        
        if qr_rotated_img is not None:
            # Lưu lại ảnh QR đã xoay chuẩn vào thư mục tạm
            temp_path = os.path.join(temp_rotated_dir, os.path.basename(img_path))
            cv2.imwrite(temp_path, qr_rotated_img)
            row_data['Full Image Path'] = temp_path
            img = qr_rotated_img # Cập nhật img để nếu fallback sang OCR thì OCR lấy luôn ảnh đã xoay
        notes = []

        if qr_string:
            row_data['Scan Type'] = 'QR_scanned'
            row_data['QR Raw'] = qr_string
            extracted, validation_notes = process_qr_string(qr_string)
            row_data.update(extracted)
            notes.extend(validation_notes)
        else:
            if err:
                notes.append(err)
            
            # Fallback to OCR
            if img is not None:
                ocr_msg = f"[yellow]⚠️ Không đọc được QR, đang thử quét OCR... (Có thể mất 30-60s/ảnh trên CPU, vui lòng đợi)[/yellow]"
                log_msgs.append(ocr_msg)
                if not IN_COLAB:
                    console.print(f"  {ocr_msg}")
                
                ocr_data, ocr_note, rotated_img = extract_ocr_data(img)
                
                if rotated_img is not None:
                    # Lưu lại ảnh đã xoay chuẩn vào thư mục tạm thay vì ghi đè file gốc
                    temp_path = os.path.join(temp_rotated_dir, os.path.basename(img_path))
                    cv2.imwrite(temp_path, rotated_img)
                    row_data['Full Image Path'] = temp_path
                
                # In thông tin OCR ra màn hình tùy theo mặt thẻ
                parts = []
                side = ocr_data.get('OCR Side')
                if side: parts.append(f"[{side}]")
                parts.append(f"CCCD: {ocr_data.get('CCCD') or '[Trống]'}")
                parts.append(f"Tên: {ocr_data.get('Họ tên') or '[Trống]'}")
                parts.append(f"NS: {ocr_data.get('Ngày sinh') or '[Trống]'}")
                parts.append(f"GT: {ocr_data.get('Giới tính') or '[Trống]'}")
                parts.append(f"Ngày cấp: {ocr_data.get('Ngày cấp CCCD') or '[Trống]'}")
                
                if side == 'Front':
                    addr = ocr_data.get('Nơi thường trú gốc') or '[Trống]'
                    parts.append(f"Địa chỉ: {addr}")
                elif side == 'Back':
                    back_raw = ocr_data.get('back_side_raw', {})
                    mrz_lines = back_raw.get('mrz_lines', [])
                    if mrz_lines:
                        parts.append(f"MRZ ({len(mrz_lines)} dòng): {' | '.join(mrz_lines[:3])}")
                    else:
                        parts.append(f"MRZ: [Trống]")
                else:
                    addr = ocr_data.get('Nơi thường trú gốc') or '[Trống]'
                    parts.append(f"Địa chỉ: {addr}")
                    back_raw = ocr_data.get('back_side_raw', {})
                    mrz_lines = back_raw.get('mrz_lines', [])
                    if mrz_lines:
                        parts.append(f"MRZ ({len(mrz_lines)} dòng): {' | '.join(mrz_lines[:3])}")
                
                ocr_print_info = ", ".join(parts)
                log_msgs.append(f"[blue]ℹ️ Kết quả OCR:[/blue] {ocr_print_info} | Note: {ocr_note}")
                
                if DEBUG_MODE and ocr_data.get('Raw Text'):
                    log_msgs.append(f"[magenta]🐛 DEBUG RAW OCR TEXT:\n{ocr_data['Raw Text']}[/magenta]")
                
                # In ra màn hình cảnh báo chói lóa nếu có
                if "chói/lóa" in ocr_note:
                    log_msgs.append(f"[red]⚠️ Ảnh bị chói/lóa sáng hoặc quá mờ không thể xử lý tốt[/red]")
                
                if ocr_data.get('CCCD'):
                    row_data['Scan Type'] = 'OCR_scanned'
                row_data.update(ocr_data)
                notes.append(ocr_note)
                
        row_data['Nơi cấp'] = get_place_of_issue(row_data.get('QR Raw', ''))
        row_data['Ngày hết hạn'] = calculate_expiry_date(row_data.get('Ngày sinh', ''))
        row_data['Phân loại'] = get_card_type(row_data.get('QR Raw', ''))
        row_data['Ghi chú'] = '; '.join(notes)
        row_data['_processing_time'] = time.time() - t0
        return row_data, log_msgs

    import re
    # Grouping and Merging Logic
    records = {} # mapping CCCD -> record
    
    # Process images in parallel and collect all returned raw data
    extracted_items = crashed_items.copy() if 'crashed_items' in locals() else []
    
    progress = Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(),
        TaskProgressColumn(),
        CountColumn(),
        TimeElapsedColumn(),
        TextColumn("⏳ ETA:"),
        ETAColumn(),
        SpeedColumn(),
        console=console,
        refresh_per_second=REFRESH_RATE,
    )
    with progress:
        task_id = progress.add_task("[cyan]Đang quét ảnh...", total=len(image_paths))
        executor = concurrent.futures.ThreadPoolExecutor(max_workers=num_threads)
        future_to_img = {executor.submit(process_single_image, path): path for path in image_paths}
        for future in concurrent.futures.as_completed(future_to_img):
            img_path = future_to_img[future]
            try:
                row_data, log_msgs = future.result(timeout=300)
            except concurrent.futures.TimeoutError:
                err = f"❌ Lỗi: Ảnh {os.path.basename(img_path)} làm treo AI quá 300s (5 phút). Đã tự động bỏ qua!"
                progress.console.print(f"[bold red]{err}[/bold red]")
                file_logs.append(err)
                progress.advance(task_id)
                continue
            except Exception as exc:
                err = f"❌ Lỗi khi xử lý ảnh {os.path.basename(img_path)}: {exc}"
                progress.console.print(f"[bold red]{err}[/bold red]")
                file_logs.append(err)
                progress.advance(task_id)
                continue
                
            extracted_items.append(row_data)
            
            # Ghi realtime backup
            try:
                import csv
                file_exists = os.path.exists(realtime_csv) and os.path.getsize(realtime_csv) > 0
                csv_keys = ['Họ tên', 'CCCD', 'CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Địa chỉ chuẩn hóa mới', 'Ngày cấp CCCD', 'Nơi cấp', 'Ngày hết hạn', 'Phân loại', 'Ghi chú', 'QR Raw', 'Image Path', 'Full Image Path', 'Scan Type', 'OCR Side', 'Raw Text Upper', 'Raw Text']
                with open(realtime_csv, 'a', newline='', encoding='utf-8-sig') as f:
                    writer = csv.DictWriter(f, fieldnames=csv_keys, extrasaction='ignore')
                    if not file_exists:
                        writer.writeheader()
                    writer.writerow(row_data)
            except Exception: pass
            
            try:
                with open(realtime_log, 'a', encoding='utf-8') as f:
                    f.write(f"[{os.path.basename(img_path)}] - {img_path}\n")
                    for msg in log_msgs:
                        f.write("  " + Text.from_markup(msg).plain + "\n")
            except Exception: pass
            
            # Print logs for this image above the progress bar
            p_time = row_data.get('_processing_time', 0)
            if not IN_COLAB:
                progress.console.print(f"[bold][{os.path.basename(img_path)}][/bold] - [dim]{img_path}[/dim] - [yellow]Timing {p_time:.1f}s[/yellow]")
            file_logs.append(f"[{os.path.basename(img_path)}] - {img_path} - Timing {p_time:.1f}s")
            for msg in log_msgs:
                if not IN_COLAB:
                    progress.console.print(f"  {msg}")
                file_logs.append("  " + Text.from_markup(msg).plain)
                
            progress.advance(task_id)
    console.print(Panel(f"[bold cyan]🔄 BẮT ĐẦU GỘP DỮ LIỆU...[/bold cyan]", border_style="green"))
    for item in extracted_items:
        cccd = item.get('CCCD')
        if not cccd:
            # Lưu lại bằng 1 ID giả để không bị mất ảnh này trong báo cáo (đẩy vào Sheet Review)
            cccd = f"UNKNOWN_{item.get('Image Path', 'NoName')}"
            item['CCCD'] = cccd
        
        is_new_record = False
        if cccd not in records:
            records[cccd] = {
                'index': len(records) + 1,
                'Họ tên': '', 'CCCD': cccd, 'CMND': '', 'Giới tính': '',
                'Ngày sinh': '', 'Nơi thường trú gốc': '', 'Địa chỉ chuẩn hóa mới': '',
                'Ngày cấp CCCD': '', 'Nơi cấp': '', 'Ngày hết hạn': '', 'Phân loại': '', 'Ghi chú': [], 'QR Raw': '',
                'Ảnh mặt trước CCCD/CC': '',
                'Ảnh mặt sau CCCD/CC': '',
                'Đổi tên Ảnh mặt trước CCCD/CC': '',
                'Đổi tên Ảnh mặt sau CCCD/CC': '',
                'Ảnh khác (SMS/Chụp màn hình/...)': '',
                'Đổi tên Ảnh khác': '',
                'Full Image Path Front': '',
                'Full Image Path Back': '',
                'OCR Image Path Front': '',
                'Full OCR Image Path Front': '',
                'OCR Image Path Back': '',
                'Full OCR Image Path Back': '',
                'OCR Image Path Unknown': '',
                'Full OCR Image Path Unknown': '',
                'has_qr_data': False,
                'has_ocr_data': False,
                'has_cong_dan_front': False,
                'has_address_front': False,
                'has_address_back': False,
                'has_cuc_truong_back': False,
                'has_bo_cong_an_back': False
            }
            is_new_record = True
        
        record = records[cccd]
        person_idx = record['index']
        
        is_qr = bool(item.get('QR Raw'))
        if is_qr:
            id_type = 'CCCD' if len(str(cccd)) == 12 else 'CMND'
            if not is_new_record:
                if record['has_ocr_data'] and not record['has_qr_data']:
                    console.print(f"   [yellow]→ [Người {person_idx}] [GỘP DỮ LIỆU][/yellow] GHI ĐÈ thông tin từ ảnh {item['Image Path']} (Đọc mã QR) lên thông tin OCR trước đó của {id_type} {cccd}")
                elif record['has_qr_data']:
                    console.print(f"   [yellow]→ [Người {person_idx}] [GỘP DỮ LIỆU][/yellow] Bỏ qua thông tin từ ảnh {item['Image Path']} vì đã quét mã QR thành công trước đó cho {id_type} {cccd}")
            
            record['has_qr_data'] = True
            # Overwrite text fields with QR accuracy
            for k in ['Họ tên', 'CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD', 'Nơi cấp', 'Ngày hết hạn', 'Phân loại', 'QR Raw']:
                if item.get(k): record[k] = item[k]
            
            # Determine card type and side
            fields = item['QR Raw'].split('|')
            if len(fields) == 7:
                if record['Ảnh mặt trước CCCD/CC']:
                    existing = record.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    record['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {item['Image Path']}".strip(', ')
                else:
                    record['Ảnh mặt trước CCCD/CC'] = item['Image Path']
                    record['Full Image Path Front'] = item['Full Image Path']
            elif len(fields) >= 10:
                if record['Ảnh mặt sau CCCD/CC']:
                    existing = record.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    record['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {item['Image Path']}".strip(', ')
                else:
                    record['Ảnh mặt sau CCCD/CC'] = item['Image Path']
                    record['Full Image Path Back'] = item['Full Image Path']
            
            if item.get('Ghi chú'):
                record['Ghi chú'].append(item['Ghi chú'])
        else: # OCR
            id_type = 'CCCD' if len(str(cccd)) == 12 else 'CMND'
            if not is_new_record:
                if record['has_qr_data']:
                    console.print(f"   [yellow]→ [Người {person_idx}] [GỘP DỮ LIỆU][/yellow] Bỏ qua thông tin trùng lặp từ ảnh {item['Image Path']} (OCR) vì đã có dữ liệu QR chuẩn xác của {id_type} {cccd}")
                elif record['has_ocr_data']:
                    console.print(f"   [yellow]→ [Người {person_idx}] [GỘP DỮ LIỆU][/yellow] Bỏ qua thông tin trùng lặp từ ảnh {item['Image Path']} (OCR) vì đã xử lý ảnh OCR trước đó cho {id_type} {cccd}")
            
            record['has_ocr_data'] = True
            
            # Fill empty text fields
            for k in ['Họ tên', 'CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD']:
                if item.get(k) and not record.get(k): record[k] = item[k]
                
            raw_text = item.get('Raw Text Upper', '')
            if item.get('OCR Side') == 'Front':
                if record.get('OCR Image Path Front') or record.get('Ảnh mặt trước CCCD/CC'):
                    existing = record.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    record['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {item['Image Path']}".strip(', ')
                else:
                    record['OCR Image Path Front'] = item['Image Path']
                    record['Full OCR Image Path Front'] = item['Full Image Path']
                if "CÔNG DÂN" in raw_text: record['has_cong_dan_front'] = True
                if item.get('Nơi thường trú gốc'): record['has_address_front'] = True
            elif item.get('OCR Side') == 'Back':
                if record.get('OCR Image Path Back') or record.get('Ảnh mặt sau CCCD/CC'):
                    existing = record.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    record['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {item['Image Path']}".strip(', ')
                else:
                    record['OCR Image Path Back'] = item['Image Path']
                    record['Full OCR Image Path Back'] = item['Full Image Path']
                if "CỤC TRƯỞNG" in raw_text: record['has_cuc_truong_back'] = True
                if "BỘ CÔNG AN" in raw_text: record['has_bo_cong_an_back'] = True
                if item.get('Nơi thường trú gốc'): record['has_address_back'] = True
            else:
                if record.get('OCR Image Path Unknown'):
                    existing = record.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    record['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {item['Image Path']}".strip(', ')
                else:
                    record['OCR Image Path Unknown'] = item['Image Path']
                    record['Full OCR Image Path Unknown'] = item['Full Image Path']
            
            if item.get('Ghi chú'):
                record['Ghi chú'].append(item['Ghi chú'])

    # ---------- GENERAL FUZZY MATCH: GỘP BẢN GHI OCR SAI LỆCH SỐ (CÙNG TÊN + TRÙNG DOB HOẶC CẬN CCCD) ----------
    import unicodedata
    import re

    def _norm_for_match(text):
        """Chuẩn hóa để so khớp: xóa dấu, in hoa, rút gọn khoảng trắng."""
        if not text: return ''
        text = text.upper().strip()
        nfkd = unicodedata.normalize('NFKD', text)
        return ' '.join(''.join(c for c in nfkd if not unicodedata.combining(c)).split())

    def _is_similar_cccd(c1, c2):
        if not c1 or not c2: return False
        c1_clean = re.sub(r'\D', '', str(c1))
        c2_clean = re.sub(r'\D', '', str(c2))
        if not c1_clean or not c2_clean: return False
        if len(c1_clean) == 12 and len(c2_clean) == 12:
            diffs = sum(1 for a, b in zip(c1_clean, c2_clean) if a != b)
            return diffs <= 3
        return c1_clean == c2_clean

    def _is_similar_name(n1, n2):
        if not n1 or not n2: return False
        import difflib
        ratio = difflib.SequenceMatcher(None, n1, n2).ratio()
        if ratio < 0.80:
            return False
        # Kiểm tra sự tương đồng của tên gọi (từ cuối cùng của họ tên, ví dụ Anh vs Sâm)
        w1 = n1.split()[-1] if n1.split() else ''
        w2 = n2.split()[-1] if n2.split() else ''
        if not w1 or not w2:
            return False
        w_ratio = difflib.SequenceMatcher(None, w1, w2).ratio()
        return w_ratio >= 0.50

    def _is_similar_dob(d1, d2):
        if not d1 or not d2: return False
        d1 = d1.strip()
        d2 = d2.strip()
        if d1 == d2: return True
        p1 = d1.split('/')
        p2 = d2.split('/')
        if len(p1) == 3 and len(p2) == 3:
            if p1[1] == p2[1] and p1[2] == p2[2]:
                return True
        return False

    def _is_invalid_cccd_placeholder(cccd):
        if not cccd: return True
        cccd_clean = re.sub(r'\D', '', str(cccd))
        if len(cccd_clean) != 12: return True
        if cccd_clean.startswith('000'): return True
        return False

    secondary_cccds = [cccd for cccd, rec in list(records.items()) if not rec['has_qr_data']]
    for sec_cccd in secondary_cccds:
        if sec_cccd not in records:
            continue
        sec_rec = records[sec_cccd]
        sec_name = _norm_for_match(sec_rec.get('Họ tên', ''))
        sec_dob = sec_rec.get('Ngày sinh', '')
        
        if not sec_name:
            continue
            
        # Tìm tất cả các ứng cử viên trùng/cận tên trong database
        candidates = []
        for target_cccd, target_rec in records.items():
            if target_cccd == sec_cccd:
                continue
            # Không gộp số CCCD hợp lệ vào số CCCD rác/placeholder (ví dụ 000000000119)
            if _is_invalid_cccd_placeholder(target_cccd) and not _is_invalid_cccd_placeholder(sec_cccd):
                continue
                
            # Ngăn gộp nếu giới tính khai báo khác nhau (tránh nhầm lẫn Nam/Nữ)
            s_gen = sec_rec.get('Giới tính', '').strip().lower()
            t_gen = target_rec.get('Giới tính', '').strip().lower()
            if s_gen and t_gen and s_gen != t_gen:
                continue
                
            target_name = _norm_for_match(target_rec.get('Họ tên', ''))
            if _is_similar_name(sec_name, target_name):
                candidates.append(target_cccd)
                
        best_target_cccd = None
        
        # Trường hợp 1: Nếu cả database chỉ có DUY NHẤT 1 người trùng tên, tự động gộp (vì không thể nhầm lẫn với ai)
        if len(candidates) == 1:
            best_target_cccd = candidates[0]
        # Trường hợp 2: Nếu có nhiều người trùng tên, bắt buộc phải lọc kỹ hơn bằng DOB hoặc số CCCD
        elif len(candidates) > 1:
            for target_cccd in candidates:
                target_rec = records[target_cccd]
                target_dob = target_rec.get('Ngày sinh', '')
                
                match_dob = (sec_dob and target_dob and _is_similar_dob(sec_dob, target_dob))
                match_cccd = _is_similar_cccd(sec_cccd, target_cccd)
                
                if match_dob or match_cccd:
                    if not best_target_cccd:
                        best_target_cccd = target_cccd
                    else:
                        curr_best = records[best_target_cccd]
                        if target_rec['has_qr_data'] and not curr_best['has_qr_data']:
                            best_target_cccd = target_cccd
                        elif _is_invalid_cccd_placeholder(best_target_cccd) and not _is_invalid_cccd_placeholder(target_cccd):
                            best_target_cccd = target_cccd
                        
        if best_target_cccd:
            target_rec = records[best_target_cccd]
            console.print(f"   [bold green]→ [GENERAL FUZZY MATCH][/bold green] Ghép bản ghi OCR {sec_cccd} ({sec_rec.get('Họ tên')}) vào bản ghi {best_target_cccd} (Trùng tên và trùng DOB/cận CCCD).")
            
            # Gộp mặt trước
            if not target_rec.get('Ảnh mặt trước CCCD/CC'):
                if sec_rec.get('Ảnh mặt trước CCCD/CC'):
                    target_rec['Ảnh mặt trước CCCD/CC'] = sec_rec['Ảnh mặt trước CCCD/CC']
                    target_rec['Full Image Path Front'] = sec_rec.get('Full Image Path Front', '')
                elif sec_rec.get('OCR Image Path Front'):
                    target_rec['Ảnh mặt trước CCCD/CC'] = sec_rec['OCR Image Path Front']
                    target_rec['Full Image Path Front'] = sec_rec.get('Full OCR Image Path Front', '')
            else:
                sec_front = sec_rec.get('Ảnh mặt trước CCCD/CC') or sec_rec.get('OCR Image Path Front')
                if sec_front:
                    existing = target_rec.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    target_rec['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {sec_front}".strip(', ')
                    
            # Gộp mặt sau
            if not target_rec.get('Ảnh mặt sau CCCD/CC'):
                if sec_rec.get('Ảnh mặt sau CCCD/CC'):
                    target_rec['Ảnh mặt sau CCCD/CC'] = sec_rec['Ảnh mặt sau CCCD/CC']
                    target_rec['Full Image Path Back'] = sec_rec.get('Full Image Path Back', '')
                elif sec_rec.get('OCR Image Path Back'):
                    target_rec['Ảnh mặt sau CCCD/CC'] = sec_rec['OCR Image Path Back']
                    target_rec['Full Image Path Back'] = sec_rec.get('Full OCR Image Path Back', '')
            else:
                sec_back = sec_rec.get('Ảnh mặt sau CCCD/CC') or sec_rec.get('OCR Image Path Back')
                if sec_back:
                    existing = target_rec.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    target_rec['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {sec_back}".strip(', ')
            
            # Gộp Ảnh khác
            sec_other = sec_rec.get('Ảnh khác (SMS/Chụp màn hình/...)') or sec_rec.get('OCR Image Path Unknown')
            if sec_other:
                existing = target_rec.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                target_rec['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {sec_other}".strip(', ')
                
            # Bổ sung thông tin văn bản còn thiếu
            for k in ['CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD', 'Nơi cấp', 'Ngày hết hạn', 'Phân loại', 'QR Raw']:
                if sec_rec.get(k) and not target_rec.get(k):
                    target_rec[k] = sec_rec[k]
                    
            for flag in ['has_qr_data', 'has_ocr_data', 'has_cong_dan_front', 'has_address_front', 'has_address_back', 'has_cuc_truong_back', 'has_bo_cong_an_back']:
                if sec_rec.get(flag):
                    target_rec[flag] = True
                    
            if sec_rec.get('Ghi chú'):
                if isinstance(sec_rec['Ghi chú'], list):
                    target_rec['Ghi chú'].extend(sec_rec['Ghi chú'])
                else:
                    target_rec['Ghi chú'].append(str(sec_rec['Ghi chú']))
                    
            del records[sec_cccd]

    # ---------- FUZZY MATCH: GỘP MẶT SAU OCR VÀO ĐÚNG BẢN GHI (2/3 TRƯỜNG) ----------
    # Mặt sau quét bằng OCR có thể đọc sai CCCD (ví dụ: 086... → 080...)
    # Nếu 2/3 trong {CCCD, Họ tên không dấu, Ngày cấp} khớp với bản ghi đã có → ghép vào đó
    import unicodedata

    def _norm_for_match(text):
        """Chuẩn hóa để so khớp: xóa dấu, in hoa, rút gọn khoảng trắng."""
        if not text: return ''
        text = text.upper().strip()
        nfkd = unicodedata.normalize('NFKD', text)
        return ' '.join(''.join(c for c in nfkd if not unicodedata.combining(c)).split())

    # Tìm các bản ghi "mặt sau mồ côi": chỉ có OCR Back, không có QR, không có mặt trước
    orphan_cccds = [
        cccd for cccd, rec in records.items()
        if (not rec['has_qr_data']
            and (rec.get('OCR Image Path Back') or rec.get('Full OCR Image Path Back'))
            and not rec.get('Full Image Path Front')
            and not rec.get('OCR Image Path Front'))
    ]

    for orphan_cccd in orphan_cccds:
        orphan = records[orphan_cccd]
        o_cccd = orphan.get('CCCD', '')
        o_name = _norm_for_match(orphan.get('Họ tên', ''))
        o_date = orphan.get('Ngày cấp CCCD', '')

        best_cccd = None
        for r_cccd, r_rec in records.items():
            if r_cccd == orphan_cccd:
                continue
            # Chỉ so với bản ghi đã có mặt trước hoặc QR
            if not (r_rec['has_qr_data']
                    or r_rec.get('Full Image Path Front')
                    or r_rec.get('OCR Image Path Front')):
                continue
            r_name = _norm_for_match(r_rec.get('Họ tên', ''))
            r_date = r_rec.get('Ngày cấp CCCD', '')

            score = 0
            if o_cccd and o_cccd == r_cccd:              score += 1
            if o_name and r_name and o_name == r_name:   score += 1
            if o_date and r_date and o_date == r_date:   score += 1
            if score >= 2:
                best_cccd = r_cccd
                break

        # Tùy chọn 2: Nếu chưa tìm được, kiểm tra 8 số liên tiếp hoặc khớp ngày cấp cho bản ghi đang thiếu mặt sau
        if not best_cccd:
            for r_cccd, r_rec in records.items():
                if r_cccd == orphan_cccd:
                    continue
                if not (r_rec['has_qr_data']
                        or r_rec.get('Full Image Path Front')
                        or r_rec.get('OCR Image Path Front')):
                    continue
                # Bản ghi đích phải thiếu mặt sau
                if r_rec.get('OCR Image Path Back') or r_rec.get('Full OCR Image Path Back') or r_rec.get('Ảnh mặt sau CCCD/CC') or r_rec.get('Full Image Path Back'):
                    continue
                
                match_8_digits = False
                if len(o_cccd) >= 8 and len(r_cccd) >= 8:
                    for i in range(len(r_cccd) - 7):
                        if r_cccd[i:i+8] in o_cccd:
                            match_8_digits = True
                            break
                            
                r_date = r_rec.get('Ngày cấp CCCD', '')
                match_date = (o_date and r_date and o_date == r_date)
                
                if match_8_digits or match_date:
                    best_cccd = r_cccd
                    break

        if best_cccd:
            target = records[best_cccd]
            # Ghép back-side image vào bản ghi đúng
            if not target.get('OCR Image Path Back'):
                target['OCR Image Path Back'] = orphan.get('OCR Image Path Back', '')
            if not target.get('Full OCR Image Path Back'):
                target['Full OCR Image Path Back'] = orphan.get('Full OCR Image Path Back', '')
            # Bổ sung thông tin còn thiếu từ OCR mặt sau
            for k in ['Họ tên', 'Ngày cấp CCCD']:
                if orphan.get(k) and not target.get(k):
                    target[k] = orphan[k]
            console.print(
                f"   [bold green]→ [FUZZY MATCH][/bold green] Ghép mặt sau (CCCD OCR: {orphan_cccd}) "
                f"vào bản ghi {best_cccd} (khớp 2/3 trường: tên/ngày cấp/cccd)."
            )
            del records[orphan_cccd]

    # ---------- FUZZY MATCH: GỘP ẢNH KHÁC VÀO ĐÚNG BẢN GHI (2/3 TRƯỜNG) ----------
    orphan_others = [
        cccd for cccd, rec in records.items()
        if (not rec['has_qr_data']
            and not rec['has_cong_dan_front']
            and not rec['has_address_front']
            and not rec['has_cuc_truong_back']
            and not rec['has_address_back']
            and not rec['has_bo_cong_an_back'])
    ]

    for orphan_cccd in orphan_others:
        orphan = records[orphan_cccd]
        o_cccd = orphan.get('CCCD', '')
        o_name = _norm_for_match(orphan.get('Họ tên', ''))
        o_dob = orphan.get('Ngày sinh', '')

        best_cccd = None
        for r_cccd, r_rec in records.items():
            if r_cccd == orphan_cccd:
                continue
            # Chỉ so với bản ghi đã có mặt trước hoặc QR
            if not (r_rec['has_qr_data'] or r_rec.get('Full Image Path Front') or r_rec.get('OCR Image Path Front') or r_rec['has_cong_dan_front'] or r_rec['has_address_front']):
                continue
                
            r_name = _norm_for_match(r_rec.get('Họ tên', ''))
            r_dob = r_rec.get('Ngày sinh', '')

            score = 0
            if o_cccd and o_cccd == r_cccd: score += 1
            if o_name and r_name and o_name == r_name: score += 1
            if o_dob and r_dob and o_dob == r_dob: score += 1
            
            if score >= 2:
                best_cccd = r_cccd
                break

        if best_cccd:
            target = records[best_cccd]
            img_path = orphan.get('OCR Image Path Unknown') or orphan.get('OCR Image Path Front') or orphan.get('OCR Image Path Back') or ''
            
            if img_path:
                existing = target.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                target['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {img_path}".strip(', ')
                
            # Bổ sung thông tin còn thiếu từ OCR
            for k in ['Họ tên', 'Ngày sinh', 'CMND', 'Giới tính']:
                if orphan.get(k) and not target.get(k):
                    target[k] = orphan[k]
                    
            console.print(
                f"   [bold green]→ [FUZZY MATCH][/bold green] Ghép ảnh Khác (CCCD OCR: {orphan_cccd}) "
                f"vào bản ghi {best_cccd} (khớp {score}/3 trường: tên/ngày sinh/cccd)."
            )
            del records[orphan_cccd]

    # ---------- FINAL CLEANUP MATCH: GỘP CÁC DÒNG MỒ CÔI (CHỈ CÓ 1 MẶT ẢNH) ----------
    def _is_incomplete_orphan(rec):
        has_front = bool(rec.get('Ảnh mặt trước CCCD/CC') or rec.get('OCR Image Path Front') or rec.get('Full Image Path Front') or rec.get('Full OCR Image Path Front'))
        has_back = bool(rec.get('Ảnh mặt sau CCCD/CC') or rec.get('OCR Image Path Back') or rec.get('Full Image Path Back') or rec.get('Full OCR Image Path Back'))
        has_other = bool(rec.get('Ảnh khác (SMS/Chụp màn hình/...)') or rec.get('OCR Image Path Unknown') or rec.get('Full OCR Image Path Unknown'))
        return sum([has_front, has_back, has_other]) <= 1

    incomplete_cccds = [cccd for cccd, rec in list(records.items()) if _is_incomplete_orphan(rec)]
    for sec_cccd in incomplete_cccds:
        if sec_cccd not in records:
            continue
        sec_rec = records[sec_cccd]
        sec_name = _norm_for_match(sec_rec.get('Họ tên', ''))
        sec_dob = sec_rec.get('Ngày sinh', '')
        sec_issue = sec_rec.get('Ngày cấp CCCD', '')
        
        if not sec_name:
            continue
            
        candidates = []
        for target_cccd, target_rec in records.items():
            if target_cccd == sec_cccd:
                continue
            # Không gộp số CCCD hợp lệ vào số CCCD rác/placeholder
            if _is_invalid_cccd_placeholder(target_cccd) and not _is_invalid_cccd_placeholder(sec_cccd):
                continue
            # Không gộp bản ghi có QR vào bản ghi không có QR
            if not target_rec['has_qr_data'] and sec_rec['has_qr_data']:
                continue
                
            # Ngăn gộp nếu giới tính khai báo khác nhau
            s_gen = sec_rec.get('Giới tính', '').strip().lower()
            t_gen = target_rec.get('Giới tính', '').strip().lower()
            if s_gen and t_gen and s_gen != t_gen:
                continue
                
            target_name = _norm_for_match(target_rec.get('Họ tên', ''))
            if _is_similar_name(sec_name, target_name):
                candidates.append(target_cccd)
                
        best_target_cccd = None
        # Trường hợp 1: Có duy nhất 1 ứng cử viên trùng/cận tên trong database
        if len(candidates) == 1:
            best_target_cccd = candidates[0]
        # Trường hợp 2: Có nhiều ứng cử viên trùng tên, lọc bằng DOB, Ngày cấp, hoặc CCCD
        elif len(candidates) > 1:
            for target_cccd in candidates:
                target_rec = records[target_cccd]
                target_dob = target_rec.get('Ngày sinh', '')
                target_issue = target_rec.get('Ngày cấp CCCD', '')
                
                match_dob = (sec_dob and target_dob and _is_similar_dob(sec_dob, target_dob))
                match_issue = (sec_issue and target_issue and _is_similar_dob(sec_issue, target_issue))
                match_cccd = _is_similar_cccd(sec_cccd, target_cccd)
                
                if match_dob or match_issue or match_cccd:
                    if not best_target_cccd:
                        best_target_cccd = target_cccd
                    else:
                        curr_best = records[best_target_cccd]
                        if target_rec['has_qr_data'] and not curr_best['has_qr_data']:
                            best_target_cccd = target_cccd
                        elif _is_invalid_cccd_placeholder(best_target_cccd) and not _is_invalid_cccd_placeholder(target_cccd):
                            best_target_cccd = target_cccd
                            
        if best_target_cccd:
            target_rec = records[best_target_cccd]
            console.print(f"   [bold green]→ [FINAL CLEANUP MATCH][/bold green] Ghép bản ghi mồ côi {sec_cccd} ({sec_rec.get('Họ tên')}) vào bản ghi {best_target_cccd} (Trùng tên & trùng DOB/Ngày cấp/Cận CCCD/Duy nhất).")
            
            # Gộp mặt trước
            if not target_rec.get('Ảnh mặt trước CCCD/CC'):
                if sec_rec.get('Ảnh mặt trước CCCD/CC'):
                    target_rec['Ảnh mặt trước CCCD/CC'] = sec_rec['Ảnh mặt trước CCCD/CC']
                    target_rec['Full Image Path Front'] = sec_rec.get('Full Image Path Front', '')
                elif sec_rec.get('OCR Image Path Front'):
                    target_rec['Ảnh mặt trước CCCD/CC'] = sec_rec['OCR Image Path Front']
                    target_rec['Full Image Path Front'] = sec_rec.get('Full OCR Image Path Front', '')
            else:
                sec_front = sec_rec.get('Ảnh mặt trước CCCD/CC') or sec_rec.get('OCR Image Path Front')
                if sec_front:
                    existing = target_rec.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    target_rec['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {sec_front}".strip(', ')
                    
            # Gộp mặt sau
            if not target_rec.get('Ảnh mặt sau CCCD/CC'):
                if sec_rec.get('Ảnh mặt sau CCCD/CC'):
                    target_rec['Ảnh mặt sau CCCD/CC'] = sec_rec['Ảnh mặt sau CCCD/CC']
                    target_rec['Full Image Path Back'] = sec_rec.get('Full Image Path Back', '')
                elif sec_rec.get('OCR Image Path Back'):
                    target_rec['Ảnh mặt sau CCCD/CC'] = sec_rec['OCR Image Path Back']
                    target_rec['Full Image Path Back'] = sec_rec.get('Full OCR Image Path Back', '')
            else:
                sec_back = sec_rec.get('Ảnh mặt sau CCCD/CC') or sec_rec.get('OCR Image Path Back')
                if sec_back:
                    existing = target_rec.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                    target_rec['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {sec_back}".strip(', ')
            
            # Gộp Ảnh khác
            sec_other = sec_rec.get('Ảnh khác (SMS/Chụp màn hình/...)') or sec_rec.get('OCR Image Path Unknown')
            if sec_other:
                existing = target_rec.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                target_rec['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {sec_other}".strip(', ')
                
            # Bổ sung thông tin văn bản còn thiếu
            for k in ['CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD', 'Nơi cấp', 'Ngày hết hạn', 'Phân loại', 'QR Raw']:
                if sec_rec.get(k) and not target_rec.get(k):
                    target_rec[k] = sec_rec[k]
                    
            for flag in ['has_qr_data', 'has_ocr_data', 'has_cong_dan_front', 'has_address_front', 'has_address_back', 'has_cuc_truong_back', 'has_bo_cong_an_back']:
                if sec_rec.get(flag):
                    target_rec[flag] = True
                    
            if sec_rec.get('Ghi chú'):
                if isinstance(sec_rec['Ghi chú'], list):
                    target_rec['Ghi chú'].extend(sec_rec['Ghi chú'])
                else:
                    target_rec['Ghi chú'].append(str(sec_rec['Ghi chú']))
                    
            del records[sec_cccd]

    # Pass 3: Assign OCR images to empty slots and Rename logic
    for cccd, record in records.items():
        # Assign Front
        if not record['Ảnh mặt trước CCCD/CC']:
            if record.get('OCR Image Path Front'):
                record['Ảnh mặt trước CCCD/CC'] = record.pop('OCR Image Path Front')
                record['Full Image Path Front'] = record.pop('Full OCR Image Path Front')
            elif record.get('OCR Image Path Unknown') and record['Ảnh mặt sau CCCD/CC']:
                record['Ảnh mặt trước CCCD/CC'] = record['OCR Image Path Unknown']
                record['Full Image Path Front'] = record['Full OCR Image Path Unknown']
                record['OCR Image Path Unknown'] = ''
                
        # Assign Back
        if not record['Ảnh mặt sau CCCD/CC']:
            if record.get('OCR Image Path Back'):
                record['Ảnh mặt sau CCCD/CC'] = record.pop('OCR Image Path Back')
                record['Full Image Path Back'] = record.pop('Full OCR Image Path Back')
            elif record.get('OCR Image Path Unknown') and record['Ảnh mặt trước CCCD/CC']:
                record['Ảnh mặt sau CCCD/CC'] = record['OCR Image Path Unknown']
                record['Full Image Path Back'] = record['Full OCR Image Path Unknown']
                record['OCR Image Path Unknown'] = ''

        if not record['Ảnh mặt trước CCCD/CC'] and not record['Ảnh mặt sau CCCD/CC']:
            if record.get('OCR Image Path Unknown'):
                record['Ghi chú'].append('Không thể phân biệt được ảnh này là mặt trước hay mặt sau do mờ và không chứa mã QR')

        hoten = record['Họ tên'] or 'KhongTen'
        cmnd = record['CMND']
        hoten_clean = hoten # keep space as user requested: Nguyễn Văn A_012_Mặt trước.jpg
        # Only remove illegal characters for filename
        hoten_clean = re.sub(r'[\\/*?:"<>|]', '', hoten_clean)
        cmnd_str = f"_{cmnd}" if cmnd else ""
        
        if record['Ảnh mặt trước CCCD/CC']:
            ext = os.path.splitext(record['Ảnh mặt trước CCCD/CC'])[1]
            record['Đổi tên Ảnh mặt trước CCCD/CC'] = f"{hoten_clean}_{cccd}{cmnd_str}_Mặt trước{ext}"
            
        if record['Ảnh mặt sau CCCD/CC']:
            ext = os.path.splitext(record['Ảnh mặt sau CCCD/CC'])[1]
            record['Đổi tên Ảnh mặt sau CCCD/CC'] = f"{hoten_clean}_{cccd}{cmnd_str}_Mặt sau{ext}"

        anh_khac = record.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
        if anh_khac:
            anh_khac_list = [p.strip() for p in str(anh_khac).split(', ') if p.strip()]
            renamed_list = []
            for i, p in enumerate(anh_khac_list):
                ext = os.path.splitext(p)[1]
                suffix = f"_Khác_{i+1}" if len(anh_khac_list) > 1 else "_Khác"
                renamed_list.append(f"{hoten_clean}_{cccd}{cmnd_str}{suffix}{ext}")
            record['Đổi tên Ảnh khác'] = ", ".join(renamed_list)

    processed_data = list(records.values())
    # Lấy danh sách địa chỉ duy nhất
    unique_addresses = list(set([item['Nơi thường trú gốc'] for item in processed_data if item.get('Nơi thường trú gốc')]))
    address_map = {}

    # Gọi API chuẩn hóa địa chỉ theo batch, advance progress bar theo từng kết quả
    if normalize_address and unique_addresses:
        console.print(Panel(f"[bold cyan]🌐 ĐANG CHUẨN BỊ GỌI API CHUẨN HÓA CHO {len(unique_addresses)} ĐỊA CHỈ DUY NHẤT VỚI {api_threads} LUỒNG...[/bold cyan]", border_style="green"))
        batch_size = 100
        total_addrs = len(unique_addresses)
        processed_count = 0
        with Progress(
            SpinnerColumn(),
            TextColumn("[progress.description]{task.description}"),
            BarColumn(),
            TaskProgressColumn(),
            CountColumn(),
            TextColumn("[bold]{task.fields[status]}"),
            TimeElapsedColumn(),
            TextColumn("⏳ ETA:"),
            TimeRemainingColumn(),
            SpeedColumn(),
            console=console,
            refresh_per_second=REFRESH_RATE,
        ) as api_progress:
            api_task = api_progress.add_task(
                "[cyan]Đang chuẩn hóa địa chỉ...",
                total=total_addrs,
                status=""
            )
            
            for i in range(0, total_addrs, batch_size):
                batch = unique_addresses[i:i+batch_size]
                
                with concurrent.futures.ThreadPoolExecutor(max_workers=api_threads) as executor:
                    future_to_addr = {executor.submit(fetch_single_address, addr): addr for addr in batch}
                    for future in concurrent.futures.as_completed(future_to_addr):
                        result = future.result()
                        processed_count += 1
                        if result and 'original' in result:
                            orig_addr = result['original']
                            address_map[orig_addr] = result
                            short_addr = orig_addr[:45]
                            
                            if result.get('success'):
                                new_addr = result.get('converted', '')
                                status_text = f"[green]✓[/green] {short_addr}"
                                file_logs.append(f"[{processed_count}/{total_addrs}] CHUẨN HÓA OK: {orig_addr} -> {new_addr}")
                                if DEBUG_MODE and not IN_COLAB:
                                    api_progress.console.print(f"[bold cyan][{processed_count}/{total_addrs}][/bold cyan] [dim]Từ:[/dim] [yellow]{orig_addr}[/yellow]\n{' '*(len(str(total_addrs))*2 + 5)}[dim]→  [/dim] [bold green]{new_addr}[/bold green]")
                            else:
                                err_msg = result.get('error', 'Lỗi không xác định')
                                status_text = f"[red]✗[/red] {short_addr}"
                                file_logs.append(f"[{processed_count}/{total_addrs}] CHUẨN HÓA LỖI: {orig_addr} -> {err_msg}")
                                if DEBUG_MODE and not IN_COLAB:
                                    api_progress.console.print(f"[bold cyan][{processed_count}/{total_addrs}][/bold cyan] [dim]Từ:[/dim] [yellow]{orig_addr}[/yellow]\n{' '*(len(str(total_addrs))*2 + 5)}[dim]→  [/dim] [bold red]{err_msg}[/bold red]")
                                
                            api_progress.update(api_task, advance=1, status=status_text)

    # Cập nhật kết quả API vào dữ liệu
    for row in processed_data:
        addr = row.get('Nơi thường trú gốc')
        if addr and addr in address_map:
            result = address_map[addr]
            new_notes = []
            if row['Ghi chú']:
                new_notes.extend(row['Ghi chú'] if isinstance(row['Ghi chú'], list) else [row['Ghi chú']])
            
            if result.get('success'):
                converted_addr = result.get('converted', '')
                row['Địa chỉ chuẩn hóa mới'] = converted_addr
                if converted_addr.lower().strip() == addr.lower().strip():
                    new_notes.append("Địa chỉ không đổi")
                if result.get('notSure'):
                    new_notes.append("Địa chỉ chuyển đổi chưa chắc chắn")
            else:
                err_msg = result.get('error', 'Lỗi không xác định khi chuẩn hóa')
                new_notes.append(err_msg)
            
            row['Ghi chú'] = '; '.join(new_notes)
            
    # Pass 4: Final formatting and cleanup
    for cccd, record in records.items():
        # Đảm bảo Ghi chú là list
        raw_notes = record['Ghi chú']
        if isinstance(raw_notes, str):
            raw_notes = raw_notes.split('; ')
        elif not isinstance(raw_notes, list):
            raw_notes = []
            
        # Clean up "QR không đọc được" và "Lấy bằng OCR" note if we actually have a QR code
        final_notes = [n for n in raw_notes if n]
        if record.get('QR Raw'):
            final_notes = [n for n in final_notes if 'QR không đọc được' not in n and 'Lấy bằng OCR' not in n]
            if 'Đọc mã QR' not in final_notes:
                final_notes.insert(0, 'Đọc mã QR')
        
        # Xử lý logic CMND (Yêu cầu mới)
        if not record['CMND']:
            if record.get('QR Raw'):
                record['CMND'] = 'Không có'
            else:
                record['CMND'] = 'Chưa xác định'

        # Tính toán ngày hết hạn dựa trên ngày sinh nếu bị khuyết (rất hay gặp ở luồng OCR)
        if not record['Ngày hết hạn'] and record.get('Ngày sinh'):
            record['Ngày hết hạn'] = calculate_expiry_date(record['Ngày sinh'])
            
        if not record.get('QR Raw'):
            if record['has_cong_dan_front'] or record['has_address_front'] or record['has_cuc_truong_back']:
                record['Phân loại'] = 'Căn cước công dân'
            elif record['has_address_back'] and not record['has_cong_dan_front'] and record['has_bo_cong_an_back']:
                record['Phân loại'] = 'Căn cước'
            else:
                record['Phân loại'] = 'Khác'
                
        # Tự động suy luận Nơi cấp dựa vào Phân loại (dành cho các thẻ hỏng QR)
        if not record.get('Nơi cấp'):
            if record['Phân loại'] == 'Căn cước công dân':
                record['Nơi cấp'] = 'CỤC TRƯỞNG CỤC CẢNH SÁT QUẢN LÝ HÀNH CHÍNH VỀ TRẬT TỰ XÃ HỘI'
            elif record['Phân loại'] == 'Căn cước':
                record['Nơi cấp'] = 'BỘ CÔNG AN'
                
        # Deduplicate notes and convert to string
        unique_notes = []
        for note in final_notes:
            # handle cases where notes were joined by '; '
            for subnote in note.split('; '):
                if subnote and subnote not in unique_notes:
                    unique_notes.append(subnote)
        record['Ghi chú'] = '; '.join(unique_notes)
            
    # Fix note formatting for those not processed by API
    for row in processed_data:
        if isinstance(row['Ghi chú'], list):
            row['Ghi chú'] = '; '.join(row['Ghi chú'])

    # Export to Excel
    print("Đang xuất file Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    headers = [
        "STT", "Họ tên", "CCCD", "CMND", "Giới tính", "Ngày sinh", 
        "Nơi thường trú gốc", "Địa chỉ chuẩn hóa mới", "Ngày cấp CCCD", "Nơi cấp", "Ngày hết hạn", "Phân loại", "Ghi chú", "QR Raw", 
        "Ảnh mặt trước CCCD/CC", "Ảnh mặt sau CCCD/CC", "Ảnh khác (SMS/Chụp màn hình/...)", "Đổi tên Ảnh mặt trước CCCD/CC", "Đổi tên Ảnh mặt sau CCCD/CC", "Đổi tên Ảnh khác"
    ]
    
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    if 'incremental_scan' in locals() and incremental_scan and old_records:
        processed_data = old_records + processed_data

    # Chuẩn hoá "Họ tên" thành chữ in hoa chữ cái đầu (VD: NGUYỄN TRỌNG HỮU -> Nguyễn Trọng Hữu)
    for row_data in processed_data:
        if row_data.get('Họ tên'):
            row_data['Họ tên'] = row_data['Họ tên'].title()

    # Xử lý logic gộp ảnh "Khác"
    for row_data in processed_data:
        if row_data.get('Phân loại') == 'Khác':
            hoten = row_data['Họ tên'] or 'KhongTen'
            cccd = row_data['CCCD']
            cmnd = row_data['CMND']
            hoten_clean = re.sub(r'[\\/*?:"<>|]', '', hoten)
            cmnd_str = f"_{cmnd}" if cmnd and cmnd not in ['Không có', 'Chưa xác định'] else ""
            base_name = f"{hoten_clean}_{cccd}{cmnd_str}"
            
            imgs = []
            if row_data.get('Ảnh mặt trước CCCD/CC'): imgs.append(row_data['Ảnh mặt trước CCCD/CC'])
            if row_data.get('Ảnh mặt sau CCCD/CC'): imgs.append(row_data['Ảnh mặt sau CCCD/CC'])
            if row_data.get('OCR Image Path Unknown'): imgs.append(row_data['OCR Image Path Unknown'])
            
            anh_khac = row_data.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
            anh_khac_list = []
            if anh_khac: anh_khac_list.extend(str(anh_khac).split(', '))
            anh_khac_list.extend(imgs)
            
            renamed_list = []
            for i, p in enumerate(anh_khac_list):
                ext = os.path.splitext(p)[1]
                suffix = f"_Khác_{i+1}" if len(anh_khac_list) > 1 else "_Khác"
                renamed_list.append(f"{base_name}{suffix}{ext}")
            
            row_data['Ảnh khác (SMS/Chụp màn hình/...)'] = ", ".join(filter(None, anh_khac_list))
            row_data['Đổi tên Ảnh khác'] = ", ".join(renamed_list)
            
            row_data['Ảnh mặt trước CCCD/CC'] = ''
            row_data['Ảnh mặt sau CCCD/CC'] = ''
            row_data['Đổi tên Ảnh mặt trước CCCD/CC'] = ''
            row_data['Đổi tên Ảnh mặt sau CCCD/CC'] = ''

    for idx, row_data in enumerate(processed_data):
        row = [
            idx + 1,
            row_data['Họ tên'],
            row_data['CCCD'],
            row_data['CMND'],
            row_data['Giới tính'],
            row_data['Ngày sinh'],
            row_data['Nơi thường trú gốc'],
            row_data['Địa chỉ chuẩn hóa mới'],
            row_data['Ngày cấp CCCD'],
            row_data['Nơi cấp'],
            row_data['Ngày hết hạn'],
            row_data['Phân loại'],
            row_data['Ghi chú'],
            row_data.get('QR Raw', ''),
            row_data.get('Ảnh mặt trước CCCD/CC', ''),
            row_data.get('Ảnh mặt sau CCCD/CC', ''),
            row_data.get('Ảnh khác (SMS/Chụp màn hình/...)', ''),
            row_data.get('Đổi tên Ảnh mặt trước CCCD/CC', ''),
            row_data.get('Đổi tên Ảnh mặt sau CCCD/CC', ''),
            row_data.get('Đổi tên Ảnh khác', '')
        ]
        ws.append(row)
        
        # Tô màu nền vàng cho các dòng lấy bằng OCR hoặc không đọc được
        note_str = row_data.get('Ghi chú', '')
        if "Lấy bằng OCR" in note_str or "không đọc được" in note_str:
            yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = yellow_fill

    # Adjust column widths slightly
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 40) # Max width 40

    console.print("\n")
    console.print(Panel("[bold cyan]✨ CHUẨN BỊ XUẤT FILE EXCEL ✨[/bold cyan]", border_style="green"))
    
    console.print("\n[bold yellow]Chọn vị trí lưu thư mục export (chứa các file kết quả):[/bold yellow]")
    console.print("  [cyan]1[/cyan]. Lưu ở cùng cấp thư mục với thư mục đang xử lý (Ví dụ: Thu_muc_anh_exports)")
    console.print("  [cyan]2[/cyan]. Lưu mặc định ở thư mục export của dự án (mặc định)")
    console.print("  [cyan]3[/cyan]. Nhập địa chỉ thư mục mong muốn")
    
    if IN_COLAB:
        export_option = "1"
    else:
        export_option = Prompt.ask("[bold cyan]Nhập lựa chọn của bạn[/bold cyan]", choices=["1", "2", "3"], default="2")
    
    if export_option == "1":
        clean_input_dir = os.path.normpath(input_dir)
        exports_dir = clean_input_dir + "_exports"
    elif export_option == "3":
        custom_dir = clean_path(Prompt.ask("[bold cyan]Nhập đường dẫn thư mục mong muốn lưu kết quả[/bold cyan]"))
        if not custom_dir:
            console.print("[yellow]Đường dẫn rỗng, quay về mặc định.[/yellow]")
            exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        else:
            exports_dir = custom_dir
    else:
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')

    try:
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
    except Exception as e:
        console.print(f"[bold red]❌ Lỗi tạo thư mục {exports_dir}: {e}. Quay về mặc định.[/bold red]")
        exports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        if not os.path.exists(exports_dir):
            os.makedirs(exports_dir)
            
    timestamp = datetime.datetime.now().strftime("%d%m%Y_%H%M%S")
    default_filename = f"ket_qua_{timestamp}.xlsx"
    
    if IN_COLAB:
        custom_name = ""
    else:
        custom_name = Prompt.ask(f"[bold cyan]Nhập tên file Excel muốn lưu[/bold cyan] (nhấn Enter để dùng tên mặc định [yellow]'{default_filename}'[/yellow])").strip()
    if not custom_name:
        output_filename = os.path.join(exports_dir, default_filename)
    else:
        if not custom_name.endswith('.xlsx'):
            custom_name += '.xlsx'
        output_filename = os.path.join(exports_dir, custom_name)

    # Khôi phục các sheet phụ theo yêu cầu
    ws_qr = wb.create_sheet(title="QR_scanned")
    ws_qr.append(["STT", "Tên file"])
    qr_idx = 1
    for item in extracted_items:
        if item.get('Scan Type') == 'QR_scanned':
            ws_qr.append([qr_idx, item['Image Path']])
            qr_idx += 1
            
    ws_ocr = wb.create_sheet(title="OCR_scanned")
    ws_ocr.append(["STT", "Tên file"])
    ocr_idx = 1
    for item in extracted_items:
        if item.get('Scan Type') == 'OCR_scanned':
            ws_ocr.append([ocr_idx, item['Image Path']])
            ocr_idx += 1
            
    # Xác định các file bị trùng lặp hoặc không được sử dụng
    used_images = set()
    used_images = set()
    cccd_to_front = {}
    cccd_to_back = {}
    for row in processed_data:
        cccd = row.get('CCCD') or row.get('CMND')
        if row.get('Full Image Path Front'): 
            used_images.add(row['Full Image Path Front'])
            if cccd: cccd_to_front[cccd] = row.get('Ảnh mặt trước CCCD/CC', '')
        if row.get('Full Image Path Back'): 
            used_images.add(row['Full Image Path Back'])
            if cccd: cccd_to_back[cccd] = row.get('Ảnh mặt sau CCCD/CC', '')
        
    duplicate_files = []
    for item in extracted_items:
        if item['Full Image Path'] not in used_images:
            cccd = item.get('CCCD') or item.get('CMND')
            dup_with = ""
            if cccd:
                is_front = False
                is_back = False
                if item.get('QR Raw'):
                    fields = item['QR Raw'].split('|')
                    if len(fields) == 7: is_front = True
                    elif len(fields) >= 10: is_back = True
                elif item.get('OCR Side') == 'Front': is_front = True
                elif item.get('OCR Side') == 'Back': is_back = True
                
                if is_front and cccd in cccd_to_front:
                    dup_with = cccd_to_front[cccd]
                elif is_back and cccd in cccd_to_back:
                    dup_with = cccd_to_back[cccd]
                else:
                    dup_with = cccd_to_front.get(cccd) or cccd_to_back.get(cccd) or ""
            item['Duplicate With'] = dup_with
            duplicate_files.append(item)

    ws_dup = wb.create_sheet(title="duplicate")
    ws_dup.append(["STT", "Tên file", "Trùng lặp với"])
    for i, item in enumerate(duplicate_files, 1):
        ws_dup.append([i, item['Image Path'], item.get('Duplicate With', '')])

    # --- Sheet "Review": dòng dữ liệu chưa đầy đủ thông tin ---
    REQUIRED_FIELDS = ['Họ tên', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD',
                       'Ảnh mặt trước CCCD/CC', 'Ảnh mặt sau CCCD/CC']
    review_rows = []
    for row in processed_data:
        missing = [f for f in REQUIRED_FIELDS if not row.get(f)]
        if 'UNKNOWN' in str(row.get('CCCD', '')):
            missing.append('CCCD')
        if missing:
            review_rows.append((row, missing))

    ws_review = wb.create_sheet(title="Review")
    ws_review.append(["STT", "CCCD", "Họ tên", "Ảnh mặt trước", "Ảnh mặt sau", "Trường còn thiếu"])
    for cell in ws_review[1]:
        cell.font = Font(bold=True)
    for stt, (row, missing) in enumerate(review_rows, 1):
        ws_review.append([
            stt,
            row.get('CCCD', ''),
            row.get('Họ tên', ''),
            row.get('Ảnh mặt trước CCCD/CC', ''),
            row.get('Ảnh mặt sau CCCD/CC', ''),
            ', '.join(missing)
        ])
    for col in ws_review.columns:
        ws_review.column_dimensions[col[0].column_letter].width = 30

    # --- Sheet "Unknown": ảnh không thuộc dòng nào (không đọc được CCCD) ---
    all_matched_basenames = set()
    for row in processed_data:
        if row.get('Full Image Path Front'): all_matched_basenames.add(os.path.basename(row['Full Image Path Front']))
        if row.get('Full Image Path Back'): all_matched_basenames.add(os.path.basename(row['Full Image Path Back']))
    for row, _ in review_rows:
        if row.get('Full Image Path Front'): all_matched_basenames.add(os.path.basename(row['Full Image Path Front']))
        if row.get('Full Image Path Back'): all_matched_basenames.add(os.path.basename(row['Full Image Path Back']))

    unknown_image_paths = [p for p in all_original_image_paths if os.path.basename(p) not in all_matched_basenames]

    ws_unknown = wb.create_sheet(title="Unknown")
    ws_unknown.append(["STT", "Tên file gốc"])
    for cell in ws_unknown[1]:
        cell.font = Font(bold=True)
    for stt, fpath in enumerate(unknown_image_paths, 1):
        ws_unknown.append([stt, os.path.basename(fpath)])
    for col in ws_unknown.columns:
        ws_unknown.column_dimensions[col[0].column_letter].width = 40

    wb.save(output_filename)
    
    with console.status("[bold green]Đang tạo các file nén zip phân loại ảnh...", spinner="dots"):
        import shutil
        
        # 1. original.zip (Chứa các file sau khi đã đánh số thứ tự)
        original_zip_path = os.path.join(exports_dir, 'original.zip')
        count_original = 0
        
        # Thu thập các file ảnh đã được đánh số thứ tự trong input_dir
        renamed_files = []
        if os.path.isdir(input_dir):
            for f in os.listdir(input_dir):
                fpath = os.path.join(input_dir, f)
                if os.path.isfile(fpath):
                    base, ext = os.path.splitext(f)
                    if base.isdigit() and ext.lower() in ('.jpg', '.jpeg', '.png', '.heic', '.webp'):
                        renamed_files.append((fpath, int(base)))
            # Sắp xếp theo thứ tự số tăng dần
            renamed_files.sort(key=lambda x: x[1])
            
        with zipfile.ZipFile(original_zip_path, 'w') as zf:
            for path, _ in renamed_files:
                zf.write(path, os.path.basename(path))
                count_original += 1
        console.print(f" [green]✓[/green] Đã tạo [bold]original.zip[/bold] với {count_original} file.")
        
        # 2. rename.zip
        rename_zip_path = os.path.join(exports_dir, 'rename.zip')
        with zipfile.ZipFile(rename_zip_path, 'w') as zf:
            count_rename = 0
            for row in processed_data + [r[0] for r in review_rows]:
                if row.get("Phân loại") == "Khác":
                    folder = "Khác"
                    if row.get('Ảnh khác (SMS/Chụp màn hình/...)') and row.get('Đổi tên Ảnh khác'):
                        orig_names = str(row['Ảnh khác (SMS/Chụp màn hình/...)']).split(', ')
                        new_names = str(row['Đổi tên Ảnh khác']).split(', ')
                        
                        full_paths = []
                        if row.get('Full Image Path Front'): full_paths.append(row['Full Image Path Front'])
                        if row.get('Full Image Path Back'): full_paths.append(row['Full Image Path Back'])
                        if row.get('Full OCR Image Path Unknown'): full_paths.append(row['Full OCR Image Path Unknown'])
                        
                        for i, orig in enumerate(orig_names):
                            if i < len(new_names):
                                new_name = new_names[i]
                                match_fp = None
                                for fp in full_paths:
                                    if os.path.basename(fp) == orig:
                                        match_fp = fp
                                        break
                                
                                if match_fp:
                                    if not os.path.exists(match_fp):
                                        fallback = os.path.join(input_dir, os.path.basename(match_fp))
                                        if os.path.exists(fallback): match_fp = fallback
                                    if os.path.exists(match_fp):
                                        zf.write(match_fp, f"{folder}/{new_name}")
                                        count_rename += 1
                else:
                    folder = "CCCD" if row.get("Phân loại") == "Căn cước công dân" else "CC"
                    
                    front_path = row.get('Full Image Path Front')
                    if front_path and not os.path.exists(front_path):
                        fallback = os.path.join(input_dir, os.path.basename(front_path))
                        if os.path.exists(fallback): front_path = fallback
                    if front_path and os.path.exists(front_path) and row.get('Đổi tên Ảnh mặt trước CCCD/CC'):
                        zf.write(front_path, f"{folder}/{row['Đổi tên Ảnh mặt trước CCCD/CC']}")
                        count_rename += 1
                        
                    back_path = row.get('Full Image Path Back')
                    if back_path and not os.path.exists(back_path):
                        fallback = os.path.join(input_dir, os.path.basename(back_path))
                        if os.path.exists(fallback): back_path = fallback
                    if back_path and os.path.exists(back_path) and row.get('Đổi tên Ảnh mặt sau CCCD/CC'):
                        zf.write(back_path, f"{folder}/{row['Đổi tên Ảnh mặt sau CCCD/CC']}")
                        count_rename += 1
                    
        console.print(f" [green]✓[/green] Đã tạo [bold]rename.zip[/bold] với {count_rename} file đã được đổi tên (trong CC và CCCD).")
        
        # 3. Khôi phục lại các file nén phân loại cũ
        def create_zip_helper(zip_name, file_paths):
            if not file_paths:
                return
            zip_path = os.path.join(exports_dir, zip_name)
            with zipfile.ZipFile(zip_path, 'w') as zf:
                for fpath in file_paths:
                    if not os.path.exists(fpath):
                        fallback = os.path.join(input_dir, os.path.basename(fpath))
                        if os.path.exists(fallback): fpath = fallback
                    if os.path.exists(fpath):
                        zf.write(fpath, os.path.basename(fpath))
            console.print(f" [green]✓[/green] Đã tạo [bold]{zip_name}[/bold] với {len(file_paths)} file.")

    qr_files = [item['Full Image Path'] for item in extracted_items if item.get('Scan Type') == 'QR_scanned']
    ocr_files = [item['Full Image Path'] for item in extracted_items if item.get('Scan Type') == 'OCR_scanned']
    dup_files = [item['Full Image Path'] for item in duplicate_files]
    
    create_zip_helper('QR_scanned.zip', qr_files)
    create_zip_helper('OCR_scanned.zip', ocr_files)
    create_zip_helper('duplicate.zip', dup_files)

    # 5. ReadQR.zip: ảnh có dòng Ghi chú chứa "Đọc mã QR"
    readqr_image_paths = []
    added_readqr = set()
    for row in processed_data + [r[0] for r in review_rows]:
        if 'Đọc mã QR' in row.get('Ghi chú', ''):
            for field in ['Full Image Path Front', 'Full Image Path Back']:
                p = row.get(field)
                if p and p not in added_readqr:
                    readqr_image_paths.append(p)
                    added_readqr.add(p)
    create_zip_helper('ReadQR.zip', readqr_image_paths)

    # 6. ReadOCR.zip: ảnh có dòng Ghi chú chứa "Lấy bằng OCR"
    readocr_image_paths = []
    added_readocr = set()
    for row in processed_data + [r[0] for r in review_rows]:
        if 'Lấy bằng OCR' in row.get('Ghi chú', ''):
            for field in ['Full Image Path Front', 'Full Image Path Back']:
                p = row.get(field)
                if p and p not in added_readocr:
                    readocr_image_paths.append(p)
                    added_readocr.add(p)
    create_zip_helper('ReadOCR.zip', readocr_image_paths)

    # 3. review.zip: ảnh của các dòng chưa đầy đủ thông tin
    review_image_paths = []
    added_review = set()
    for row, _ in review_rows:
        for field in ['Full Image Path Front', 'Full Image Path Back']:
            p = row.get(field)
            if p and p not in added_review:
                review_image_paths.append(p)
                added_review.add(p)
    create_zip_helper('review.zip', review_image_paths)

    # 4. unknown.zip: ảnh không thuộc dòng nào
    create_zip_helper('unknown.zip', unknown_image_paths)
    
    console.print("\n" + "🎉"*15)
    console.print(f"[bold green]ĐÃ HOÀN TẤT THÀNH CÔNG![/bold green]")
    console.print(f"File kết quả được lưu tại: [yellow]{os.path.abspath(output_filename)}[/yellow]")
    
    # Xuất file log
    log_filename = os.path.join(exports_dir, f"log_{timestamp}.txt")
    file_logs.append("\nĐÃ HOÀN TẤT THÀNH CÔNG!")
    file_logs.append(f"File kết quả: {os.path.abspath(output_filename)}")
    with open(log_filename, "w", encoding="utf-8") as f:
        f.write("\n".join(file_logs))
    console.print(f"File log chi tiết được lưu tại: [yellow]{os.path.abspath(log_filename)}[/yellow]")
    console.print("🎉"*15 + "\n")
    
    from rich.table import Table
    table = Table(title="📊 BÁO CÁO THỐNG KÊ KẾT QUẢ", show_header=True, header_style="bold magenta")
    table.add_column("Chỉ số đo lường", style="cyan", width=42)
    table.add_column("Số lượng", justify="right", style="green", width=12)
    table.add_column("Tỉ lệ", justify="right", style="yellow", width=10)
    
    table.add_row("[bold]1. THỐNG KÊ ẢNH ĐẦU VÀO[/bold]", "", "")
    total_imgs = total_raw_images
    table.add_row("Tổng số file ảnh đã đọc", str(total_imgs), "100.0%")
    table.add_row("Ảnh trùng lặp bị tự động loại bỏ", str(duplicates_count), f"{(duplicates_count/total_imgs*100):.1f}%" if total_imgs else "0.0%")
    table.add_row("Ảnh rác / Không thể phân loại (Unknown)", str(len(unknown_image_paths)), f"{(len(unknown_image_paths)/total_imgs*100):.1f}%" if total_imgs else "0.0%")
    valid_imgs = total_imgs - duplicates_count - len(unknown_image_paths)
    table.add_row("Ảnh đưa vào trích xuất thành công", str(valid_imgs), f"{(valid_imgs/total_imgs*100):.1f}%" if total_imgs else "0.0%")
    
    table.add_row("", "", "")
    table.add_row("[bold]2. THỐNG KÊ HỒ SƠ (CÔNG DÂN)[/bold]", "", "")
    total_profiles = len(processed_data) + len(review_rows)
    table.add_row("Tổng số hồ sơ (công dân) quét được", str(total_profiles), "100.0%" if total_profiles else "0.0%")
    table.add_row("Hồ sơ hoàn thiện đầy đủ", str(len(processed_data)), f"{(len(processed_data)/total_profiles*100):.1f}%" if total_profiles else "0.0%")
    table.add_row("Hồ sơ thiếu/lỗi cần kiểm tra (Review)", str(len(review_rows)), f"{(len(review_rows)/total_profiles*100):.1f}%" if total_profiles else "0.0%")
    
    table.add_row("", "", "")
    table.add_row("[bold]3. THỐNG KÊ THEO CÔNG NGHỆ[/bold]", "", "")
    qr_count = sum(1 for r in processed_data if 'Lấy bằng OCR' not in str(r.get('Ghi chú', ''))) + sum(1 for r, _ in review_rows if 'Lấy bằng OCR' not in str(r.get('Ghi chú', '')))
    ocr_count = sum(1 for r in processed_data if 'Lấy bằng OCR' in str(r.get('Ghi chú', ''))) + sum(1 for r, _ in review_rows if 'Lấy bằng OCR' in str(r.get('Ghi chú', '')))
    
    table.add_row("Dữ liệu trích xuất chuẩn xác bằng mã QR", str(qr_count), f"{(qr_count/total_profiles*100):.1f}%" if total_profiles else "0.0%")
    table.add_row("Dữ liệu cứu hộ thành công bằng AI OCR", str(ocr_count), f"{(ocr_count/total_profiles*100):.1f}%" if total_profiles else "0.0%")
    
    end_time = time.time()
    elapsed = end_time - start_time
    hours, rem = divmod(elapsed, 3600)
    minutes, seconds = divmod(rem, 60)
    time_str = f"{int(hours)}h {int(minutes)}m {int(seconds)}s" if hours > 0 else f"{int(minutes)}m {int(seconds)}s"
    
    table.add_row("", "", "")
    table.add_row("[bold]4. HIỆU SUẤT XỬ LÝ[/bold]", "", "")
    table.add_row("Tổng thời gian thực thi", time_str, "")
    
    processing_times = [r.get('_processing_time', 0) for r in extracted_items if r.get('_processing_time')]
    if processing_times:
        table.add_row("Tốc độ OCR chậm nhất / 1 ảnh", f"{max(processing_times):.1f}s", "")
        table.add_row("Tốc độ OCR nhanh nhất / 1 ảnh", f"{min(processing_times):.1f}s", "")
        table.add_row("Tốc độ OCR trung bình / 1 ảnh", f"{(sum(processing_times) / len(processing_times)):.1f}s", "")
        
    if normalize_address and 'address_map' in locals() and address_map:
        api_processing_times = [r.get('_processing_time', 0) for r in address_map.values() if r.get('_processing_time')]
        if api_processing_times:
            table.add_row("Tốc độ gọi API chậm nhất / 1 dòng", f"{max(api_processing_times):.2f}s", "")
            table.add_row("Tốc độ gọi API nhanh nhất / 1 dòng", f"{min(api_processing_times):.2f}s", "")
            table.add_row("Tốc độ gọi API trung bình / 1 dòng", f"{(sum(api_processing_times) / len(api_processing_times)):.2f}s", "")
    
    console.print(table)
    console.print()

    
    # Xoá file backup tạm sau khi hoàn tất thành công
    try:
        if 'realtime_csv' in locals() and os.path.exists(realtime_csv): os.remove(realtime_csv)
        if 'realtime_log' in locals() and os.path.exists(realtime_log): os.remove(realtime_log)
    except: pass



def clean_name_string(name):
    if not name: return ""
    import re
    rules = load_ocr_rules()
    clean_line = str(name)
    
    for pattern in rules.get("blacklist_patterns", []):
        clean_line = re.sub(pattern, '', clean_line)
        
    clean_line = re.sub(r'(?i)^(họ và tên|ho ten|kho và tên)[:\s]*', '', clean_line)
    
    # Lọc bỏ các từ chỉ có chữ thường (do nhiễu OCR) vì tên VN luôn IN HOA
    words = []
    for w in clean_line.split():
        # Xóa các ký tự không phải chữ
        cw = re.sub(r'[^a-zA-ZÀ-ɏḀ-ỿ]', '', w)
        if cw.isupper():
            words.append(cw)
            
    return " ".join(words).strip()

def clean_address_string(addr):
    if not addr: return ""
    import re
    import difflib
    
    rules = load_ocr_rules()
    valid_locs = load_valid_locations()
    clean_line = str(addr)
    
    # 1. Xóa rác bằng regex
    for pattern in rules.get("blacklist_patterns", []):
        clean_line = re.sub(pattern, '', clean_line)
        
    # 2. Thay thế typo mapping chuẩn
    for bad, good in rules.get("exact_typo_mapping", {}).items():
        clean_line = clean_line.replace(bad, good)
        
    # 3. Chuẩn hóa khoảng trắng, dấu phẩy
    clean_line = re.sub(r'\s+', ' ', clean_line).strip(', ')
    clean_line = re.sub(r',\s*,', ',', clean_line)
    clean_line = re.sub(r'^\s*,\s*', '', clean_line)
    clean_line = re.sub(r'\s*,\s*$', '', clean_line)
    
    # 4. Viết tắt (dùng regex tránh thay thế nhầm giữa từ)
    for pattern, replacement in rules.get("abbreviation_mapping", {}).items():
        clean_line = re.sub(pattern, replacement, clean_line)
        
    # 5. Xóa các cụm từ bị lặp lại liên tiếp (trường hợp OCR đọc nhiều lần)
    parts = [p.strip() for p in clean_line.split(',') if p.strip()]
    deduped = []
    for p in parts:
        if not deduped or p.lower() != deduped[-1].lower():
            deduped.append(p)
    clean_line = ', '.join(deduped)

    # 6. Fuzzy Matching từng đoạn qua dấu phẩy
    segments = [s.strip() for s in clean_line.split(',') if s.strip() and "pir" not in s.lower()]
    corrected_segments = []
    
    # Caching strategies for optimization
    global VALID_LOCS_LOWER_MAP, CLEAN_ADDRESS_SEGMENT_CACHE
    if 'VALID_LOCS_LOWER_MAP' not in globals() or VALID_LOCS_LOWER_MAP is None:
        VALID_LOCS_LOWER_MAP = {loc.lower(): loc for loc in (valid_locs or [])}
    if 'CLEAN_ADDRESS_SEGMENT_CACHE' not in globals() or CLEAN_ADDRESS_SEGMENT_CACHE is None:
        CLEAN_ADDRESS_SEGMENT_CACHE = {}
        
    for seg in segments:
        if valid_locs and 3 <= len(seg) <= 45:
            if seg in CLEAN_ADDRESS_SEGMENT_CACHE:
                corrected_segments.append(CLEAN_ADDRESS_SEGMENT_CACHE[seg])
                continue
                
            seg_lower = seg.lower()
            if VALID_LOCS_LOWER_MAP and seg_lower in VALID_LOCS_LOWER_MAP:
                matched = VALID_LOCS_LOWER_MAP[seg_lower]
                CLEAN_ADDRESS_SEGMENT_CACHE[seg] = matched
                corrected_segments.append(matched)
                continue
                
            matches = difflib.get_close_matches(seg, valid_locs, n=1, cutoff=0.85)
            if matches:
                CLEAN_ADDRESS_SEGMENT_CACHE[seg] = matches[0]
                corrected_segments.append(matches[0])
            else:
                CLEAN_ADDRESS_SEGMENT_CACHE[seg] = seg
                corrected_segments.append(seg)
        else:
            corrected_segments.append(seg)
            
    clean_line = ", ".join(corrected_segments)
    
    return clean_line

def run_reprocess(excel_path, mode="1", process_all_rows=False, normalize_address=True):
    import time
    start_time = time.time()
    from rich.text import Text
    import datetime
    file_logs = []
    
    img_map = {}
    if mode == "1":
        excel_dir = os.path.dirname(os.path.abspath(excel_path))
        parent_dir = os.path.dirname(excel_dir)
        image_dir = parent_dir
        
        user_img_dir = clean_path(Prompt.ask(f"[bold cyan]Nhập đường dẫn thư mục chứa ảnh gốc (Ấn Enter nếu là: {image_dir})[/bold cyan]"))
        if user_img_dir:
            image_dir = user_img_dir
            
        if not os.path.isdir(image_dir):
            console.print(f"[bold red]❌ Thư mục ảnh '{image_dir}' không tồn tại![/bold red]")
            return
            
        image_paths = []
        for ext in ('*.jpg', '*.jpeg', '*.png', '*.heic', '*.webp', '*.JPG', '*.JPEG', '*.PNG', '*.HEIC', '*.WEBP'):
            image_paths.extend(glob.glob(os.path.join(image_dir, ext)))
            
        image_paths = get_unique_images(image_paths)
            
        img_map = {os.path.basename(p): p for p in image_paths}
    
    console.print(f"[cyan]Đang đọc file Excel: {excel_path}[/cyan]")
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        console.print(f"[bold red]❌ Lỗi đọc file Excel: {e}[/bold red]")
        return
        
    headers = [cell.value for cell in ws[1]]
    if not headers or "Họ tên" not in headers:
        console.print("[bold red]❌ File Excel không đúng định dạng (không tìm thấy cột 'Họ tên').[/bold red]")
        return
        
    col_idx = {name: i for i, name in enumerate(headers)}
    
    required_cols = ['Họ tên', 'CCCD', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD', 'Nơi cấp', 'Ngày hết hạn']
    img_front_col = col_idx.get('Ảnh mặt trước CCCD/CC')
    img_back_col = col_idx.get('Ảnh mặt sau CCCD/CC')
    img_other_col = col_idx.get('Ảnh khác (SMS/Chụp màn hình/...)')
    
    if img_front_col is None or img_back_col is None:
        console.print("[bold red]❌ Không tìm thấy cột chứa tên file ảnh trong Excel.[/bold red]")
        return

    rows_to_process = []
    assigned_images = set()
    all_rows_info = []
    cccd_col = col_idx.get('CCCD')
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        front_name = row[img_front_col].value
        back_name = row[img_back_col].value
        other_name = row[img_other_col].value if img_other_col is not None else None
        
        if front_name: assigned_images.add(os.path.basename(str(front_name)))
        if back_name: assigned_images.add(os.path.basename(str(back_name)))
        if other_name:
            for part in str(other_name).split(','):
                part_clean = os.path.basename(part.strip())
                if part_clean:
                    assigned_images.add(part_clean)
        
        row_info = {
            'row_idx': row_idx,
            'front_name': front_name,
            'back_name': back_name,
            'other_name': other_name,
            'row_cells': row,
            'is_missing': False,
            'force_ocr_reprocess': False
        }
        all_rows_info.append(row_info)
        
        note_idx = col_idx.get('Ghi chú')
        note_val = str(row[note_idx].value) if note_idx is not None and row[note_idx].value else ""
        if "Lấy bằng OCR" in note_val:
            row_info['force_ocr_reprocess'] = True
            row_info['is_missing'] = True
            
        for col_name in required_cols:
            idx = col_idx.get(col_name)
            if idx is not None:
                val = str(row[idx].value).strip() if row[idx].value else ""
                if not val or val == "None" or val == "[Trống]":
                    row_info['is_missing'] = True
                    break
        
        if row_info['is_missing']:
            rows_to_process.append(row_info)

    unassigned_images = []
    if mode == "1" and 'image_paths' in locals():
        unassigned_images = [p for p in image_paths if os.path.basename(p) not in assigned_images]

    num_threads = 4
    if mode == "1":
        if not rows_to_process and not unassigned_images:
            console.print("[bold green]✅ File Excel đã đầy đủ thông tin và không có ảnh mới, chuyển sang bước chuẩn hóa địa chỉ![/bold green]")
        else:
            if rows_to_process:
                console.print(f"[bold yellow]⚠️ Tìm thấy {len(rows_to_process)} dòng bị thiếu thông tin.[/bold yellow]")
            if unassigned_images:
                console.print(f"[bold cyan]🔍 Tìm thấy {len(unassigned_images)} ảnh mới trong thư mục, sẽ quét để bổ sung.[/bold cyan]")
            
            # Cấu hình luồng xử lý ảnh
            num_threads_input = Prompt.ask("\n[cyan]Nhập số luồng xử lý ảnh song song[/cyan] (Enter để mặc định là 4)", default="4").strip()
            try:
                num_threads = int(num_threads_input) if num_threads_input else 4
            except ValueError:
                console.print("[yellow]⚠️ Giá trị không hợp lệ, sử dụng mặc định: 4 luồng.[/yellow]")
                num_threads = 4

    api_threads = 4
    if normalize_address:
        api_threads_input = Prompt.ask("[cyan]Nhập số luồng gọi API địa chỉ song song[/cyan] (Enter để mặc định là 4)", default="4").strip()
        try:
            api_threads = int(api_threads_input) if api_threads_input else 4
        except ValueError:
            console.print("[yellow]Số luồng không hợp lệ, dùng mặc định 4[/yellow]")
            api_threads = 4
        
    if IN_COLAB:
        confirm = True
    else:
        confirm = Confirm.ask("\n[bold yellow]Bạn có muốn bắt đầu TÁI XỬ LÝ ngay bây giờ không?[/bold yellow]", default=True)
    if not confirm:
        console.print("[yellow]Đã hủy quá trình.[/yellow]")
        return
    
    # Đặt lại start_time để loại bỏ thời gian chờ người dùng nhập lệnh
    start_time = time.time()
    
    # Process images for missing rows
    import concurrent.futures
    
    # Helper to reprocess a single image
    def process_single_image(img_path):
        import time
        t0 = time.time()
        qr_string, engine, err, img, qr_rotated_img = extract_qr_data(img_path)
        log_msgs = []
        row_data = {}
        if qr_string:
            log_msgs.append(f"[green]✅ [Đã quét mã QR bằng {engine}]:[/green] {qr_string}")
            extracted, validation_notes = process_qr_string(qr_string)
            row_data.update(extracted)
        else:
            if img is not None:
                log_msgs.append(f"[yellow]⚠️ Không đọc được QR, đang thử quét OCR...[/yellow]")
                # Mỗi thread có instance model riêng (thread-local) nên không cần lock
                ocr_data, ocr_note, rotated_img = extract_ocr_data(img)
                
                parts = []
                side = ocr_data.get('OCR Side')
                if side: parts.append(f"[{side}]")
                parts.append(f"CCCD: {ocr_data.get('CCCD') or '[Trống]'}")
                parts.append(f"Tên: {ocr_data.get('Họ tên') or '[Trống]'}")
                parts.append(f"NS: {ocr_data.get('Ngày sinh') or '[Trống]'}")
                parts.append(f"GT: {ocr_data.get('Giới tính') or '[Trống]'}")
                parts.append(f"Ngày cấp: {ocr_data.get('Ngày cấp CCCD') or '[Trống]'}")
                
                if side == 'Front':
                    addr = ocr_data.get('Nơi thường trú gốc') or '[Trống]'
                    parts.append(f"Địa chỉ: {addr}")
                elif side == 'Back':
                    back_raw = ocr_data.get('back_side_raw', {})
                    mrz_lines = back_raw.get('mrz_lines', [])
                    if mrz_lines:
                        parts.append(f"MRZ ({len(mrz_lines)} dòng): {' | '.join(mrz_lines[:3])}")
                    else:
                        parts.append(f"MRZ: [Trống]")
                else:
                    addr = ocr_data.get('Nơi thường trú gốc') or '[Trống]'
                    parts.append(f"Địa chỉ: {addr}")
                    back_raw = ocr_data.get('back_side_raw', {})
                    mrz_lines = back_raw.get('mrz_lines', [])
                    if mrz_lines:
                        parts.append(f"MRZ ({len(mrz_lines)} dòng): {' | '.join(mrz_lines[:3])}")
                
                ocr_print_info = ", ".join(parts)
                log_msgs.append(f"[blue]ℹ️ Kết quả OCR:[/blue] {ocr_print_info} | Note: {ocr_note}")
                
                if DEBUG_MODE and ocr_data.get('Raw Text'):
                    log_msgs.append(f"[magenta]🐛 DEBUG RAW OCR TEXT:\n{ocr_data['Raw Text']}[/magenta]")
                
                row_data.update(ocr_data)
        
        row_data['Nơi cấp'] = get_place_of_issue(row_data.get('QR Raw', ''))
        row_data['Ngày hết hạn'] = calculate_expiry_date(row_data.get('Ngày sinh', ''))
        row_data['_processing_time'] = time.time() - t0
        return row_data, log_msgs

    # We need to process both front and back images for each row
    all_images_to_process = set()
    img_results = {}
    recovered_data = {}
    
    reprocess_tmp = excel_path.replace('.xlsx', '_reprocess_recovery.jsonl')
    if mode == "1" and os.path.exists(reprocess_tmp):
        try:
            with open(reprocess_tmp, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.strip():
                        data = json.loads(line.strip())
                        recovered_data.update(data)
            if recovered_data:
                console.print(f"[bold green]✅ Đã phục hồi {len(recovered_data)} kết quả quét ảnh bị gián đoạn từ {os.path.basename(reprocess_tmp)}.[/bold green]")
        except Exception as e:
            console.print(f"[yellow]⚠️ Lỗi đọc file phục hồi ảnh: {e}[/yellow]")
            
    api_recovery_file = excel_path.replace('.xlsx', '_api_recovery.jsonl')
    recovered_api = {}
    if os.path.exists(api_recovery_file):
        try:
            with open(api_recovery_file, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.strip():
                        data = json.loads(line.strip())
                        recovered_api.update(data)
            if recovered_api:
                console.print(f"[bold green]✅ Đã phục hồi {len(recovered_api)} kết quả API bị gián đoạn từ {os.path.basename(api_recovery_file)}.[/bold green]")
        except Exception as e:
            console.print(f"[yellow]⚠️ Lỗi đọc file phục hồi API: {e}[/yellow]")
    
    if mode == "1":
        for row in rows_to_process:
            if row['front_name'] and row['front_name'] in img_map:
                p = img_map[row['front_name']]
                if p in recovered_data: img_results[p] = recovered_data[p]
                else: all_images_to_process.add(p)
            if row['back_name'] and row['back_name'] in img_map:
                p = img_map[row['back_name']]
                if p in recovered_data: img_results[p] = recovered_data[p]
                else: all_images_to_process.add(p)
            if row.get('other_name'):
                for part in str(row['other_name']).split(','):
                    part_clean = os.path.basename(part.strip())
                    if part_clean and part_clean in img_map:
                        p = img_map[part_clean]
                        if p in recovered_data: img_results[p] = recovered_data[p]
                        else: all_images_to_process.add(p)
                
        # Thêm ảnh unassigned vào danh sách xử lý
        for p in unassigned_images:
            if p in recovered_data: img_results[p] = recovered_data[p]
            else: all_images_to_process.add(p)
                
        with Progress(
            SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
            BarColumn(), TaskProgressColumn(), CountColumn(), TimeElapsedColumn(),
            TextColumn("⏳ ETA:"), ETAColumn(), SpeedColumn(),
            console=console,
        ) as progress:
            task_id = progress.add_task("[cyan]Đang quét ảnh...", total=len(all_images_to_process))
            executor = concurrent.futures.ThreadPoolExecutor(max_workers=num_threads)
            future_to_img = {executor.submit(process_single_image, path): path for path in all_images_to_process}
            for future in concurrent.futures.as_completed(future_to_img):
                img_path = future_to_img[future]
                try:
                    row_data, log_msgs = future.result(timeout=600)
                except concurrent.futures.TimeoutError:
                    err = f"❌ Lỗi: Ảnh {os.path.basename(img_path)} làm treo AI quá 600s (10 phút). Bỏ qua!"
                    progress.console.print(f"[bold red]{err}[/bold red]")
                    file_logs.append(err)
                    progress.advance(task_id)
                    continue
                except Exception as exc:
                    err = f"❌ Lỗi khi xử lý ảnh {os.path.basename(img_path)}: {exc}"
                    progress.console.print(f"[bold red]{err}[/bold red]")
                    file_logs.append(err)
                    progress.advance(task_id)
                    continue
                    
                img_results[img_path] = row_data
                    
                try:
                    with open(reprocess_tmp, 'a', encoding='utf-8') as f:
                        json.dump({img_path: row_data}, f, ensure_ascii=False)
                        f.write('\n')
                except: pass
                
                p_time = row_data.get('_processing_time', 0)
                progress.console.print(f"[bold][{os.path.basename(img_path)}][/bold] - [dim]{img_path}[/dim] - [yellow]Timing {p_time:.1f}s[/yellow]")
                file_logs.append(f"[{os.path.basename(img_path)}] - {img_path} - Timing {p_time:.1f}s")
                for msg in log_msgs:
                    progress.console.print(f"  {msg}")
                    file_logs.append("  " + Text.from_markup(msg).plain)
                    
                progress.advance(task_id)
                
            executor.shutdown(wait=False, cancel_futures=True)

    # Merge results back into Excel rows for Mode 1
    if mode == "1":
        # We prioritize keeping existing non-empty values, but overwrite if OCR found new data
        for row_info in rows_to_process:
            front_path = img_map.get(row_info['front_name']) if row_info['front_name'] else None
            back_path = img_map.get(row_info['back_name']) if row_info['back_name'] else None
            
            front_data = img_results.get(front_path, {})
            back_data = img_results.get(back_path, {})
            
            other_data = {}
            if row_info.get('other_name'):
                for part in str(row_info['other_name']).split(','):
                    part_clean = os.path.basename(part.strip())
                    if part_clean and part_clean in img_map:
                        p = img_map[part_clean]
                        if p in img_results:
                            for k, v in img_results[p].items():
                                if v and not other_data.get(k):
                                    other_data[k] = v
            
            # Merge logic
            for col_name in required_cols:
                idx = col_idx.get(col_name)
                if idx is None: continue
                
                existing_val = str(row_info['row_cells'][idx].value).strip() if row_info['row_cells'][idx].value else ""
                is_empty = not existing_val or existing_val == "None" or existing_val == "[Trống]"
                
                # Cố lấy từ front_data, back_data hoặc other_data
                new_val = front_data.get(col_name) or back_data.get(col_name) or other_data.get(col_name)
                
                # Giữ nguyên giá trị cũ nếu đã có, TRỪ KHI dòng này được đánh dấu force_ocr_reprocess và có data mới
                if not is_empty and not row_info.get('force_ocr_reprocess'):
                    continue
                    
                if new_val:
                    row_info['row_cells'][idx].value = new_val
                    row_info['row_cells'][idx].font = Font(color="FF0000")

    # LÀM SẠCH VÀ CHUẨN BỊ GỌI API CHO FILE EXCEL
    address_to_normalize = set()
    orig_idx = col_idx.get('Nơi thường trú gốc')
    norm_idx = col_idx.get('Địa chỉ chuẩn hóa mới')
    note_idx = col_idx.get('Ghi chú')

    name_idx = col_idx.get('Họ tên')
    
    if orig_idx is not None:
        for row in ws.iter_rows(min_row=2, values_only=False):
            val = row[orig_idx].value
            note_val = str(row[note_idx].value) if note_idx and row[note_idx].value else ""
            
            # Tô nền vàng nếu lấy bằng OCR hoặc không đọc được
            if "Lấy bằng OCR" in note_val or "Không đọc được" in note_val:
                yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
                for cell in row:
                    cell.fill = yellow_fill
            
            if mode == "1":
                # Làm sạch Họ tên nếu lấy bằng OCR
                if "Lấy bằng OCR" in note_val and name_idx is not None:
                    name_val = row[name_idx].value
                    if name_val and str(name_val).strip() and str(name_val).strip() != "None":
                        cleaned_name = clean_name_string(str(name_val))
                        if cleaned_name:
                            if cleaned_name != str(name_val):
                                file_logs.append(f"[LÀM SẠCH HỌ TÊN] {str(name_val)} -> {cleaned_name}")
                            row[name_idx].value = cleaned_name
                            row[name_idx].font = Font(color="FF0000")

            if val and str(val).strip() and str(val).strip() != "None":
                raw_addr = str(val).strip()
                norm_val = str(row[norm_idx].value).strip() if norm_idx and row[norm_idx].value else ""
                is_norm_empty = not norm_val or norm_val == "None" or norm_val == ""
                
                is_ocr = "Lấy bằng OCR" in note_val
                
                if mode == "1":
                    if is_ocr:
                        cleaned_val = clean_address_string(raw_addr)
                        if cleaned_val != raw_addr:
                            file_logs.append(f"[LÀM SẠCH ĐỊA CHỈ GỐC] {raw_addr} -> {cleaned_val}")
                        row[orig_idx].value = cleaned_val
                        row[orig_idx].font = Font(color="FF0000")
                        address_to_normalize.add(cleaned_val)
                    else:
                        if is_norm_empty:
                            address_to_normalize.add(raw_addr)
                elif mode == "2":
                    if is_ocr:
                        should_process_address = process_all_rows or is_norm_empty
                        if should_process_address:
                            cleaned_val = clean_address_string(raw_addr)
                            if cleaned_val != raw_addr:
                                file_logs.append(f"[LÀM SẠCH ĐỊA CHỈ GỐC] {raw_addr} -> {cleaned_val}")
                            row[orig_idx].value = cleaned_val
                            row[orig_idx].font = Font(color="FF0000")
                            address_to_normalize.add(cleaned_val)
                    else:
                        # QR: KHÔNG làm sạch. CHỈ gọi API nếu thiếu
                        if is_norm_empty:
                            address_to_normalize.add(raw_addr)
                elif mode == "3":
                    if is_ocr:
                        should_process_address = process_all_rows or is_norm_empty
                        if should_process_address:
                            address_to_normalize.add(raw_addr)
                    else:
                        # QR: KHÔNG làm sạch. CHỈ gọi API nếu thiếu
                        if is_norm_empty:
                            address_to_normalize.add(raw_addr)

    # Nơi chuẩn hóa địa chỉ
    address_map = recovered_api.copy()
    if normalize_address and address_to_normalize:
        
        # Chỉ gọi API cho những địa chỉ chưa có kết quả (bao gồm cả thành công và thất bại)
        pending_addresses = [addr for addr in address_to_normalize if addr not in address_map]
        
        if pending_addresses:
            console.print(Panel(f"[bold cyan]🌐 ĐANG CHUẨN BỊ GỌI API CHUẨN HÓA CHO {len(pending_addresses)} ĐỊA CHỈ DUY NHẤT VỚI {api_threads} LUỒNG...[/bold cyan]", border_style="green"))
            unique_addresses = pending_addresses
            batch_size = 100
            total_addrs = len(unique_addresses)
        processed_count = 0
        with Progress(
            SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
            BarColumn(), TaskProgressColumn(), CountColumn(), TextColumn("[bold]{task.fields[status]}"),
            TimeElapsedColumn(), TextColumn("⏳ ETA:"), TimeRemainingColumn(), SpeedColumn(), console=console,
        ) as api_progress:
            api_task = api_progress.add_task("[cyan]Đang chuẩn hóa địa chỉ...", total=total_addrs, status="")
            for i in range(0, total_addrs, batch_size):
                batch = unique_addresses[i:i+batch_size]
                with concurrent.futures.ThreadPoolExecutor(max_workers=api_threads) as executor:
                    future_to_addr = {executor.submit(fetch_single_address, addr): addr for addr in batch}
                    for future in concurrent.futures.as_completed(future_to_addr):
                        result = future.result()
                        processed_count += 1
                        if result and 'original' in result:
                            orig_addr = result['original']
                            address_map[orig_addr] = result
                            short_addr = orig_addr[:45]
                            
                            try:
                                with open(api_recovery_file, 'a', encoding='utf-8') as f:
                                    json.dump({orig_addr: result}, f, ensure_ascii=False)
                                    f.write('\n')
                            except: pass
                            
                            if result.get('success'):
                                new_addr = result.get('converted', '')
                                status_text = f"[green]✓[/green] {short_addr}"
                                file_logs.append(f"[{processed_count}/{total_addrs}] CHUẨN HÓA OK: {orig_addr} -> {new_addr}")
                                if DEBUG_MODE and not IN_COLAB:
                                    api_progress.console.print(f"[bold cyan][{processed_count}/{total_addrs}][/bold cyan] [dim]Từ:[/dim] [yellow]{orig_addr}[/yellow]\n{' '*(len(str(total_addrs))*2 + 5)}[dim]→  [/dim] [bold green]{new_addr}[/bold green]")
                            else:
                                err_msg = result.get('error', 'Lỗi không xác định')
                                status_text = f"[red]✗[/red] {short_addr}"
                                file_logs.append(f"[{processed_count}/{total_addrs}] CHUẨN HÓA LỖI: {orig_addr} -> {err_msg}")
                                if DEBUG_MODE and not IN_COLAB:
                                    api_progress.console.print(f"[bold cyan][{processed_count}/{total_addrs}][/bold cyan] [dim]Từ:[/dim] [yellow]{orig_addr}[/yellow]\n{' '*(len(str(total_addrs))*2 + 5)}[dim]→  [/dim] [bold red]{err_msg}[/bold red]")
                                
                            api_progress.update(api_task, advance=1, status=status_text)
                            
        # Xử lý Retry: Làm đẹp các địa chỉ gọi API thất bại và thử lại
        retry_addresses = set()
        retry_to_orig = {}
        for addr, result in address_map.items():
            if not result.get('success'):
                cleaned_addr = clean_address_string(addr)
                if cleaned_addr and cleaned_addr != addr:
                    retry_addresses.add(cleaned_addr)
                    if cleaned_addr not in retry_to_orig:
                        retry_to_orig[cleaned_addr] = []
                    retry_to_orig[cleaned_addr].append(addr)
                    
        if retry_addresses:
            console.print(Panel(f"[bold yellow]🔄 ĐANG GỌI API LẠI CHO {len(retry_addresses)} ĐỊA CHỈ SAU KHI LÀM ĐẸP...[/bold yellow]", border_style="yellow"))
            unique_retries = list(retry_addresses)
            total_retries = len(unique_retries)
            processed_count = 0
            with Progress(
                SpinnerColumn(), TextColumn("[progress.description]{task.description}"),
                BarColumn(), TaskProgressColumn(), CountColumn(), TextColumn("[bold]{task.fields[status]}"),
                TimeElapsedColumn(), TextColumn("⏳ ETA:"), TimeRemainingColumn(), SpeedColumn(), console=console,
            ) as retry_progress:
                retry_task = retry_progress.add_task("[yellow]Đang thử lại...", total=total_retries, status="")
                for i in range(0, total_retries, batch_size):
                    batch = unique_retries[i:i+batch_size]
                    with concurrent.futures.ThreadPoolExecutor(max_workers=api_threads) as executor:
                        future_to_addr = {executor.submit(fetch_single_address, addr): addr for addr in batch}
                        for future in concurrent.futures.as_completed(future_to_addr):
                            result = future.result()
                            processed_count += 1
                            if result and 'original' in result:
                                cleaned_addr = result['original']
                                short_addr = cleaned_addr[:45]
                                
                                try:
                                    with open(api_recovery_file, 'a', encoding='utf-8') as f:
                                        json.dump({cleaned_addr: result}, f, ensure_ascii=False)
                                        f.write('\n')
                                except: pass
                                
                                if result.get('success'):
                                    new_addr = result.get('converted', '')
                                    status_text = f"[green]✓[/green] {short_addr}"
                                    file_logs.append(f"[LÀM SẠCH] [{processed_count}/{total_retries}] LÀM ĐẸP & CHUẨN HÓA OK: {cleaned_addr} -> {new_addr}")
                                    if DEBUG_MODE and not IN_COLAB:
                                        retry_progress.console.print(f"[bold cyan][{processed_count}/{total_retries}][/bold cyan] [dim]Từ:[/dim] [yellow]{cleaned_addr}[/yellow]\n{' '*(len(str(total_retries))*2 + 5)}[dim]→  [/dim] [bold green]{new_addr}[/bold green]")
                                    result['final_cleaned_orig'] = cleaned_addr
                                    for orig_addr in retry_to_orig.get(cleaned_addr, []):
                                        address_map[orig_addr] = result
                                else:
                                    err_msg = result.get('error', 'Lỗi không xác định')
                                    status_text = f"[red]✗[/red] {short_addr}"
                                    file_logs.append(f"[LÀM SẠCH] [{processed_count}/{total_retries}] LÀM ĐẸP & CHUẨN HÓA LỖI: {cleaned_addr} -> {err_msg}")
                                    if DEBUG_MODE and not IN_COLAB:
                                        retry_progress.console.print(f"[bold cyan][{processed_count}/{total_retries}][/bold cyan] [dim]Từ:[/dim] [yellow]{cleaned_addr}[/yellow]\n{' '*(len(str(total_retries))*2 + 5)}[dim]→  [/dim] [bold red]{err_msg}[/bold red]")
                                retry_progress.update(retry_task, advance=1, status=status_text)
                                
        # Điền lại địa chỉ chuẩn hóa vào Excel
        norm_idx = col_idx.get('Địa chỉ chuẩn hóa mới')
        orig_idx = col_idx.get('Nơi thường trú gốc')
        note_idx = col_idx.get('Ghi chú')
        if norm_idx is not None and orig_idx is not None:
            for row in ws.iter_rows(min_row=2, values_only=False):
                orig_val = row[orig_idx].value
                # Chỉ điền nếu gọi API thành công
                if orig_val and orig_val in address_map and address_map[orig_val].get('success'):
                    norm_val = row[norm_idx].value
                    note_val = str(row[note_idx].value) if note_idx and row[note_idx].value else ""
                    # Đè lên nếu trước đó chưa có, HOẶC nếu dòng đó được Lấy bằng OCR
                    if "Lấy bằng OCR" in note_val or not norm_val or str(norm_val).strip() == "" or str(norm_val).strip() == "None":
                        row[norm_idx].value = address_map[orig_val].get('converted', '')
                        row[norm_idx].font = Font(color="FF0000")
                        
                        # Nếu địa chỉ này được cứu nhờ bước làm sạch (retry), cập nhật luôn cột gốc
                        if address_map[orig_val].get('final_cleaned_orig'):
                            row[orig_idx].value = address_map[orig_val]['final_cleaned_orig']
                            row[orig_idx].font = Font(color="FF0000")

    # ---------- AUTOMATIC ROW MERGING FOR REPROCESSING (MODE 1, 2, 3) ----------
    if mode in ("1", "2", "3"):
        console.print("[cyan]🔄 Đang tự động gộp các dòng trùng lặp / mồ côi...[/cyan]")
        import unicodedata
        import re
        import difflib

        def _norm_for_match(text):
            if not text: return ''
            text = str(text).upper().strip()
            nfkd = unicodedata.normalize('NFKD', text)
            return ' '.join(''.join(c for c in nfkd if not unicodedata.combining(c)).split())

        def _is_similar_cccd(c1, c2):
            if not c1 or not c2: return False
            c1_clean = re.sub(r'\D', '', str(c1))
            c2_clean = re.sub(r'\D', '', str(c2))
            if not c1_clean or not c2_clean: return False
            if len(c1_clean) == 12 and len(c2_clean) == 12:
                diffs = sum(1 for a, b in zip(c1_clean, c2_clean) if a != b)
                return diffs <= 3
            return c1_clean == c2_clean

        def _is_similar_name(n1, n2):
            if not n1 or not n2: return False
            ratio = difflib.SequenceMatcher(None, n1, n2).ratio()
            if ratio < 0.80:
                return False
            w1 = n1.split()[-1] if n1.split() else ''
            w2 = n2.split()[-1] if n2.split() else ''
            if not w1 or not w2:
                return False
            w_ratio = difflib.SequenceMatcher(None, w1, w2).ratio()
            return w_ratio >= 0.50

        def _is_similar_dob(d1, d2):
            if not d1 or not d2: return False
            d1_str = str(d1).strip()
            d2_str = str(d2).strip()
            if d1_str == d2_str: return True
            p1 = d1_str.split('/')
            p2 = d2_str.split('/')
            if len(p1) == 3 and len(p2) == 3:
                # So khớp cùng tháng/năm
                if p1[1] == p2[1] and p1[2] == p2[2]:
                    return True
            return False

        def _is_invalid_cccd_placeholder(cccd):
            if not cccd: return True
            cccd_clean = re.sub(r'\D', '', str(cccd))
            if len(cccd_clean) != 12: return True
            if cccd_clean.startswith('000'): return True
            return False

        def _is_red_font(cell):
            if not cell or not cell.font or not cell.font.color:
                return False
            color = cell.font.color
            rgb_str = str(color.rgb or color.value or '')
            return rgb_str.endswith('FF0000') or rgb_str == 'FF0000'

        # Đọc dữ liệu hiện tại từ ws thành list các dict
        records_list = []
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            rec = {}
            for name in headers:
                idx = col_idx.get(name)
                cell = row[idx] if idx is not None else None
                rec[name] = cell.value if cell else None
                rec[f'_red_{name}'] = _is_red_font(cell)
            
            # Bỏ qua dòng trống hoàn toàn
            has_data = any(rec.get(h) is not None and str(rec.get(h)).strip() != "" for h in headers if h != 'STT')
            if not has_data:
                continue

            qr_raw = str(rec.get('QR Raw') or '').strip()
            rec['has_qr_data'] = bool(qr_raw and qr_raw != 'None' and qr_raw != '[Trống]')
            
            note_val = str(rec.get('Ghi chú') or '').strip()
            rec['has_ocr_data'] = 'Lấy bằng OCR' in note_val
            
            raw_notes = rec.get('Ghi chú') or ""
            rec['_notes_list'] = [n.strip() for n in str(raw_notes).split(';') if n.strip()]
            
            records_list.append(rec)

        def _is_better_record(r1, r2):
            if r1.get('has_qr_data') and not r2.get('has_qr_data'):
                return True
            if r2.get('has_qr_data') and not r1.get('has_qr_data'):
                return False
                
            c1_invalid = _is_invalid_cccd_placeholder(r1.get('CCCD'))
            c2_invalid = _is_invalid_cccd_placeholder(r2.get('CCCD'))
            if not c1_invalid and c2_invalid:
                return True
            if not c2_invalid and c1_invalid:
                return False
                
            fields_to_check = ['Họ tên', 'Ngày sinh', 'Nơi thường trú gốc', 'Ngày cấp CCCD']
            score1 = sum(1 for f in fields_to_check if r1.get(f) and str(r1.get(f)).strip() not in ['', 'None', 'Chưa xác định'])
            score2 = sum(1 for f in fields_to_check if r2.get(f) and str(r2.get(f)).strip() not in ['', 'None', 'Chưa xác định'])
            if score1 != score2:
                return score1 > score2
                
            img_fields = ['Ảnh mặt trước CCCD/CC', 'Ảnh mặt sau CCCD/CC', 'Ảnh khác (SMS/Chụp màn hình/...)']
            img_score1 = sum(1 for f in img_fields if r1.get(f) and str(r1.get(f)).strip() not in ['', 'None'])
            img_score2 = sum(1 for f in img_fields if r2.get(f) and str(r2.get(f)).strip() not in ['', 'None'])
            if img_score1 != img_score2:
                return img_score1 > img_score2
                
            return True

        def _should_merge(r1, r2, all_records):
            gen1 = str(r1.get('Giới tính') or '').strip().lower()
            gen2 = str(r2.get('Giới tính') or '').strip().lower()
            if gen1 and gen2 and gen1 in ['nam', 'nữ'] and gen2 in ['nam', 'nữ'] and gen1 != gen2:
                return False
                
            n1 = _norm_for_match(r1.get('Họ tên', ''))
            n2 = _norm_for_match(r2.get('Họ tên', ''))
            if not n1 or not n2:
                return False
                
            if not _is_similar_name(n1, n2):
                return False
                
            dob_match = _is_similar_dob(r1.get('Ngày sinh'), r2.get('Ngày sinh'))
            issue_match = _is_similar_dob(r1.get('Ngày cấp CCCD'), r2.get('Ngày cấp CCCD'))
            cccd_match = _is_similar_cccd(r1.get('CCCD'), r2.get('CCCD'))
            
            # Kiểm tra tính duy nhất của tên trong database
            similar_count = 0
            for r in all_records:
                rn = _norm_for_match(r.get('Họ tên', ''))
                if rn and _is_similar_name(n1, rn):
                    similar_count += 1
            is_unique_name = (similar_count == 2)
            
            if dob_match or issue_match or cccd_match or is_unique_name:
                c1 = r1.get('CCCD')
                c2 = r2.get('CCCD')
                if not _is_invalid_cccd_placeholder(c1) and not _is_invalid_cccd_placeholder(c2):
                    if not _is_similar_cccd(c1, c2):
                        return False
                if r1.get('has_qr_data') and r2.get('has_qr_data'):
                    if not _is_similar_cccd(c1, c2):
                        return False
                return True
                
            return False

        def merge_two_records(target, source):
            for k in ['Họ tên', 'CCCD', 'CMND', 'Giới tính', 'Ngày sinh', 'Nơi thường trú gốc', 'Địa chỉ chuẩn hóa mới', 'Ngày cấp CCCD', 'Nơi cấp', 'Ngày hết hạn', 'Phân loại', 'QR Raw']:
                target_val = str(target.get(k) or '').strip()
                source_val = str(source.get(k) or '').strip()
                
                is_target_empty = not target_val or target_val.lower() in ['none', '[trống]', 'chưa xác định']
                is_source_empty = not source_val or source_val.lower() in ['none', '[trống]', 'chưa xác định']
                
                if k == 'CCCD':
                    if _is_invalid_cccd_placeholder(target.get('CCCD')) and not _is_invalid_cccd_placeholder(source.get('CCCD')):
                        target['CCCD'] = source['CCCD']
                        target['_red_CCCD'] = source.get('_red_CCCD', False) or target.get('_red_CCCD', False)
                        continue
                        
                if is_target_empty and not is_source_empty:
                    target[k] = source[k]
                    target[f'_red_{k}'] = source.get(f'_red_{k}', False)
                    
            # Gộp ảnh
            target_front = str(target.get('Ảnh mặt trước CCCD/CC') or '').strip()
            source_front = str(source.get('Ảnh mặt trước CCCD/CC') or '').strip()
            if not target_front or target_front.lower() == 'none':
                target['Ảnh mặt trước CCCD/CC'] = source_front
                target['_red_Ảnh mặt trước CCCD/CC'] = source.get('_red_Ảnh mặt trước CCCD/CC', False)
            elif source_front and source_front.lower() != 'none' and target_front != source_front:
                existing = str(target.get('Ảnh khác (SMS/Chụp màn hình/...)') or '').strip()
                if existing and existing.lower() != 'none':
                    target['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {source_front}"
                else:
                    target['Ảnh khác (SMS/Chụp màn hình/...)'] = source_front
                    
            target_back = str(target.get('Ảnh mặt sau CCCD/CC') or '').strip()
            source_back = str(source.get('Ảnh mặt sau CCCD/CC') or '').strip()
            if not target_back or target_back.lower() == 'none':
                target['Ảnh mặt sau CCCD/CC'] = source_back
                target['_red_Ảnh mặt sau CCCD/CC'] = source.get('_red_Ảnh mặt sau CCCD/CC', False)
            elif source_back and source_back.lower() != 'none' and target_back != source_back:
                existing = str(target.get('Ảnh khác (SMS/Chụp màn hình/...)') or '').strip()
                if existing and existing.lower() != 'none':
                    target['Ảnh khác (SMS/Chụp màn hình/...)'] = f"{existing}, {source_back}"
                else:
                    target['Ảnh khác (SMS/Chụp màn hình/...)'] = source_back

            target_other = str(target.get('Ảnh khác (SMS/Chụp màn hình/...)') or '').strip()
            source_other = str(source.get('Ảnh khác (SMS/Chụp màn hình/...)') or '').strip()
            unique_others = []
            if target_other and target_other.lower() != 'none':
                for o in target_other.split(','):
                    o_clean = o.strip()
                    if o_clean and o_clean not in unique_others:
                        unique_others.append(o_clean)
            if source_other and source_other.lower() != 'none':
                for o in source_other.split(','):
                    o_clean = o.strip()
                    if o_clean and o_clean not in unique_others:
                        unique_others.append(o_clean)
            target['Ảnh khác (SMS/Chụp màn hình/...)'] = ", ".join(unique_others)
            
            target_notes = target.get('_notes_list', [])
            source_notes = source.get('_notes_list', [])
            for n in source_notes:
                if n not in target_notes:
                    target_notes.append(n)
            target['_notes_list'] = target_notes
            target['Ghi chú'] = "; ".join(target_notes)
            
            for flag in ['has_qr_data', 'has_ocr_data']:
                if source.get(flag):
                    target[flag] = True

        merged_any = True
        while merged_any:
            merged_any = False
            n_rec = len(records_list)
            pair_found = None
            for i in range(n_rec):
                for j in range(i + 1, n_rec):
                    if _should_merge(records_list[i], records_list[j], records_list):
                        pair_found = (i, j)
                        break
                if pair_found:
                    break
            if pair_found:
                i, j = pair_found
                rec1 = records_list[i]
                rec2 = records_list[j]
                if _is_better_record(rec1, rec2):
                    target, source = rec1, rec2
                    source_idx = j
                else:
                    target, source = rec2, rec1
                    source_idx = i
                
                console.print(f"   [bold green]→ [Reprocess Merge][/bold green] Ghép bản ghi {source.get('CCCD')} ({source.get('Họ tên')}) vào bản ghi {target.get('CCCD')} ({target.get('Họ tên')})")
                file_logs.append(f"[GỘP TÁI XỬ LÝ] Ghép bản ghi {source.get('CCCD')} ({source.get('Họ tên')}) vào bản ghi {target.get('CCCD')} ({target.get('Họ tên')})")
                
                merge_two_records(target, source)
                records_list.pop(source_idx)
                merged_any = True

        # Regenerate renaming columns for merged records
        for record in records_list:
            hoten = record.get('Họ tên') or 'KhongTen'
            if hoten and hoten != 'KhongTen':
                hoten = hoten.title()
                record['Họ tên'] = hoten
                
            cccd = record.get('CCCD') or ''
            cmnd = record.get('CMND') or ''
            hoten_clean = re.sub(r'[\\/*?:"<>|]', '', hoten)
            cmnd_clean = str(cmnd or '').strip()
            if cmnd_clean.lower() in ['', 'none', 'chưa xác định', 'không có']:
                cmnd_str = ""
            else:
                cmnd_str = f"_{cmnd_clean}"
                
            front_img = record.get('Ảnh mặt trước CCCD/CC')
            if front_img and str(front_img).strip() and str(front_img).strip().lower() != 'none':
                ext = os.path.splitext(str(front_img).strip())[1]
                record['Đổi tên Ảnh mặt trước CCCD/CC'] = f"{hoten_clean}_{cccd}{cmnd_str}_Mặt trước{ext}"
            else:
                record['Đổi tên Ảnh mặt trước CCCD/CC'] = ''
                
            back_img = record.get('Ảnh mặt sau CCCD/CC')
            if back_img and str(back_img).strip() and str(back_img).strip().lower() != 'none':
                ext = os.path.splitext(str(back_img).strip())[1]
                record['Đổi tên Ảnh mặt sau CCCD/CC'] = f"{hoten_clean}_{cccd}{cmnd_str}_Mặt sau{ext}"
            else:
                record['Đổi tên Ảnh mặt sau CCCD/CC'] = ''

            # Special logic for "Khác" category
            if record.get('Phân loại') == 'Khác':
                base_name = f"{hoten_clean}_{cccd}{cmnd_str}"
                imgs = []
                if record.get('Ảnh mặt trước CCCD/CC'): imgs.append(record['Ảnh mặt trước CCCD/CC'])
                if record.get('Ảnh mặt sau CCCD/CC'): imgs.append(record['Ảnh mặt sau CCCD/CC'])
                
                anh_khac = record.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                anh_khac_list = []
                if anh_khac: anh_khac_list.extend(str(anh_khac).split(', '))
                for img in imgs:
                    if img not in anh_khac_list:
                        anh_khac_list.append(img)
                
                renamed_list = []
                for i, p in enumerate(anh_khac_list):
                    ext = os.path.splitext(p)[1]
                    suffix = f"_Khác_{i+1}" if len(anh_khac_list) > 1 else "_Khác"
                    renamed_list.append(f"{base_name}{suffix}{ext}")
                
                record['Ảnh khác (SMS/Chụp màn hình/...)'] = ", ".join(filter(None, anh_khac_list))
                record['Đổi tên Ảnh khác'] = ", ".join(renamed_list)
                record['Ảnh mặt trước CCCD/CC'] = ''
                record['Ảnh mặt sau CCCD/CC'] = ''
                record['Đổi tên Ảnh mặt trước CCCD/CC'] = ''
                record['Đổi tên Ảnh mặt sau CCCD/CC'] = ''
            else:
                anh_khac = record.get('Ảnh khác (SMS/Chụp màn hình/...)', '')
                if anh_khac and str(anh_khac).strip() and str(anh_khac).strip().lower() != 'none':
                    anh_khac_list = [p.strip() for p in str(anh_khac).split(', ') if p.strip()]
                    renamed_list = []
                    for i, p in enumerate(anh_khac_list):
                        ext = os.path.splitext(p)[1]
                        suffix = f"_Khác_{i+1}" if len(anh_khac_list) > 1 else "_Khác"
                        renamed_list.append(f"{hoten_clean}_{cccd}{cmnd_str}{suffix}{ext}")
                    record['Đổi tên Ảnh khác'] = ", ".join(renamed_list)
                else:
                    record['Đổi tên Ảnh khác'] = ''

        # Clear worksheet content (retaining headers)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
            
        # Write merged records back
        for idx, record in enumerate(records_list):
            row_data = []
            for name in headers:
                if name == 'STT':
                    row_data.append(idx + 1)
                else:
                    row_data.append(record.get(name))
            ws.append(row_data)
            
            current_row_idx = ws.max_row
            note_val = str(record.get('Ghi chú') or '')
            if "Lấy bằng OCR" in note_val or "Không đọc được" in note_val:
                yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
                for cell in ws[current_row_idx]:
                    cell.fill = yellow_fill
                    
            for name in headers:
                if record.get(f'_red_{name}'):
                    col_num = col_idx.get(name) + 1
                    ws.cell(row=current_row_idx, column=col_num).font = Font(color="FF0000")

    timestamp = datetime.datetime.now().strftime("%d%m%Y_%H%M%S")
    reprocess_out = excel_path.replace('.xlsx', f'_reprocessed_{timestamp}.xlsx')
    wb.save(reprocess_out)
    
    log_filename = excel_path.replace('.xlsx', f'_reprocess_log_{timestamp}.txt')
    file_logs.append("\nĐÃ HOÀN TẤT THÀNH CÔNG TÁI XỬ LÝ!")
    file_logs.append(f"File kết quả: {os.path.abspath(reprocess_out)}")
    with open(log_filename, "w", encoding="utf-8") as f:
        f.write("\n".join(file_logs))
        
    console.print("\n" + "🎉"*15)
    console.print(f"[bold green]ĐÃ HOÀN TẤT THÀNH CÔNG TÁI XỬ LÝ![/bold green]")
    console.print(f"File kết quả được lưu tại: [yellow]{os.path.abspath(reprocess_out)}[/yellow]")
    console.print(f"File log chi tiết được lưu tại: [yellow]{os.path.abspath(log_filename)}[/yellow]")
    console.print("🎉"*15 + "\n")
    
    # Xóa file tạm
    if os.path.exists(reprocess_tmp): os.remove(reprocess_tmp)
    if os.path.exists(api_recovery_file): os.remove(api_recovery_file)

    from rich.table import Table
    table = Table(title="📊 BÁO CÁO THỐNG KÊ TÁI XỬ LÝ", show_header=True, header_style="bold magenta")
    table.add_column("Chỉ số đo lường", style="cyan", width=42)
    table.add_column("Số lượng", justify="right", style="green", width=12)
    table.add_column("Tỉ lệ", justify="right", style="yellow", width=10)
    
    total_imgs = len(all_images_to_process) if mode == "1" else 0
    table.add_row("Tổng số dòng trong Excel", str(len(rows_to_process) if mode == "1" else ws.max_row - 1), "")
    
    if mode == "1":
        table.add_row("Tổng số ảnh đã đưa vào quét lại", str(total_imgs), "100.0%" if total_imgs else "0.0%")
        qr_count = sum(1 for r in img_results.values() if 'Lấy bằng OCR' not in str(r.get('Ghi chú', '')))
        ocr_count = sum(1 for r in img_results.values() if 'Lấy bằng OCR' in str(r.get('Ghi chú', '')))
        table.add_row("Số ảnh đọc mã QR thành công", str(qr_count), f"{(qr_count/total_imgs*100):.1f}%" if total_imgs else "0.0%")
        table.add_row("Số ảnh phải dùng AI OCR để cứu", str(ocr_count), f"{(ocr_count/total_imgs*100):.1f}%" if total_imgs else "0.0%")
    else:
        table.add_row("Số địa chỉ gốc đã xử lý", str(len(address_to_normalize)), "")
    
    end_time = time.time()
    elapsed = end_time - start_time
    hours, rem = divmod(elapsed, 3600)
    minutes, seconds = divmod(rem, 60)
    time_str = f"{int(hours)}h {int(minutes)}m {int(seconds)}s" if hours > 0 else f"{int(minutes)}m {int(seconds)}s"
    
    table.add_row("", "", "")
    table.add_row("[bold]HIỆU SUẤT XỬ LÝ[/bold]", "", "")
    table.add_row("Tổng thời gian thực thi", time_str, "")
    
    if mode == "1":
        processing_times = [r.get('_processing_time', 0) for r in img_results.values() if r.get('_processing_time')]
        if processing_times:
            table.add_row("Tốc độ OCR chậm nhất / 1 ảnh", f"{max(processing_times):.1f}s", "")
            table.add_row("Tốc độ OCR nhanh nhất / 1 ảnh", f"{min(processing_times):.1f}s", "")
            table.add_row("Tốc độ OCR trung bình / 1 ảnh", f"{(sum(processing_times) / len(processing_times)):.1f}s", "")
            
    if normalize_address and address_map:
        api_processing_times = [r.get('_processing_time', 0) for r in address_map.values() if r.get('_processing_time')]
        success_api_count = sum(1 for r in address_map.values() if r.get('success'))
        total_api_count = len(address_map)
        table.add_row("Số địa chỉ chuẩn hóa thành công", str(success_api_count), f"{(success_api_count/total_api_count*100):.1f}%" if total_api_count else "0.0%")
        
        if api_processing_times:
            table.add_row("Tốc độ gọi API chậm nhất / 1 dòng", f"{max(api_processing_times):.2f}s", "")
            table.add_row("Tốc độ gọi API nhanh nhất / 1 dòng", f"{min(api_processing_times):.2f}s", "")
            table.add_row("Tốc độ gọi API trung bình / 1 dòng", f"{(sum(api_processing_times) / len(api_processing_times)):.2f}s", "")
    
    console.print(table)
    console.print()

import threading
GEOVINA_TOKEN_LOCK = threading.Lock()

def check_and_prompt_geovina_token(current_failed_token=None):
    """Kiểm tra token Geovina xem còn hạn không và hướng dẫn người dùng nhập mới nếu hết hạn."""
    import os, requests
    
    with GEOVINA_TOKEN_LOCK:
        if current_failed_token is not None:
            latest_token = os.environ.get('GEOVINA_DEMO_TOKEN', '')
            if latest_token and latest_token != current_failed_token:
                return True # Đã có thread khác cập nhật token thành công
                
        def _test_token(token):
            try:
                geovina_headers = {
                    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:152.0) Gecko/20100101 Firefox/152.0',
                    'Accept': 'application/json',
                    'Accept-Language': 'en-US,vi;q=0.9,en;q=0.8',
                    'Content-Type': 'application/json',
                    'X-Demo-Token': token,
                    'X-Api-Key': 'gvn_5740dceda5cb2424b787f1153da3802a721ae3f6',
                    'Referer': 'https://www.geovina.io.vn/',
                    'Origin': 'https://www.geovina.io.vn'
                }
                res = requests.post(
                    'https://www.geovina.io.vn/parse',
                    headers=geovina_headers,
                    json={"address": "Phường Bến Nghé, Quận 1, TP. Hồ Chí Minh"},
                    timeout=5
                )
                return res.json().get('success', False)
            except:
                return False

        def _auto_fetch_token():
            try:
                html_resp = requests.get(
                    'https://www.geovina.io.vn/',
                    headers={
                        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:152.0) Gecko/20100101 Firefox/152.0',
                        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8'
                    },
                    timeout=10
                )
                import re
                match = re.search(r'window\.__DEMO_TOKEN__\s*=\s*[\'"]([^\'"]+)[\'"]', html_resp.text)
                if match:
                    return match.group(1)
            except:
                pass
            return None

    current_token = os.environ.get('GEOVINA_DEMO_TOKEN', '')
    
    # Lúc đầu chạy, nếu token trống, ta thử auto-fetch luôn
    if not current_token:
        with console.status("[bold green]Đang tự động lấy Token Geovina...", spinner="dots"):
            auto_tok = _auto_fetch_token()
            if auto_tok and _test_token(auto_tok):
                os.environ['GEOVINA_DEMO_TOKEN'] = auto_tok
                return True

    current_token = os.environ.get('GEOVINA_DEMO_TOKEN', '1782095814853.ab6fc10225b874be936bd6fe9a020c6e0a5418e03a1215c0463d5628c91083e7')

    with console.status("[bold green]Đang kiểm tra trạng thái API Geovina...", spinner="dots"):
        is_ok = _test_token(current_token)

    while not is_ok:
        console.print("\n[bold red]⚠️ Token dự phòng Geovina đã hết hạn hoặc không hợp lệ![/bold red]")
        
        with console.status("[bold green]Đang thử tự động gia hạn Token mới...", spinner="dots"):
            auto_tok = _auto_fetch_token()
            if auto_tok and _test_token(auto_tok):
                os.environ['GEOVINA_DEMO_TOKEN'] = auto_tok
                console.print("[bold green]✅ Đã tự động gia hạn thành công Token Geovina![/bold green]\n")
                return True
                
        console.print("[yellow]Tự động gia hạn thất bại. Hướng dẫn lấy Token thủ công:[/yellow]")
        console.print("1. Mở trình duyệt web truy cập: [cyan]https://www.geovina.io.vn/[/cyan]")
        console.print("2. Nhấn F12 (hoặc Chuột phải -> Inspect) để mở Developer Tools, chuyển sang tab [cyan]Network[/cyan].")
        console.print("3. Bấm nút [cyan]Tách địa chỉ[/cyan] trên trang web.")
        console.print("4. Trong tab Network, click vào request có tên là [cyan]parse[/cyan].")
        console.print("5. Trong phần [cyan]Request Headers[/cyan], tìm dòng [bold]X-Demo-Token[/bold] và copy giá trị của nó.")
        console.print("\n[dim]Lưu ý: Nếu bạn không có mạng hoặc không muốn cập nhật bây giờ, hãy gõ 'skip' để bỏ qua (Hệ thống sẽ chỉ dùng VNHub).[/dim]")
        
        in_colab = 'google.colab' in sys.modules or 'IPython' in sys.modules
        
        if in_colab:
            console.print("\n[bold cyan]⚠️ PHÁT HIỆN ĐANG CHẠY TRÊN COLAB/JUPYTER![/bold cyan]")
            console.print("[yellow]Do Colab không hỗ trợ nhập liệu trực tiếp lúc đang chạy luồng ngầm, vui lòng làm theo cách sau:[/yellow]")
            console.print("👉 Tạo một file tên là [bold green]geovina_token.txt[/bold green] ở thư mục gốc (nơi chứa source code).")
            console.print("👉 Mở file đó ra, dán Token mới vào, sau đó lưu lại (Ctrl+S / Cmd+S).")
            console.print("👉 Hệ thống đang theo dõi file này và sẽ tự động đọc Token ngay khi bạn lưu...")
            console.print("[dim](Gõ chữ 'skip' vào file và lưu lại nếu muốn bỏ qua Geovina)[/dim]\n")
            
            # Tạo sẵn file rỗng cho người dùng dễ thấy
        if in_colab:
            env_token = os.environ.get('GEOVINA_TOKEN', '')
            if env_token:
                is_ok = _test_token(env_token)
                if is_ok:
                    os.environ['GEOVINA_DEMO_TOKEN'] = env_token
                    console.print("[bold green]✅ Đã load GEOVINA_TOKEN từ biến môi trường![/bold green]\n")
                    return True
            
            # Kiểm tra file geovina_token.txt 1 lần duy nhất thay vì loop
            if os.path.exists("geovina_token.txt"):
                try:
                    with open("geovina_token.txt", "r", encoding="utf-8") as f:
                        new_token = f.read().strip().strip('\'"')
                    if new_token and new_token.lower() != 'skip':
                        if _test_token(new_token):
                            os.environ['GEOVINA_DEMO_TOKEN'] = new_token
                            return True
                except: pass
                
            console.print("[yellow]Bỏ qua Geovina trên Colab (Có thể set biến môi trường os.environ['GEOVINA_TOKEN']). Chuyển sang dùng VNHub.[/yellow]\n")
            return False
        else:
            LOG("WARNING: Token expired. Cannot prompt in background thread. Failing gracefully.")
            console.print("[bold red]⚠️ Token Geovina hết hạn! Để nhập Token mới, vui lòng chạy lệnh ở terminal, hoặc cập nhật biến môi trường GEOVINA_TOKEN.[/bold red]")
            return False
            
            if new_token.lower() == 'skip':
                console.print("[yellow]Đã bỏ qua kiểm tra Geovina. Hệ thống sẽ tiếp tục chạy với VNHub.[/yellow]\n")
                return False
            elif new_token:
                with console.status("[bold green]Đang kiểm tra lại token mới...", spinner="dots"):
                    is_ok = _test_token(new_token)
                if is_ok:
                    os.environ['GEOVINA_DEMO_TOKEN'] = new_token
                    console.print("[bold green]✅ Token mới hợp lệ! Hệ thống sẽ tiếp tục quá trình xử lý.[/bold green]\n")
                    return True
                else:
                    console.print("[bold red]❌ Token mới vẫn không hợp lệ. Vui lòng thử lại![/bold red]")
    return True


def main():
    global DEBUG_MODE
    console.print(Panel.fit("[bold green]🚀 PHẦN MỀM TRÍCH XUẤT DỮ LIỆU TỪ ẢNH CCCD RA EXCEL[/bold green]", border_style="cyan", padding=(1, 5)))
    
    with console.status("[bold green]Đang khởi tạo model AI...", spinner="dots"):
        init_models()
        
    first_run = True

    while True:
        input_dir = ""
        do_normalize = True
        # Bỏ tham số --debug khỏi sys.argv để không bị nhầm thành thư mục đầu vào
        args = [a for a in sys.argv if a != '--debug']
        if first_run and len(args) >= 2:
            input_dir = args[1]
            first_run = False
            
            if IN_COLAB:
                do_normalize = True
            else:
                DEBUG_MODE = Confirm.ask("[bold yellow]Bạn có muốn bật chế độ Gỡ lỗi (ghi toàn bộ Raw OCR Text vào file log) không?[/bold yellow]", default=DEBUG_MODE)
                do_normalize = Confirm.ask("\n[bold yellow]Bạn có muốn KIỂM TRA & CHUẨN HÓA ĐỊA CHỈ (quá trình này cần kết nối mạng) không?[/bold yellow]", default=True)
            
            if input_dir.endswith('.xlsx') and os.path.isfile(input_dir):
                if do_normalize: check_and_prompt_geovina_token()
                run_reprocess(input_dir, normalize_address=do_normalize)
                if IN_COLAB:
                    break
                if not Confirm.ask("\n[bold yellow]Bạn có muốn tiếp tục xử lý thư mục khác không?[/bold yellow]"):
                    console.print("\n[bold green]Cảm ơn bạn đã sử dụng phần mềm. Tạm biệt![/bold green]")
                    break
                continue
            else:
                if do_normalize: check_and_prompt_geovina_token()
                run_wizard(input_dir, normalize_address=do_normalize)
                if IN_COLAB:
                    break
        else:
            console.print("\n[bold cyan]--- VUI LÒNG CHỌN LUỒNG XỬ LÝ ---[/bold cyan]")
            console.print("[1] Quét mới / Quét nối tiếp (Xử lý một thư mục ảnh)")
            console.print("[2] Tái xử lý (Chỉ quét lại các ảnh bị lỗi/thiếu thông tin từ file Excel cũ)")
            console.print("[3] Thoát")
            
            choice = Prompt.ask("\n[bold yellow]Nhập lựa chọn của bạn (1/2/3)[/bold yellow]", default="1").strip()
            while choice not in ["1", "2", "3"]:
                console.print("[red]Lựa chọn không hợp lệ, vui lòng nhập 1, 2 hoặc 3.[/red]")
                choice = Prompt.ask("\n[bold yellow]Nhập lựa chọn của bạn (1/2/3)[/bold yellow]", default="1").strip()
            
            if choice == "3":
                console.print("\n[bold green]Cảm ơn bạn đã sử dụng phần mềm. Tạm biệt![/bold green]")
                break
                
            is_reprocess = (choice == "2")
            
            if is_reprocess:
                console.print("\n[bold cyan]--- CHỌN CHẾ ĐỘ TÁI XỬ LÝ ---[/bold cyan]")
                console.print("[1] Tái bổ sung thông tin còn thiếu (Quét lại ảnh bằng AI OCR)")
                console.print("[2] Làm đẹp địa chỉ gốc & Chuẩn hóa địa chỉ mới (Không quét OCR)")
                console.print("[3] Chỉ cập nhật lại địa chỉ chuẩn hóa từ địa chỉ gốc (Không quét OCR)")
                
                reprocess_mode = Prompt.ask("\n[bold yellow]Nhập lựa chọn của bạn (1/2/3)[/bold yellow]", default="1").strip()
                while reprocess_mode not in ["1", "2", "3"]:
                    console.print("[red]Lựa chọn không hợp lệ, vui lòng nhập 1, 2 hoặc 3.[/red]")
                    reprocess_mode = Prompt.ask("\n[bold yellow]Nhập lựa chọn của bạn (1/2/3)[/bold yellow]", default="1").strip()
                
                process_all_rows = False
                if reprocess_mode in ["2", "3"]:
                    console.print("\n[bold yellow]Bạn muốn xử lý cho toàn bộ danh sách, hay chỉ những dòng chưa có Địa chỉ chuẩn hóa?[/bold yellow]")
                    console.print("[1] Toàn bộ danh sách\n[2] Chỉ dòng bị trống")
                    process_all_choice = Prompt.ask("Chọn (1/2)", default="1").strip()
                    while process_all_choice not in ["1", "2"]:
                        console.print("[red]Lựa chọn không hợp lệ, vui lòng nhập 1 hoặc 2.[/red]")
                        process_all_choice = Prompt.ask("Chọn (1/2)", default="1").strip()
                    process_all_rows = (process_all_choice == "1")
                
                excel_path = clean_path(Prompt.ask("\n[bold cyan]Nhập đường dẫn file Excel cũ (hoặc gõ 'q' để quay lại menu)[/bold cyan]"))
                if excel_path.lower() in ('q', 'quit', 'exit'):
                    continue
                if not os.path.isfile(excel_path) or not excel_path.endswith('.xlsx'):
                    console.print(f"\n[bold red]❌ Lỗi: File '{excel_path}' không hợp lệ hoặc không tồn tại.[/bold red]")
                    continue
                    
                DEBUG_MODE = Confirm.ask("\n[bold yellow]Bạn có muốn bật chế độ Gỡ lỗi (in chi tiết log quá trình xử lý) không?[/bold yellow]", default=DEBUG_MODE)
                if reprocess_mode == "1":
                    do_normalize = Confirm.ask("\n[bold yellow]Bạn có muốn KIỂM TRA & CHUẨN HÓA ĐỊA CHỈ (quá trình này cần kết nối mạng) không?[/bold yellow]", default=True)
                else:
                    do_normalize = True # Bắt buộc phải chuẩn hóa trong mode 2, 3
                
                if do_normalize: check_and_prompt_geovina_token()
                run_reprocess(excel_path, mode=reprocess_mode, process_all_rows=process_all_rows, normalize_address=do_normalize)
            else:
                console.print("\n[yellow][Hướng dẫn][/yellow]: Kéo thả thư mục chứa ảnh vào cửa sổ này, hoặc copy đường dẫn thư mục và dán vào đây.")
                input_dir = clean_path(Prompt.ask("[bold cyan]Nhập đường dẫn thư mục chứa ảnh CCCD (hoặc gõ 'q' để quay lại menu)[/bold cyan]"))
                
                if input_dir.lower() in ('q', 'quit', 'exit'):
                    continue
                    
                DEBUG_MODE = Confirm.ask("\n[bold yellow]Bạn có muốn bật chế độ Gỡ lỗi (ghi toàn bộ Raw OCR Text vào file log) không?[/bold yellow]", default=DEBUG_MODE)
                        
                do_normalize = Confirm.ask("\n[bold yellow]Bạn có muốn KIỂM TRA & CHUẨN HÓA ĐỊA CHỈ (quá trình này cần kết nối mạng và tốn thêm thời gian) không?[/bold yellow]", default=True)
                
                if do_normalize: check_and_prompt_geovina_token()
                run_wizard(input_dir, normalize_address=do_normalize)
                
            first_run = False
        
        if not Confirm.ask("\n[bold yellow]Bạn có muốn tiếp tục xử lý thư mục khác không?[/bold yellow]"):
            console.print("\n[bold green]Cảm ơn bạn đã sử dụng phần mềm. Tạm biệt![/bold green]")
            break

if __name__ == '__main__':
    main()
